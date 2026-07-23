#!/usr/bin/env python
"""Self-contained synthetic fixture files for the eval battery.

The feature sweep's data-dependent actions (BOQ processing, quantity takeoff,
RFI/VO registers, spec analysis, schedules) were "failing" on prod evals for
one honest reason: the fixture project had no input documents, so the actions
returned their honest "upload X first" refusals and the harness read the short
answers as FAILs. The real client files can't ship in the repo — these
generators build REALISTIC, WHOLLY SYNTHETIC stand-ins on demand instead, so
any environment can seed a working fixture estate with zero client data and
zero out-of-repo dependencies.

Every generator is deterministic (same bytes each run — no timestamps, no
randomness) so re-seeding is idempotent at the content level.

Usage:
  python scripts/synthetic_fixtures.py --out data/fixtures      # write all
  python scripts/synthetic_fixtures.py --list                   # names only

Programmatic: ``build_all(out_dir)`` or the per-file ``build_*`` functions.
seed_fixtures.py calls these as a fallback for any required fixture file
missing from FIXTURES_DIR (see its --synthetic flag).
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

# Filenames match what the harness manifest / seeder reference.
BOQ_XLSX = "synthetic_boq.xlsx"
RFI_REGISTER_XLSX = "synthetic_rfi_register.xlsx"
VO_LOG_XLSX = "synthetic_variation_order_log.xlsx"
SPEC_SECTION_TXT = "synthetic_spec_section_03_concrete.txt"
GROUND_FLOOR_DXF = "ground_floor_plan.dxf"
PROJECT_PROGRAMME_XER = "project_programme.xer"

_REPO_ROOT = Path(__file__).resolve().parents[1]
# The committed resource-loaded schedule the block tests already trust — one
# source of truth for a synthetic P6 programme.
_XER_SOURCE = _REPO_ROOT / "tests" / "fixtures" / "resource_loaded.xer"


# ── BOQ workbook ─────────────────────────────────────────────────────────────

_BOQ_ROWS = [
    # (section, item, description, unit, qty, rate)
    ("1 — Earthworks", "1.01", "Excavation to reduced level", "m3", 12500, 18.50),
    ("1 — Earthworks", "1.02", "Imported fill compacted in layers", "m3", 8200, 42.00),
    ("1 — Earthworks", "1.03", "Disposal of surplus material off site", "m3", 4300, 25.00),
    ("2 — Concrete", "2.01", "Blinding concrete C12/15, 75mm thick", "m2", 3100, 28.00),
    ("2 — Concrete", "2.02", "Reinforced concrete C32/40 in foundations", "m3", 1850, 385.00),
    ("2 — Concrete", "2.03", "Reinforced concrete C32/40 in ground slabs", "m3", 2400, 410.00),
    ("2 — Concrete", "2.04", "High-yield steel reinforcement", "tonne", 520, 3150.00),
    ("3 — Drainage", "3.01", "300mm dia uPVC storm drain laid in trench", "m", 2650, 145.00),
    ("3 — Drainage", "3.02", "Precast concrete manhole, 1200mm dia", "no", 48, 4850.00),
    ("3 — Drainage", "3.03", "Gully and connection to storm network", "no", 96, 980.00),
    ("4 — Roads", "4.01", "Granular sub-base course, 200mm", "m2", 18500, 32.00),
    ("4 — Roads", "4.02", "Asphalt binder course, 60mm", "m2", 17800, 48.00),
    ("4 — Roads", "4.03", "Asphalt wearing course, 40mm", "m2", 17800, 38.00),
]


def build_boq_xlsx(path: Path) -> Path:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"
    ws.append(["Section", "Item", "Description", "Unit", "Qty", "Rate (SAR)", "Amount (SAR)"])
    for section, item, desc, unit, qty, rate in _BOQ_ROWS:
        ws.append([section, item, desc, unit, qty, rate, round(qty * rate, 2)])
    total = round(sum(q * r for *_, q, r in [(s, i, d, u, q, r) for s, i, d, u, q, r in _BOQ_ROWS]), 2)
    ws.append(["", "", "TOTAL", "", "", "", total])
    wb.save(path)
    return path


# ── RFI register ─────────────────────────────────────────────────────────────

_RFI_ROWS = [
    # (rfi, subject, raised, due, status)
    ("RFI-001", "Clarify rebar cover at ground slab edge", "2026-05-04", "2026-05-11", "Closed"),
    ("RFI-002", "Storm drain invert level conflict at MH-12", "2026-05-18", "2026-05-25", "Closed"),
    ("RFI-003", "Confirm asphalt binder grade for service road", "2026-06-02", "2026-06-09", "Open"),
    ("RFI-004", "Manhole cover load class in landscaped areas", "2026-06-15", "2026-06-22", "Open — overdue"),
    ("RFI-005", "Joint spacing for 2.4m ground slab pours", "2026-06-28", "2026-07-05", "Open — overdue"),
    ("RFI-006", "Substation duct bank routing vs tree pits", "2026-07-06", "2026-07-13", "Open"),
]


def build_rfi_register_xlsx(path: Path) -> Path:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "RFI Register"
    ws.append(["RFI No", "Subject", "Date Raised", "Response Due", "Status"])
    for row in _RFI_ROWS:
        ws.append(list(row))
    wb.save(path)
    return path


# ── Variation order log ──────────────────────────────────────────────────────

_VO_ROWS = [
    ("VO-01", "Additional 300m storm drain diversion", 486000, "Approved"),
    ("VO-02", "Upgrade manhole covers to D400", 118000, "Approved"),
    ("VO-03", "Omit landscaping to phase 2 boundary", -220000, "Pending"),
    ("VO-04", "Rock excavation encountered at CH 2+150", 640000, "Pending"),
    ("VO-05", "Acceleration measures — night shift paving", 380000, "Rejected"),
]


def build_vo_log_xlsx(path: Path) -> Path:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "VO Log"
    ws.append(["VO No", "Description", "Value (SAR)", "Status"])
    for row in _VO_ROWS:
        ws.append(list(row))
    open_value = sum(v for _, _, v, s in _VO_ROWS if s == "Pending")
    ws.append(["", "OPEN (Pending) VALUE", open_value, ""])
    wb.save(path)
    return path


# ── Specification section (CSI div 03) ───────────────────────────────────────

_SPEC_TEXT = """SECTION 03 30 00 — CAST-IN-PLACE CONCRETE

PART 1 — GENERAL
1.1 SUMMARY
    A. Section includes cast-in-place concrete for foundations, ground slabs,
       walls, and pavements, including formwork, reinforcement, mix design,
       placement, curing, and testing.
1.2 SUBMITTALS
    A. Mix designs for each concrete class at least 21 days before first pour.
    B. Reinforcement shop drawings and bar bending schedules.
    C. Curing method statements per exposed element type.
1.3 QUALITY ASSURANCE
    A. Batch plant certified to SASO/ASTM C94. One set of three cylinders per
       50 m3 or per pour, whichever is more frequent.

PART 2 — PRODUCTS
2.1 CONCRETE MATERIALS
    A. Cement: Type I/II Portland cement, ASTM C150.
    B. Class C32/40: minimum cementitious content 380 kg/m3, w/c ratio not
       exceeding 0.45, target slump 100 mm ± 25 mm.
    C. Blinding class C12/15 may use 20% fly ash replacement.
2.2 REINFORCEMENT
    A. High-yield deformed bars grade B500B to BS 4449. Cover: 50 mm to
       ground-contact faces, 40 mm elsewhere unless noted.

PART 3 — EXECUTION
3.1 PLACEMENT
    A. Do not place concrete when ambient temperature exceeds 40 degC unless
       hot-weather measures are in effect (chilled water, ice, night pours).
    B. Maximum free-fall height 1.5 m; use tremie or chutes beyond.
3.2 CURING
    A. Water-cure ground slabs minimum 7 days; membrane curing compound
       permitted on vertical faces.
3.3 TOLERANCES
    A. Slab surface flatness FF25 minimum; column verticality within h/500.

END OF SECTION
"""


def build_spec_section_txt(path: Path) -> Path:
    path.write_text(_SPEC_TEXT, encoding="utf-8")
    return path


# ── Ground floor plan DXF ────────────────────────────────────────────────────

def build_ground_floor_dxf(path: Path) -> Path:
    """A small plan with closed slab outlines and drainage runs, in metres.

    Geometry drawing_qto can actually measure: closed LWPOLYLINEs on layer
    SLAB (areas) and open polylines on layer DRAIN (lengths), plus circles on
    layer MANHOLE (counts).
    """
    import ezdxf

    doc = ezdxf.new("R2010", setup=True)
    doc.units = ezdxf.units.M
    msp = doc.modelspace()
    for layer in ("SLAB", "DRAIN", "MANHOLE"):
        if layer not in doc.layers:
            doc.layers.add(layer)

    # Two rectangular slab bays: 30x20 m and 18x12 m.
    msp.add_lwpolyline(
        [(0, 0), (30, 0), (30, 20), (0, 20)], close=True, dxfattribs={"layer": "SLAB"}
    )
    msp.add_lwpolyline(
        [(35, 0), (53, 0), (53, 12), (35, 12)], close=True, dxfattribs={"layer": "SLAB"}
    )
    # Storm drain runs: 60 m + 25 m.
    msp.add_lwpolyline([(0, -5), (60, -5)], dxfattribs={"layer": "DRAIN"})
    msp.add_lwpolyline([(60, -5), (60, 20)], dxfattribs={"layer": "DRAIN"})
    # Manholes at the bends/ends.
    for x, y in ((0, -5), (30, -5), (60, -5), (60, 20)):
        msp.add_circle((x, y), radius=0.6, dxfattribs={"layer": "MANHOLE"})

    doc.saveas(path)
    return path


# ── Primavera programme (copied from the committed block-test fixture) ───────

def build_project_programme_xer(path: Path) -> Path:
    path.write_bytes(_XER_SOURCE.read_bytes())
    return path


# ── registry / CLI ───────────────────────────────────────────────────────────

BUILDERS = {
    BOQ_XLSX: build_boq_xlsx,
    RFI_REGISTER_XLSX: build_rfi_register_xlsx,
    VO_LOG_XLSX: build_vo_log_xlsx,
    SPEC_SECTION_TXT: build_spec_section_txt,
    GROUND_FLOOR_DXF: build_ground_floor_dxf,
    PROJECT_PROGRAMME_XER: build_project_programme_xer,
}


def build_all(out_dir: Path) -> list:
    out_dir.mkdir(parents=True, exist_ok=True)
    written = []
    for name, builder in BUILDERS.items():
        written.append(builder(out_dir / name))
    return written


def main() -> int:
    ap = argparse.ArgumentParser(description="generate synthetic eval fixture files")
    ap.add_argument("--out", default="data/fixtures", help="output directory")
    ap.add_argument("--list", action="store_true", help="print fixture names and exit")
    args = ap.parse_args()
    if args.list:
        for name in BUILDERS:
            print(name)
        return 0
    written = build_all(Path(args.out))
    for p in written:
        print(f"wrote {p} ({p.stat().st_size} bytes)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
