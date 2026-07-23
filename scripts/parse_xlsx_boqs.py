"""Parse five digital BOQ spreadsheets into structured per-item rows with arithmetic QA.

Clean xlsx only (no OCR). Writes one CSV per file plus a markdown QA summary into
the boq_batch scratchpad folder. See boq_batch_qa_summary.md for structure notes.

Run: python scripts/parse_xlsx_boqs.py
"""
import os
import re
import csv
from numbers import Number
import openpyxl

# acacia bill/section header code, e.g. "A6 EMPLOYER'S FACILITIES", "B9 EXCAVATION"
ACACIA_SECTION_RE = re.compile(r"^[A-Z]\d+\s+[A-Z]")

BATCH = (
    r"C:\Users\shimm\AppData\Local\Temp\claude\C--Users-shimm"
    r"\436703f7-0a30-48b6-a650-29d75aac4fa5\scratchpad\boq_batch"
)

CSV_COLS = [
    "Section", "Description", "Quantity", "Unit",
    "Unit Price", "Total Price", "priced", "qa_row_ok",
]


def s(x):
    """ASCII-safe string for the cp1252 console."""
    return ("" if x is None else str(x)).encode("ascii", "replace").decode()


def num(x):
    return x if isinstance(x, Number) and not isinstance(x, bool) else None


def is_num(x):
    return isinstance(x, Number) and not isinstance(x, bool)


def qa_ok(q, r, a):
    """quantity*unit_price == total_price within tolerance."""
    if q is None or r is None or a is None:
        return None
    return abs(q * r - a) <= max(1.0, 0.005 * abs(a))


def load(fname, sheet):
    wb = openpyxl.load_workbook(os.path.join(BATCH, fname), read_only=True, data_only=True)
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()
    return rows


def cell(row, j):
    return row[j] if len(row) > j else None


def write_csv(slug, records, extra_cols=None):
    extra_cols = extra_cols or []
    path = os.path.join(BATCH, f"{slug}_items.csv")
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(CSV_COLS + extra_cols)
        for rec in records:
            w.writerow([rec.get(c, "") for c in CSV_COLS] + [rec.get(c, "") for c in extra_cols])
    return path


# --------------------------------------------------------------------------- #
# 1) ROSHN  (Janadriyah Phase 4)  -- currency n/a, entirely UNPRICED template
# --------------------------------------------------------------------------- #
def parse_roshn():
    rows = load("roshn.xlsx", "BOQ")
    # header row index 7: POMI|CSI|Item|Description|Qty.|Unit|Rate|Amount, then
    # per-variant Quantity/Cost blocks (cols 9-90). All Cost sub-columns empty.
    recs = []
    section = ""
    for r in rows[9:]:
        item = cell(r, 2)
        desc = cell(r, 3)
        qty = num(cell(r, 4))
        unit = cell(r, 5)
        rate = num(cell(r, 6))
        amt = num(cell(r, 7))
        # Literal zeros in Rate/Amount are not real prices -> treat as absent.
        if rate == 0:
            rate = None
        if amt == 0:
            amt = None
        # Section header: text in Item col, no description / no quantity.
        if isinstance(item, str) and item.strip() and not (isinstance(desc, str) and desc.strip()):
            section = item.strip()
            continue
        if not (isinstance(desc, str) and desc.strip()):
            continue
        priced = rate is not None or amt is not None
        recs.append({
            "Section": section,
            "Description": desc.strip(),
            "Quantity": qty if qty is not None else "",
            "Unit": (unit or "") if isinstance(unit, str) else "",
            "Unit Price": rate if rate is not None else "",
            "Total Price": amt if amt is not None else "",
            "priced": priced,
            "qa_row_ok": qa_ok(qty, rate, amt) if priced else "",
        })
    return recs


# --------------------------------------------------------------------------- #
# 2) ACACIA 1  (MBR - DH-Acacia 1)  -- currency AED, PRICED
# --------------------------------------------------------------------------- #
ACACIA_AGG = ("collection", "summary", "total", "carried", "brought",
              "page", "grand", "carry forward")


def parse_acacia():
    rows = load("acacia1.xlsx", "BOQ Bill")
    # desc=3, unit=11, qty=12, rate=13, amount(Cost AED)=14
    recs = []
    section = ""
    measured = 0
    measured_pass = 0
    for r in rows:
        desc = cell(r, 3)
        unit = cell(r, 11)
        qty = num(cell(r, 12))
        rate = num(cell(r, 13))
        amt = num(cell(r, 14))
        row_txt = " ".join(str(c) for c in r if isinstance(c, str)).lower()
        is_agg = any(k in row_txt for k in ACACIA_AGG)
        # Section header: clause/bill title matching the code pattern.
        if isinstance(desc, str) and ACACIA_SECTION_RE.match(desc.strip()) and amt is None:
            section = desc.strip()
            continue
        if is_agg:
            continue
        # Line item = a row carrying a money amount AND a description.
        if amt is None or not (isinstance(desc, str) and desc.strip()):
            continue
        q, rt = qty, rate
        row_qa = ""
        if q is not None and rt is not None and amt is not None:
            ok = qa_ok(q, rt, amt)
            row_qa = ok
            measured += 1
            if ok:
                measured_pass += 1
        recs.append({
            "Section": section,
            "Description": (desc.strip() if isinstance(desc, str) else ""),
            "Quantity": q if q is not None else "",
            "Unit": (unit or "") if isinstance(unit, str) else "",
            "Unit Price": rt if rt is not None else "",
            "Total Price": amt,
            "priced": True,
            "qa_row_ok": row_qa,
        })
    return recs, measured, measured_pass


# --------------------------------------------------------------------------- #
# 3) JEBEL SIFAH Farms-3 (hlx_diff)  -- OMR, UNPRICED qty-comparison BOQ
# --------------------------------------------------------------------------- #
def parse_jebel(fname, slug, mod_label):
    rows = load(fname, openpyxl.load_workbook(
        os.path.join(BATCH, fname), read_only=True).sheetnames[0])
    # No.=0, DESCRIPTION=1, UNIT=2, Original qty=3, Modified/Contracted qty=4,
    # (hlx_diff only) derived col5 = modified*0.7, "AMOUNT (OMR)" col=text/none
    recs = []
    section = ""
    for r in rows[9:]:
        no = cell(r, 0)
        desc = cell(r, 1)
        unit = cell(r, 2)
        oq = num(cell(r, 3))
        mq = num(cell(r, 4))
        # Section header: description present, no unit, no quantities.
        if isinstance(desc, str) and desc.strip() and not (isinstance(unit, str) and unit.strip()) \
                and oq is None and mq is None:
            section = desc.strip()
            continue
        if not (isinstance(desc, str) and desc.strip()):
            continue
        if oq is None and mq is None:
            continue
        recs.append({
            "Section": section,
            "Description": desc.strip(),
            "Quantity": mq if mq is not None else "",  # current/modified qty
            "Unit": (unit or "") if isinstance(unit, str) else "",
            "Unit Price": "",
            "Total Price": "",
            "priced": False,
            "qa_row_ok": "",
            "Item No": (no if no is not None else ""),
            "Original Qty": oq if oq is not None else "",
            mod_label: mq if mq is not None else "",
        })
    return recs


# --------------------------------------------------------------------------- #
# 5) TGH - The Grand Hanoi  -- VND, QUANTITY-ONLY furniture count schedule
# --------------------------------------------------------------------------- #
def parse_tgh():
    rows = load("tgh.xlsx", "Schedule")
    # header row 3: DESCRIPTION=1, SIZE/TYPE=2, REF.CODE=3, REF.IMAGE=4, TOTAL=5,
    # then per-apartment-type counts cols 6-15. No prices in this sheet.
    recs = []
    section = ""
    desc_ff = ""
    for r in rows[5:]:
        desc = cell(r, 1)
        size = cell(r, 2)
        ref = cell(r, 3)
        total = num(cell(r, 5))
        # Section header (e.g. LIVING - DINING): desc present, no total, no ref.
        if isinstance(desc, str) and desc.strip() and total is None and \
                not (isinstance(ref, str) and ref.strip()) and \
                not (isinstance(size, str) and size.strip()):
            section = desc.strip()
            continue
        if isinstance(desc, str) and desc.strip():
            desc_ff = desc.strip()
        if total is None:
            continue
        full_desc = desc_ff
        if isinstance(size, str) and size.strip():
            full_desc = (desc_ff + " - " + size.strip().replace("\n", " ")).strip(" -")
        recs.append({
            "Section": section,
            "Description": full_desc,
            "Quantity": total,
            "Unit": "pcs",  # implied (furniture count); Schedule has no unit col
            "Unit Price": "",
            "Total Price": "",
            "priced": False,
            "qa_row_ok": "",
            "Ref Code": (ref if isinstance(ref, str) else ""),
        })
    return recs


def pct(n, d):
    return "n/a" if not d else f"{100.0*n/d:.1f}%"


def main():
    out = []

    roshn = parse_roshn()
    p1 = write_csv("roshn", roshn)
    r_priced = sum(1 for x in roshn if x["priced"])
    out.append(("roshn", p1, len(roshn), r_priced, len(roshn) - r_priced, 0, 0))

    acacia, meas, meas_pass = parse_acacia()
    p2 = write_csv("acacia1", acacia)
    out.append(("acacia1", p2, len(acacia), len(acacia), 0, meas, meas_pass))

    jebel = parse_jebel("hlx_diff.xlsx", "jebel_sifah", "Modified Qty")
    p3 = write_csv("jebel_sifah", jebel, extra_cols=["Item No", "Original Qty", "Modified Qty"])
    out.append(("jebel_sifah", p3, len(jebel), 0, len(jebel), 0, 0))

    villas = parse_jebel("hlx_80villas.xlsx", "hlx_80villas", "Contracted Qty")
    p4 = write_csv("hlx_80villas", villas, extra_cols=["Item No", "Original Qty", "Contracted Qty"])
    out.append(("hlx_80villas", p4, len(villas), 0, len(villas), 0, 0))

    tgh = parse_tgh()
    p5 = write_csv("tgh", tgh, extra_cols=["Ref Code"])
    tgh_qty = sum(x["Quantity"] for x in tgh if is_num(x["Quantity"]))
    out.append(("tgh", p5, len(tgh), 0, len(tgh), 0, 0))

    for slug, path, n, pr, un, meas, mp in out:
        line = (f"{slug:14s} items={n:5d} priced={pr:5d} unpriced={un:5d} "
                f"measured_QA={meas} pass={mp} ({pct(mp, meas)}) -> {os.path.basename(path)}")
        print(s(line))
    print(s(f"tgh total furniture pieces (sum of TOTAL) = {int(tgh_qty)}"))


if __name__ == "__main__":
    main()
