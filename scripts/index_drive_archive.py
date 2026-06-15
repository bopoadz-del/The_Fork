"""Drive-archive indexer for The_Fork.

Walks a Google-Drive mirror (G:\\My Drive\\ by default), filters to construction
documents by extension + path-keyword, extracts text, chunks via
``app.core.doc_index.chunk_text_with_overlap``, and upserts into the project's
vector store via ``app.core.rag.retriever.index_chunks``.

Run modes
---------
* ``--inventory-out PATH``  Walk + filter, write JSONL of {path, ext, size}, exit.
* ``--inventory-in PATH``   Skip walking; read file list from JSONL.
* ``--dry-run``             Extract + chunk but do NOT embed/upsert.
* ``--limit N``             Cap total files processed.
* ``--resume-state PATH``   Track completed paths so reruns skip them.

Designed for Windows hosts. No emojis.
"""

from __future__ import annotations

import argparse
import asyncio
import hashlib
import io
import json
import multiprocessing
import os
import shlex
import sys
import time
import traceback
from datetime import datetime, timezone
from typing import Any, Dict, Iterable, List, Optional, Tuple

# Make ``app`` importable regardless of CWD.
_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)


# ── platform-block singletons ────────────────────────────────────────────────
# Instantiate ONCE at module scope (not per-doc). Per-doc instantiation re-runs
# wire() every call and reaches into the dependency dict each time. The async
# `process()` calls run on a single asyncio loop owned by main().
#
# Wiring matters: pdf_v2.process() calls self.get_dep("ocr") for image-only
# PDFs. get_dep reads self._dependencies, which is populated only by .wire().
# Without the explicit wire below, every scanned drawing returns empty text.
#
# Caps: pdf_v2 defaults are text_limit=20000 chars and max_pages=100. The old
# script had no caps; for an archive of construction docs (80+ page specs,
# multi-hundred-page submittals) we must lift them or chunk counts collapse.
_PDF_BLOCK = None
_OCR_BLOCK = None
_DRAWING_BLOCK = None
_LOOP: Optional[asyncio.AbstractEventLoop] = None


def _init_platform_blocks() -> None:
    """Lazy-init the pdf_v2 + ocr + drawing_qto block singletons wired together.

    pdf_v2 is the newer block (has the OCR-fallback path for image-only PDFs).
    Caps are lifted so we don't truncate long specs or stop at page 100.
    drawing_qto handles CAD PDFs (returns title-block + notes + cross-refs as
    `result["text"]` per Phase 1+1.5+1.6).
    """
    global _PDF_BLOCK, _OCR_BLOCK, _DRAWING_BLOCK, _LOOP
    if _PDF_BLOCK is not None:
        return
    from app.blocks.pdf_v2 import PDFBlockV2
    from app.blocks.ocr import OCRBlock
    from app.blocks.drawing_qto import DrawingQTOBlock

    _OCR_BLOCK = OCRBlock()
    _PDF_BLOCK = PDFBlockV2(config={
        # No truncation: a long spec must produce all its pages.
        "text_limit": 10 ** 9,
        "max_pages": 10 ** 9,
        # When the text layer is below 200 chars, treat as image-only and OCR.
        "ocr_fallback_min_chars": 200,
    })
    # Explicit dependency wiring — required so pdf_v2.get_dep("ocr") returns
    # a real block instead of None (no DI container in a script).
    _PDF_BLOCK.wire("ocr", _OCR_BLOCK)
    _DRAWING_BLOCK = DrawingQTOBlock()
    _LOOP = asyncio.new_event_loop()


def _run(coro):
    """Run an async coroutine on the singleton loop."""
    assert _LOOP is not None, "loop not initialised; call _init_platform_blocks() first"
    return _LOOP.run_until_complete(coro)


# ── filter config ────────────────────────────────────────────────────────────

INCLUDE_EXTS = {
    ".pdf", ".docx", ".doc", ".xlsx", ".xls", ".pptx", ".ppt",
    ".txt", ".md", ".csv",
}

# Lower-cased keywords; match if any appears anywhere in the full path.
INCLUDE_KEYWORDS = [
    "construction", "project", "boq", "bod", "rfp", "rfq", "tender", "bid",
    "drawing", "spec", "specification", "prc-", "cesmm", "evm", "wbs",
    "schedule", "manpower", "cost", "estimate", "procurement", "cad", "bim",
    "revit", "civil", "structural", "mep", "architect", "engineering",
    "contract", "subcontract", "submittal", "rfi", "nce", "variation",
    "change order", "as-built", "commissioning", "data center", "data centre",
    "mechanical", "electrical", "plumbing", "hvac", "building", "floorplan",
    "road", "concrete", "asphalt", "pavement", "earthworks", "bridge",
    "tunnel", "scaned", "scanned",
]

# Path-component substrings that, if present, exclude the file outright.
EXCLUDE_PATH_FRAGMENTS = [
    "/photos/", "/pictures/", "/music/", "/videos/", "/family/",
    "/personal/", "/income tax/", "/recipes/",
]

MAX_BYTES = 200 * 1024 * 1024  # 200 MB hard ceiling
MIN_BYTES = 1024               # 1 KB floor — corrupt / empty


def _norm_path_for_match(p: str) -> str:
    """Lower-case + forward-slash so EXCLUDE_PATH_FRAGMENTS match cleanly."""
    return p.replace("\\", "/").lower()


def matches_include(path: str) -> bool:
    """Return True iff the path passes ext + keyword filters and is not excluded."""
    ext = os.path.splitext(path)[1].lower()
    if ext not in INCLUDE_EXTS:
        return False
    norm = _norm_path_for_match(path)
    for frag in EXCLUDE_PATH_FRAGMENTS:
        if frag in norm:
            return False
    return any(kw in norm for kw in INCLUDE_KEYWORDS)


# ── inventory ────────────────────────────────────────────────────────────────

def walk_inventory(root: str, verbose: bool = False) -> List[Dict]:
    """Walk ``root`` recursively, yield filtered entries with size + ext."""
    out: List[Dict] = []
    scanned = 0
    for dirpath, dirnames, filenames in os.walk(root):
        # Cheap pre-prune: skip excluded directory branches outright.
        norm_dir = _norm_path_for_match(dirpath) + "/"
        if any(frag in norm_dir for frag in EXCLUDE_PATH_FRAGMENTS):
            dirnames[:] = []
            continue
        for fn in filenames:
            scanned += 1
            full = os.path.join(dirpath, fn)
            if not matches_include(full):
                continue
            try:
                size = os.path.getsize(full)
            except OSError:
                continue
            if size > MAX_BYTES or size < MIN_BYTES:
                continue
            out.append({
                "path": full,
                "ext": os.path.splitext(fn)[1].lower(),
                "size": size,
            })
            if verbose and len(out) % 500 == 0:
                print(f"  scanned={scanned} kept={len(out)}", flush=True)
    if verbose:
        print(f"walk complete: scanned={scanned} kept={len(out)}", flush=True)
    return out


def write_inventory_jsonl(entries: Iterable[Dict], path: str) -> int:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    n = 0
    with open(path, "w", encoding="utf-8") as fh:
        for e in entries:
            fh.write(json.dumps(e, ensure_ascii=False) + "\n")
            n += 1
    return n


def read_inventory_jsonl(path: str) -> List[Dict]:
    out: List[Dict] = []
    with open(path, "r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            out.append(json.loads(line))
    return out


# ── text extraction ──────────────────────────────────────────────────────────

def _read_text_file(path: str) -> str:
    with open(path, "rb") as fh:
        raw = fh.read()
    return raw.decode("utf-8", errors="replace")


def _ocr_lang() -> str:
    """Tesseract language code. The OCR platform block reads RAG_OCR_LANG
    too; setting it here keeps the env contract documented at the script
    surface."""
    return os.getenv("RAG_OCR_LANG", "eng")


def _is_drawing_path(path: str) -> bool:
    """True iff the path looks like a CAD/drawing PDF.

    Drawings route through DrawingQTOBlock (Phase 1+1.5+1.6) which returns
    a text-shaped chunk (`result["text"]`) built from title-block + notes
    + cross-refs, plus a structured `result["drawing"]` namespace for audit.
    """
    norm = path.replace("\\", "/").lower()
    name = os.path.basename(norm)
    return (
        "02-drawings" in norm
        or "/drawings/" in norm
        or "-dwg-" in name
        or name.startswith("dwg-")
    )


def _extract_drawing_via_qto(path: str) -> Tuple[str, Optional[Dict], List[str], str]:
    """Run DrawingQTOBlock's text path on a CAD PDF.

    Returns (text, drawing_dict_or_None, errors_list, block_status). On
    `status: error` we DO NOT fall back to pdf+OCR — scanned drawings
    without a text layer are deferred to a future v2 / OCR-bbox pass per
    the spec.

    Implementation note: we call `_extract_drawing_text` directly (the
    pdfplumber-based text path) instead of `.process()`. `.process()`
    additionally runs `_extract_from_pdf` (PyMuPDF `get_drawings()`
    vector-geometry pass) — that produces measurements/areas/volumes the
    INDEXER discards, while on vector-dense CAD PDFs it's a memory/CPU
    bomb (observed ~1 GB resident on a 10 MB SW sheet). The four keys
    the indexer needs (`status`, `text`, `drawing`, `errors`) all come
    from the text path, so we skip the geometry pass entirely.
    """
    _init_platform_blocks()
    try:
        # Synchronous — _extract_drawing_text is not a coroutine, so we
        # call it directly (no _run / asyncio loop).
        result = _DRAWING_BLOCK._extract_drawing_text(path)  # noqa: SLF001
    except Exception as exc:  # noqa: BLE001
        return "", None, [f"drawing_qto_exception: {repr(exc)[:200]}"], "drawing_qto_exception"

    if not isinstance(result, dict):
        return "", None, ["drawing_qto_bad_return"], "drawing_qto_bad_return"

    text = result.get("text") or ""
    drawing = result.get("drawing") if isinstance(result.get("drawing"), dict) else None
    errors = list(result.get("errors") or [])
    status = result.get("status") or "unknown"
    if status == "error" and "error" in result and not errors:
        errors.append(f"drawing_qto_error: {str(result.get('error'))[:200]}")
    block_status = f"drawing_qto:{status}"
    return text, drawing, errors, block_status


def _extract_pdf_via_platform(path: str) -> Tuple[str, str, int, Optional[str]]:
    """Run pdf_v2 → returns (text, block_source, pages, error).

    pdf_v2's `source` field is one of "pdf" (text-layer extraction)
    or "pdf+ocr" (text layer was below threshold, OCR fallback ran).
    We expose this verbatim as the audit row's `block_status`.
    """
    _init_platform_blocks()
    try:
        result = _run(_PDF_BLOCK.process({"file_path": path}))
    except Exception as exc:  # noqa: BLE001
        return "", "pdf_v2_exception", 0, repr(exc)[:200]

    if not isinstance(result, dict):
        return "", "pdf_v2_bad_return", 0, f"unexpected return type: {type(result).__name__}"

    text = result.get("text") or ""
    source = result.get("source") or "pdf"
    pages = int(result.get("pages") or 0)
    err = (result.get("metadata") or {}).get("error")
    return text, source, pages, err


def _extract_docx(path: str) -> str:
    import docx  # python-docx
    document = docx.Document(path)
    return "\n".join(p.text for p in document.paragraphs if p.text)


def _extract_xlsx(path: str) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet_blobs: List[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: List[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(v) for v in row if v is not None and str(v).strip()]
            if cells:
                rows.append("\t".join(cells))
        if rows:
            sheet_blobs.append(f"[sheet: {sheet_name}]\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(sheet_blobs)


class ExtractResult:
    """Container for extraction outcomes (avoids tuple sprawl in callers).

    Fields:
        text             extracted text (possibly empty)
        ocr_required     True iff OCR ran (pdf+ocr source, or set by caller)
        skipped_reason   set when the format is not handled
        pages_ocrd       count of pages OCR'd; with whole-doc OCR via
                         pdf_v2 this equals total pages when source=='pdf+ocr'
        ocr_error        last error string from extraction, or None
        extractor_used   which platform path produced the text
                         (pdf_v2 / docx / xlsx / text)
        block_status     the underlying block's `source` / status value
                         (e.g. pdf / pdf+ocr); None when not from a block
        is_drawing       True iff path matches a CAD-drawing heuristic
    """
    __slots__ = (
        "text", "ocr_required", "skipped_reason", "pages_ocrd", "ocr_error",
        "extractor_used", "block_status", "is_drawing",
        "drawing", "drawing_errors",
    )

    def __init__(
        self,
        text: str = "",
        ocr_required: bool = False,
        skipped_reason: Optional[str] = None,
        pages_ocrd: int = 0,
        ocr_error: Optional[str] = None,
        extractor_used: Optional[str] = None,
        block_status: Optional[str] = None,
        is_drawing: bool = False,
        drawing: Optional[Dict] = None,
        drawing_errors: Optional[List[str]] = None,
    ) -> None:
        self.text = text
        self.ocr_required = ocr_required
        self.skipped_reason = skipped_reason
        self.pages_ocrd = pages_ocrd
        self.ocr_error = ocr_error
        self.extractor_used = extractor_used
        self.block_status = block_status
        self.is_drawing = is_drawing
        self.drawing = drawing
        self.drawing_errors = drawing_errors or []


def extract(path: str, ext: str) -> ExtractResult:
    """Extract text from ``path``. Returns ExtractResult.

    Routing:
      .pdf  → pdf_v2 platform block (OCR-fallback for image-only PDFs).
              Drawings (02-Drawings / -DWG- / filename heuristic) follow
              the same path: pdf_v2 captures the title-block + callout
              text that vector retrieval needs. Structured QTO (areas,
              measurements) is a SEPARATE ingestion pass through
              DrawingQTOBlock — not done here, because that block returns
              measurements (m², m³, layer counts), not text.
      .docx → python-docx direct. document_engine.DocumentEngineBlock is
              a Parse → Reason → Map pipeline that internally calls its
              own DOCXParser; we don't need its reasoning layers for
              vector indexing, and routing through it would drag in the
              full blocks/document_engine/* tree as a script dependency.
      .xlsx → openpyxl direct (no platform block exists for spreadsheets).

    skipped_reason set when the format is not handled (no installed lib).
    Raises only on truly unexpected errors; the caller logs them as errors.
    """
    if ext == ".pdf":
        is_drawing = _is_drawing_path(path)
        if is_drawing:
            # Drawings: DrawingQTOBlock returns a text-shaped chunk built from
            # title-block + notes + cross-refs (Phase 1+1.5+1.6). No OCR
            # fallback — scanned drawings without text layer are deferred.
            text, drawing, qto_errors, block_status = _extract_drawing_via_qto(path)
            return ExtractResult(
                text=text,
                ocr_required=False,
                pages_ocrd=0,
                ocr_error=None,
                extractor_used="drawing_qto",
                block_status=block_status,
                is_drawing=True,
                drawing=drawing,
                drawing_errors=qto_errors,
            )
        text, source, pages, err = _extract_pdf_via_platform(path)
        ocr_required = (source == "pdf+ocr")
        # pdf_v2 runs OCR over the WHOLE doc as a unit when it falls back —
        # there is no per-page page-by-page counter coming back. The closest
        # truthful signal for `pages_ocrd` is "all pages if OCR ran, else 0".
        pages_ocrd = pages if ocr_required else 0
        return ExtractResult(
            text=text, ocr_required=ocr_required,
            pages_ocrd=pages_ocrd, ocr_error=err,
            extractor_used="pdf_v2",
            block_status=source,
            is_drawing=is_drawing,
        )
    if ext == ".docx":
        return ExtractResult(
            text=_extract_docx(path),
            extractor_used="docx",
            block_status="docx_python_docx",
        )
    if ext == ".xlsx":
        return ExtractResult(
            text=_extract_xlsx(path),
            extractor_used="xlsx",
            block_status="xlsx_openpyxl",
        )
    if ext in (".txt", ".md", ".csv"):
        return ExtractResult(
            text=_read_text_file(path),
            extractor_used="text",
            block_status="text_raw",
        )
    if ext in (".doc",):
        return ExtractResult(skipped_reason="legacy_doc_no_extractor")
    if ext in (".ppt", ".pptx"):
        return ExtractResult(skipped_reason="pptx_no_extractor")
    if ext in (".xls",):
        return ExtractResult(skipped_reason="xls_no_extractor")
    return ExtractResult(skipped_reason=f"unsupported_ext:{ext}")


# ── doc_id + chunk header ────────────────────────────────────────────────────

def doc_id_for(path: str) -> str:
    """Stable 8-hex doc_id from absolute path."""
    h = hashlib.sha256(path.encode("utf-8")).hexdigest()
    return h[:8]


def prefix_chunks_with_path(chunks: List[str], path: str) -> List[str]:
    """Prepend a one-line ``[source: <path>]`` header to each chunk so that
    the vector store (which doesn't carry per-chunk metadata) still surfaces
    the source location to humans inspecting hits."""
    header = f"[source: {path}]\n"
    return [header + c for c in chunks]


# ── resume state + audit ─────────────────────────────────────────────────────

def load_resume(path: str) -> Dict:
    if not path or not os.path.exists(path):
        return {"done": []}
    try:
        with open(path, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        if "done" not in data or not isinstance(data["done"], list):
            data["done"] = []
        return data
    except Exception:
        return {"done": []}


def save_resume(path: str, data: Dict) -> None:
    if not path:
        return
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False)
        fh.flush()
        os.fsync(fh.fileno())
    os.replace(tmp, path)


def append_audit(path: str, row: Dict) -> None:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "a", encoding="utf-8") as fh:
        fh.write(json.dumps(row, ensure_ascii=False) + "\n")


def _drawing_audit_extras(
    is_drawing: bool,
    drawing_meta: Optional[Dict],
    drawing_errors: List[str],
) -> Dict:
    """Build the drawing-specific audit fields. For non-drawing rows this
    just records `is_drawing=False` so column shape stays uniform across
    the audit log. For drawings, lift drawing_number / title / discipline
    / revision / cross-ref + note counts from `result["drawing"]` and
    surface the DrawingQTOBlock errors list."""
    extras: Dict[str, Any] = {
        "is_drawing": is_drawing,
        "drawing_qto_deferred": False if is_drawing else False,
    }
    if not is_drawing:
        return extras
    dm = drawing_meta or {}
    extras.update({
        "drawing_number": dm.get("drawing_number"),
        "drawing_title": dm.get("drawing_title"),
        "discipline": dm.get("discipline"),
        "revision": dm.get("revision"),
        "n_cross_refs": len(dm.get("cross_refs") or []),
        "n_notes": len(dm.get("notes") or []),
        "drawing_errors": list(drawing_errors or []),
    })
    return extras


# ── watchdog (per-doc hard timeout) ──────────────────────────────────────────
#
# Why a worker process and not a thread: a stuck PDF parser inside fitz /
# pdf_v2 is uninterruptible from Python. `Thread.join(timeout=...)` would
# unblock the main loop, but the worker keeps running and holds the GIL
# during native calls, leaving a real zombie + memory leak across files.
# A worker process can be force-killed with SIGTERM/TerminateProcess via
# ``pool.terminate()``, which truly releases the file handle and memory.
#
# Why only ``extract()`` runs in the worker: the hang is in PDF parsing.
# Chunking + embedding + audit + state stay in main, so the vector store
# has a single owner (no cross-process write / partial-write-on-kill),
# and worker restarts after a timeout do not reload the embedder. This
# is a slight deviation from a literal "5 minutes per *doc*" reading —
# the embedder call is uncapped — but matches where the hang risk lives.
#
# Worker payload is a plain dict, NOT the ``ExtractResult`` slotted class,
# to avoid pickle surprises across the process boundary.


def _extract_worker(path: str, ext: str) -> Dict:
    """Run ``extract()`` in a worker process; return a JSON-safe dict.

    Module-scope so ``multiprocessing`` can pickle it on Windows
    (spawn-method default). Imports and block init happen lazily inside
    the worker on first call; the parent submits a one-shot warmup
    before the timed loop so cold-start cost is not charged to a real
    file's per-doc budget.
    """
    res = extract(path, ext)
    return {
        "text": res.text,
        "ocr_required": res.ocr_required,
        "skipped_reason": res.skipped_reason,
        "pages_ocrd": res.pages_ocrd,
        "ocr_error": res.ocr_error,
        "extractor_used": res.extractor_used,
        "block_status": res.block_status,
        "is_drawing": res.is_drawing,
        "drawing": res.drawing,
        "drawing_errors": res.drawing_errors,
    }


def _warmup_worker() -> str:
    """Force the worker to load fitz + pdf_v2 + drawing_qto so the first
    real extract() inside the timed loop isn't charged for cold-start
    init. Returns a literal string so the parent can assert success."""
    _init_platform_blocks()
    return "ok"


class _Watchdog:
    """One-worker process pool with per-call hard timeout.

    On TimeoutError, ``call()`` calls ``pool.terminate()`` (TerminateProcess
    on Windows) and rebuilds the pool so the next call gets a fresh
    worker. Workers are otherwise long-lived: the embedder + block
    singletons load once per worker and are amortized across all
    healthy files.
    """

    def __init__(self, timeout_seconds: int) -> None:
        self.timeout = timeout_seconds
        self._pool: Optional[multiprocessing.pool.Pool] = None
        self._open_pool_and_warm()

    def _open_pool_and_warm(self) -> None:
        # spawn context so the worker starts clean — no inherited fitz
        # state, no inherited asyncio loop from the parent. Matches
        # Windows default behaviour explicitly.
        ctx = multiprocessing.get_context("spawn")
        self._pool = ctx.Pool(processes=1)
        # Warmup is untimed — first-time fitz/PyMuPDF init can take a
        # few seconds on Windows and we do not want that charged to a
        # `--per-doc-timeout-seconds 5` smoke test.
        try:
            self._pool.apply(_warmup_worker)
        except Exception:
            # If even warmup fails, keep going — the per-file call will
            # surface the real error. We don't want to abort the batch.
            pass

    def call(self, path: str, ext: str) -> Tuple[Optional[Dict], Optional[float]]:
        """Run extract() in the worker with a hard timeout.

        Returns (result_dict, None) on success; (None, elapsed_seconds)
        on timeout. Any non-timeout exception from the worker is
        re-raised to the main loop's existing ``except Exception``
        handler so it lands in the audit log as an error row.
        """
        assert self._pool is not None
        t0 = time.monotonic()
        async_res = self._pool.apply_async(_extract_worker, (path, ext))
        try:
            data = async_res.get(timeout=self.timeout)
            return data, None
        except multiprocessing.TimeoutError:
            elapsed = time.monotonic() - t0
            # Force-kill the stuck worker and rebuild so the next file
            # starts fresh. apply_async's worker IS the only worker in
            # this pool of size 1, so terminate() targets it.
            self._pool.terminate()
            self._pool.join()
            self._pool = None
            self._open_pool_and_warm()
            return None, elapsed

    def close(self) -> None:
        if self._pool is not None:
            self._pool.close()
            self._pool.join()
            self._pool = None


def _append_timeout_path(timeouts_path: str, path: str) -> None:
    """One line per timeout, for the operator to inspect after the run."""
    os.makedirs(os.path.dirname(timeouts_path) or ".", exist_ok=True)
    with open(timeouts_path, "a", encoding="utf-8") as fh:
        fh.write(path + "\n")


# ── main ─────────────────────────────────────────────────────────────────────

def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description="Drive-archive indexer")
    ap.add_argument("--root", default=r"G:\My Drive")
    ap.add_argument("--project-id", default="drive_archive")
    ap.add_argument("--limit", type=int, default=None)
    ap.add_argument("--inventory-out", default=None)
    ap.add_argument("--inventory-in", default=None)
    ap.add_argument("--resume-state",
                    default=os.path.join("data", "logs", "drive_indexer_state.json"))
    ap.add_argument("--audit", default=os.path.join("data", "logs", "drive_indexer_audit.jsonl"))
    ap.add_argument("--skipped", default=os.path.join("data", "logs", "drive_indexer_skipped.jsonl"))
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument(
        "--budget-seconds", type=int, default=7200,
        help="Wall-clock budget in seconds; on overrun, save state and exit 0.",
    )
    ap.add_argument(
        "--per-doc-timeout-seconds", type=int, default=300,
        help=(
            "Hard timeout per file. On timeout the file is logged with "
            "status=timeout, appended to data/logs/drive_indexer_timeouts.txt, "
            "and marked DONE in resume-state so daily reruns skip it. "
            "Default 300 (5 minutes)."
        ),
    )
    ap.add_argument(
        "--timeouts-log",
        default=os.path.join("data", "logs", "drive_indexer_timeouts.txt"),
        help="Append-only one-line-per-path log of files that timed out.",
    )
    ap.add_argument(
        "--package-name", default=None,
        help="Tag stamped into each audit row (defaults to today's UTC date).",
    )
    ap.add_argument("-v", "--verbose", action="store_true")
    args = ap.parse_args(argv)

    package_name = args.package_name or datetime.now(timezone.utc).strftime("%Y-%m-%d")

    t_start = time.time()
    t_mono_start = time.monotonic()

    # ── build / load inventory ────────────────────────────────────────────────
    if args.inventory_in:
        if args.verbose:
            print(f"loading inventory from {args.inventory_in}", flush=True)
        entries = read_inventory_jsonl(args.inventory_in)
    else:
        if args.verbose:
            print(f"walking {args.root}", flush=True)
        entries = walk_inventory(args.root, verbose=args.verbose)

    if args.inventory_out:
        n = write_inventory_jsonl(entries, args.inventory_out)
        print(f"inventory written: {args.inventory_out} rows={n}")
        return 0

    if args.limit is not None:
        entries = entries[: args.limit]

    # ── resume state ──────────────────────────────────────────────────────────
    resume = load_resume(args.resume_state)
    done_set = set(resume.get("done", []))

    # ── lazy import of indexer pieces (slow first-time) ──────────────────────
    from app.core.doc_index import chunk_text_with_overlap
    if not args.dry_run:
        from app.core.rag.retriever import index_chunks, available as rag_available
        if not rag_available():
            print("ERROR: RAG embedder not available; aborting.", file=sys.stderr)
            return 2

    total = len(entries)
    processed = 0
    skipped = 0
    errors = 0
    timeouts = 0
    chunks_total = 0
    ocr_required_count = 0
    pages_ocrd_total = 0

    # Per-doc watchdog: one worker process, hard timeout on each extract().
    # Init AFTER inventory is loaded so a bad inventory path fails fast in
    # the parent rather than after spawning a worker.
    watchdog = _Watchdog(args.per_doc_timeout_seconds)

    for i, e in enumerate(entries, start=1):
        # Wall-clock budget check at the TOP of every iteration so a slow doc
        # cannot push us arbitrarily over budget after it lands.
        if args.budget_seconds and (time.monotonic() - t_mono_start) >= args.budget_seconds:
            save_resume(args.resume_state, {"done": sorted(done_set)})
            watchdog.close()
            print(
                f"budget exhausted: processed={processed} / {total}, "
                f"resume next run.",
                flush=True,
            )
            return 0

        path = e["path"]
        ext = e["ext"]
        size = e["size"]

        if path in done_set:
            continue

        t0 = time.time()
        n_chunks = 0
        extract_chars = 0
        err_repr: Optional[str] = None
        ocr_required = False
        pages_ocrd = 0
        ocr_error: Optional[str] = None

        extractor_used: Optional[str] = None
        block_status: Optional[str] = None
        is_drawing: bool = False
        drawing_meta: Optional[Dict] = None
        drawing_errors: List[str] = []

        try:
            # Per-doc watchdog: hard timeout enforced in a worker process.
            res_dict, timeout_elapsed = watchdog.call(path, ext)
            if res_dict is None:
                # Timeout. Build a status=timeout audit row, append to the
                # timeouts log, mark DONE in resume-state, save state
                # immediately (don't wait for the every-50 checkpoint —
                # an unattended crash here would replay the hang), and
                # move on. Action=skipped per the operator's spec.
                _append_timeout_path(args.timeouts_log, path)
                timeout_row = {
                    "path": path,
                    "doc_id": doc_id_for(path),
                    "ext": ext,
                    "size_bytes": size,
                    "package": package_name,
                    "status": "timeout",
                    "action": "skipped",
                    "elapsed_s": round(timeout_elapsed or 0.0, 3),
                    "per_doc_timeout_seconds": args.per_doc_timeout_seconds,
                    "timestamp": now_iso(),
                }
                append_audit(args.audit, timeout_row)
                timeouts += 1
                done_set.add(path)
                save_resume(args.resume_state, {"done": sorted(done_set)})
                if args.verbose:
                    print(
                        f"  TIMEOUT {path} after {timeout_elapsed:.1f}s "
                        f"(limit {args.per_doc_timeout_seconds}s)",
                        flush=True,
                    )
                continue

            # Re-hydrate ExtractResult from the worker's plain dict so
            # the rest of the loop's `res.<field>` accesses are unchanged.
            res = ExtractResult(
                text=res_dict.get("text") or "",
                ocr_required=bool(res_dict.get("ocr_required")),
                skipped_reason=res_dict.get("skipped_reason"),
                pages_ocrd=int(res_dict.get("pages_ocrd") or 0),
                ocr_error=res_dict.get("ocr_error"),
                extractor_used=res_dict.get("extractor_used"),
                block_status=res_dict.get("block_status"),
                is_drawing=bool(res_dict.get("is_drawing")),
                drawing=res_dict.get("drawing"),
                drawing_errors=res_dict.get("drawing_errors") or [],
            )
            extractor_used = res.extractor_used
            block_status = res.block_status
            is_drawing = res.is_drawing
            drawing_meta = res.drawing
            drawing_errors = res.drawing_errors
            if res.skipped_reason:
                # log to skipped.jsonl and audit
                append_audit(args.skipped, {
                    "path": path,
                    "ext": ext,
                    "size_bytes": size,
                    "reason": res.skipped_reason,
                    "package": package_name,
                    "timestamp": now_iso(),
                })
                skipped += 1
                done_set.add(path)
                row = {
                    "path": path,
                    "doc_id": doc_id_for(path),
                    "ext": ext,
                    "size_bytes": size,
                    "n_chunks": 0,
                    "extract_chars": 0,
                    "ocr_required": False,
                    "pages_ocrd": 0,
                    "error": None,
                    "errors": [],
                    "skipped_reason": res.skipped_reason,
                    "extractor_used": extractor_used,
                    "block_status": block_status,
                    **_drawing_audit_extras(is_drawing, drawing_meta, drawing_errors),
                    "package": package_name,
                    "elapsed_ms": int((time.time() - t0) * 1000),
                    "timestamp": now_iso(),
                }
                append_audit(args.audit, row)
                continue

            text = res.text or ""
            ocr_required = res.ocr_required
            pages_ocrd = res.pages_ocrd
            ocr_error = res.ocr_error
            if ocr_required:
                ocr_required_count += 1
            pages_ocrd_total += pages_ocrd
            extract_chars = len(text)

            if not text.strip():
                # nothing to index, but treat as processed (not error)
                row = {
                    "path": path,
                    "doc_id": doc_id_for(path),
                    "ext": ext,
                    "size_bytes": size,
                    "n_chunks": 0,
                    "extract_chars": 0,
                    "ocr_required": ocr_required,
                    "pages_ocrd": pages_ocrd,
                    "error": ocr_error,
                    "errors": list(drawing_errors or []),
                    "extractor_used": extractor_used,
                    "block_status": block_status,
                    **_drawing_audit_extras(is_drawing, drawing_meta, drawing_errors),
                    "package": package_name,
                    "elapsed_ms": int((time.time() - t0) * 1000),
                    "timestamp": now_iso(),
                }
                append_audit(args.audit, row)
                done_set.add(path)
                processed += 1
                continue

            chunks = chunk_text_with_overlap(
                text, target_chars=500, overlap=50, max_chars=800,
            )
            chunks = prefix_chunks_with_path(chunks, path)
            n_chunks = len(chunks)

            doc_id = doc_id_for(path)
            if not args.dry_run and chunks:
                index_chunks(args.project_id, doc_id, chunks)

            chunks_total += n_chunks
            processed += 1
            done_set.add(path)

            row = {
                "path": path,
                "doc_id": doc_id,
                "ext": ext,
                "size_bytes": size,
                "n_chunks": n_chunks,
                "extract_chars": extract_chars,
                "ocr_required": ocr_required,
                "pages_ocrd": pages_ocrd,
                "error": ocr_error,
                "errors": list(drawing_errors or []),
                "extractor_used": extractor_used,
                "block_status": block_status,
                # Drawings now route through DrawingQTOBlock (text + structured
                # title-block + cross-refs). drawing_qto_deferred is False.
                **_drawing_audit_extras(is_drawing, drawing_meta, drawing_errors),
                "package": package_name,
                "elapsed_ms": int((time.time() - t0) * 1000),
                "timestamp": now_iso(),
            }
            append_audit(args.audit, row)
        except KeyboardInterrupt:
            save_resume(args.resume_state, {"done": sorted(done_set)})
            raise
        except Exception as exc:  # noqa: BLE001
            err_repr = repr(exc)[:500]
            errors += 1
            row = {
                "path": path,
                "doc_id": doc_id_for(path),
                "ext": ext,
                "size_bytes": size,
                "n_chunks": 0,
                "extract_chars": extract_chars,
                "ocr_required": ocr_required,
                "pages_ocrd": pages_ocrd,
                "error": err_repr,
                "errors": list(drawing_errors or []),
                "extractor_used": extractor_used,
                "block_status": block_status,
                **_drawing_audit_extras(is_drawing, drawing_meta, drawing_errors),
                "package": package_name,
                "elapsed_ms": int((time.time() - t0) * 1000),
                "timestamp": now_iso(),
            }
            append_audit(args.audit, row)
            done_set.add(path)  # don't retry forever on the same blowup
            if args.verbose:
                print(f"  ERROR {path}: {err_repr}", flush=True)

        if args.verbose or i % 25 == 0:
            name = os.path.basename(path)
            print(
                f"[{i}/{total}] {name} -> {n_chunks} chunks "
                f"(ocr_pages={pages_ocrd})",
                flush=True,
            )
        if i % 50 == 0:
            save_resume(args.resume_state, {"done": sorted(done_set)})

    # final state save
    save_resume(args.resume_state, {"done": sorted(done_set)})
    watchdog.close()
    elapsed = time.time() - t_start
    print(
        f"done. processed={processed} skipped={skipped} errors={errors} "
        f"timeouts={timeouts} "
        f"chunks_total={chunks_total} ocr_required={ocr_required_count} "
        f"pages_ocrd_total={pages_ocrd_total} elapsed={elapsed:.1f}s"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
