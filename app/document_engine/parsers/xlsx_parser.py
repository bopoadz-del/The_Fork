"""XLSX Parser — Layer 1 syntactic extraction for document engine."""
from dataclasses import dataclass, field
from typing import List, Dict, Any, Optional
from pathlib import Path


@dataclass
class XLSXDocument:
    source: str
    sheets: Dict[str, List[List[Any]]] = field(default_factory=dict)
    headers: Dict[str, List[str]] = field(default_factory=dict)


class XLSXParser:
    """Extract sheets, headers, and schedule-template fields from Excel."""

    def __init__(self, config: Optional[Dict] = None):
        self.config = config or {}

    def parse(self, file_path: str) -> XLSXDocument:
        path = Path(file_path)
        doc = XLSXDocument(source=str(path))

        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append([str(cell) if cell is not None else "" for cell in row])
                doc.sheets[sheet_name] = rows
                if rows:
                    doc.headers[sheet_name] = [str(c).strip() for c in rows[0]]
        except ImportError:
            pass

        return doc
