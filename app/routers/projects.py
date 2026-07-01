"""Project API — create projects, attach documents, gated progress tracking.

Roadmap V2 · Part 0:
  0.1  Project entity
  0.2  Readiness gate — progress tracking refuses to run on an unready project
  0.3  Execution-intent model — attaching a document stores it and runs NOTHING;
       analysis happens only when explicitly requested.
"""

import io
import math
import os
import uuid
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import Response
from pydantic import BaseModel

from app.core import audit, doc_index, file_crypto, projects as store
from app.blocks import BLOCK_REGISTRY, get_block
from app.dependencies import (
    require_user,
    block_instances,
    _create_block_instance,
)

router = APIRouter()

DATA_DIR = os.getenv("DATA_DIR", "./data")
# Cap document uploads so one large file can't OOM the shared instance — the
# whole file is read into memory here (and copied again to encrypt). Larger
# than the generic 10MB /upload cap because this path accepts BIM/schedule
# formats (.rvt/.ifc/.xer); raise MAX_DOC_UPLOAD_SIZE on a bigger box.
MAX_DOC_UPLOAD_SIZE = int(os.getenv("MAX_DOC_UPLOAD_SIZE", str(50 * 1024 * 1024)))
# Pilot guardrail: approved Drive projects with this many or fewer indexed
# documents are treated as incomplete shells and suppressed from non-admin
# project lists so pilot users land on the master corpus instead.
PILOT_INCOMPLETE_SHELL_DOC_THRESHOLD = int(
    os.getenv("PILOT_INCOMPLETE_SHELL_DOC_THRESHOLD", "1")
)
# Above this document count, project-detail skips the full-index chunk_count
# enrichment (it deserializes a tens-of-MB index blob → ~11s on the master
# corpus). The fields it produces are cosmetic; chat + listing don't need them.
_DETAIL_ENRICHMENT_DOC_LIMIT = int(
    os.getenv("DETAIL_ENRICHMENT_DOC_LIMIT", "500")
)
# Project-detail returns only the first page of documents (newest first); the
# rest load on demand via GET .../documents. Keeps a 2700-doc corpus from
# serializing every row on every workspace open.
_DOC_FIRST_PAGE = int(os.getenv("PROJECT_DETAIL_DOC_PAGE", "100"))
try:
    os.makedirs(DATA_DIR, exist_ok=True)
except PermissionError:
    import tempfile
    DATA_DIR = tempfile.gettempdir()

ALLOWED_DOC_EXTENSIONS = {
    ".pdf", ".jpg", ".jpeg", ".png", ".gif", ".webp", ".tif", ".tiff",
    ".txt", ".md", ".csv", ".json", ".xml",
    ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx",
    # Construction-domain formats the registered blocks know how to parse.
    # ezdxf reads .dxf; ifcopenshell reads .ifc; xer/mpp are schedule exports;
    # .dwg is kept even though drawing_qto rejects it with a "convert to DXF"
    # message so the upload doesn't 400 before the user sees that guidance.
    ".dxf", ".dwg", ".ifc", ".xer", ".mpp", ".rvt",
}


def _owned_or_404(
    project_id: str,
    user_id: str,
    *,
    read_only: bool = False,
    doc_limit: Optional[int] = None,
    doc_offset: int = 0,
):
    """Load a project the caller can access, or 404 (never leak existence).

    PR D — when ``read_only=True``, non-owners are also allowed to load
    admin-approved platform projects (origin='admin_drive_approved' +
    is_approved=True). Used by the read-only GET handler so users can
    open shared platform projects without owning them. Mutating
    handlers must use the default (owner-only).

    ``doc_limit``/``doc_offset`` paginate the returned documents (default None
    = all, for mutating callers that need the full set).
    """
    proj = store.get_project(
        project_id, user_id=user_id, include_admin_approved=read_only,
        doc_limit=doc_limit, doc_offset=doc_offset,
    )
    if not proj:
        raise HTTPException(404, f"Project '{project_id}' not found")
    return proj


# ── request models ──────────────────────────────────────────────────────────

class CreateProjectRequest(BaseModel):
    name: str
    client: Optional[str] = None


class CreateProjectFromDriveRequest(BaseModel):
    """PR C — user-facing variant of approve-from-drive.

    Lets any authenticated user create a personal project from a Drive
    folder they own (not just admins from a scanned cascade). The
    created row is owned by the calling user and stamped
    origin='user_drive_import' so admins can distinguish it from their
    own 'admin_drive_approved' rows and from blank 'user_create' rows.
    """
    folder_id: str
    name: str
    client: Optional[str] = None
    max_files: int = 500
    max_depth: int = 6
    role: str = "other"


class ConnectorRequest(BaseModel):
    connected: bool = True


class ProgressRequest(BaseModel):
    planned_percent: float = 0
    actual_percent: float = 0
    contract_value: float = 0
    reporting_period: Optional[str] = None
    activities: List[Dict[str, Any]] = []
    photos: List[str] = []


# ── projects ────────────────────────────────────────────────────────────────

@router.post("/v1/projects", status_code=201)
async def create_project(req: CreateProjectRequest, auth: dict = Depends(require_user)):
    """Create a project. Documents and analytics hang off this entity."""
    name = (req.name or "").strip()
    if not name:
        raise HTTPException(400, "Project name is required")
    proj = store.create_project(name, (req.client or "").strip() or None, user_id=auth["user_id"])
    audit.record("project.created", project_id=proj["id"], name=name, user_id=auth["user_id"])
    return proj


@router.post("/v1/projects/from-drive", status_code=201)
async def create_project_from_drive(
    req: CreateProjectFromDriveRequest,
    background_tasks: BackgroundTasks,
    auth: dict = Depends(require_user),
):
    """Create a user-owned project seeded from a Drive folder.

    Mirrors /v1/admin/projects/approve-from-drive but:
      * Open to any authenticated user (no admin gate).
      * Project row is owned by the caller, not the admin.
      * origin='user_drive_import' (not 'admin_drive_approved').

    Validates Drive auth eagerly (fails fast 409 if not connected),
    slugs the project name with a uniqueness suffix, persists the row,
    and queues the recursive Drive folder walk as a background task —
    the import progresses asynchronously; the response returns once
    the row is on disk so the UI can navigate into the new project
    immediately.
    """
    import re
    from app.core import drive_auth
    from app.routers.admin import _run_drive_folder_import

    folder_id = (req.folder_id or "").strip()
    name = (req.name or "").strip()
    if not folder_id:
        raise HTTPException(400, "folder_id is required")
    if not name:
        raise HTTPException(400, "name is required")

    try:
        await drive_auth.get_access_token(auth["user_id"])
    except drive_auth.DriveNotConnected:
        raise HTTPException(409, "Google Drive is not connected for your account.")
    except drive_auth.DriveAuthError as e:
        raise HTTPException(409, f"{e} Reconnect Google Drive.")

    slug = re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")[:48] or "project"
    if store.get_project(slug) is not None:
        suffix = 2
        while store.get_project(f"{slug}_{suffix}") is not None:
            suffix += 1
        slug = f"{slug}_{suffix}"

    proj = store.create_project(
        name=name,
        client=(req.client or "").strip() or None,
        user_id=auth["user_id"],
        is_approved=True,
        project_id=slug,
        origin="user_drive_import",
    )
    audit.record(
        "project.created_from_drive",
        project_id=proj["id"], name=name, user_id=auth["user_id"],
        folder_id=folder_id,
    )

    background_tasks.add_task(
        _run_drive_folder_import,
        project_id=slug, user_id=auth["user_id"],
        folder_id=folder_id, max_files=req.max_files,
        max_depth=req.max_depth, role=req.role,
    )

    return {
        "status": "queued",
        "project": proj,
        "import": {
            "folder_id": folder_id,
            "max_files": req.max_files,
            "max_depth": req.max_depth,
            "role": req.role,
        },
    }


@router.get("/v1/projects")
async def list_projects(auth: dict = Depends(require_user)):
    """List projects visible to the caller.

    PR D visibility model:
      * Admins see every project — they need the full picture to
        approve, re-index, and delete.
      * Non-admins see their own projects PLUS the admin-curated
        platform projects (origin='admin_drive_approved' AND
        is_approved=True). Other users' personal projects stay hidden.
      * ``is_approved=False`` rows never appear for non-owners — the
        column is reserved for future "detected but pending" candidates.
    """
    role = (auth.get("role") or "user").lower()
    if role == "admin":
        rows = store.list_projects()  # full set
    else:
        rows = store.list_projects(
            user_id=auth["user_id"], include_admin_approved=True,
        )
        # Pilot: hide incomplete approved shells from non-admins so they
        # gravitate to the Dar Al Arkan Master Corpus.
        rows = [
            r for r in rows
            if not (
                r.get("origin") == "admin_drive_approved"
                and not r.get("is_master_corpus")
                and r.get("document_count", 0) <= PILOT_INCOMPLETE_SHELL_DOC_THRESHOLD
            )
        ]
    return {"projects": rows}


@router.get("/v1/projects/{project_id}")
async def get_project(project_id: str, auth: dict = Depends(require_user)):
    """Project detail — documents + the computed readiness gate.

    Documents are enriched with ``chunk_count`` so the frontend can render
    a "Not indexed" badge for docs the extractor failed on (count == 0)
    without making N extra round-trips.
    """
    proj = _owned_or_404(
        project_id, auth["user_id"], read_only=True, doc_limit=_DOC_FIRST_PAGE,
    )
    # `documents` is now the first page only; `document_count` is the true total
    # (a cheap COUNT). Gate enrichment on the TOTAL, not the page length.
    doc_count = proj.get("document_count") or 0
    proj["documents_truncated"] = len(proj.get("documents") or []) < doc_count

    # Per-document chunk_count enrichment deserializes the ENTIRE doc index
    # (one JSON blob holding every document + its chunk list). On a large corpus
    # — the 2713-doc master corpus — that blob is tens of MB and deserializing
    # it dominated the project-detail response (~11s warm; latency scaled with
    # document count, sub-second for small projects). It only feeds the cosmetic
    # "Not indexed" per-doc badge, which is meaningless for the master corpus, so
    # skip the whole enrichment past a threshold and keep the load snappy.
    if doc_count <= _DETAIL_ENRICHMENT_DOC_LIMIT:
        try:
            from app.core import doc_index as _doc_index
            index = _doc_index._load_index(project_id)  # noqa: SLF001 — internal use
            chunk_counts: dict[str, int] = {}
            if index and isinstance(index.get("documents"), list):
                for entry in index["documents"]:
                    doc_id = entry.get("document_id")
                    if doc_id:
                        chunk_counts[doc_id] = len(entry.get("chunks", []))
            for doc in proj.get("documents", []) or []:
                doc["chunk_count"] = chunk_counts.get(doc.get("id"), 0)
        except Exception:
            # Enrichment is best-effort — never break the project load on it.
            pass

        # Expose the live indexed-chunk count so the UI/admin can flag
        # projects that have documents but no searchable corpus. Same
        # threshold: the count is one query but pairs with the badge above.
        try:
            from app.core.rag.embeddings import get_embedder
            from app.core.rag.vector_store import get_store
            resolved_id = store._master_corpus_source(project_id) or project_id
            embedder = get_embedder()
            chunk_store = get_store(dim=embedder.dim)
            indexed_chunks = chunk_store.count(resolved_id)
            proj["indexed_chunks"] = indexed_chunks
            proj["has_indexed_chunks"] = indexed_chunks > 0
        except Exception:
            # Vector store may not be configured in all test environments.
            proj["indexed_chunks"] = None
            proj["has_indexed_chunks"] = None
    else:
        # Large corpus: skip both enrichments. Fields stay absent/None so the
        # UI simply doesn't render the per-doc badge (correct for a corpus we
        # know is indexed). Chat + document listing are unaffected.
        proj["indexed_chunks"] = None
        proj["has_indexed_chunks"] = None

    return proj


@router.get("/v1/projects/{project_id}/documents")
async def list_project_documents(
    project_id: str,
    offset: int = 0,
    limit: int = 100,
    auth: dict = Depends(require_user),
):
    """A page of a project's documents (newest first).

    The workspace gets page 1 with the project detail and loads further pages
    here on demand, so a large corpus never serializes every row at once.
    """
    # Access check (read-only: platform projects are visible to all). doc_limit=0
    # makes the check itself cheap — it doesn't load the document rows.
    proj = _owned_or_404(
        project_id, auth["user_id"], read_only=True, doc_limit=0,
    )
    limit = max(1, min(limit, 200))
    offset = max(0, offset)
    return {
        "documents": store.list_documents(
            project_id, limit=limit, offset=offset, newest_first=True,
        ),
        "total": proj.get("document_count") or store.count_documents(project_id),
        "offset": offset,
        "limit": limit,
    }


# ── inline document preview ─────────────────────────────────────────────────
# Render generated / uploaded artifacts (schedules, BOQs, procurement lists,
# reports) in the right-panel WITHOUT forcing a download. Read-only; shared and
# master-corpus projects are previewable like the read GET.

# Per-sheet caps so a huge workbook can't serialize megabytes of cells into the
# preview payload. The frontend notes truncation to the user.
PREVIEW_MAX_ROWS = 200
PREVIEW_MAX_COLS = 40
# First N chars of extracted text (docx/txt/md) — enough to preview, bounded.
PREVIEW_TEXT_CHARS = 20_000

_TABLE_EXTS = {".xlsx", ".xls", ".csv"}
_TEXT_EXTS = {".docx", ".doc", ".txt", ".md"}


def _cell_str(value: Any) -> str:
    """Stringify a cell for the JSON payload. None and pandas NaN → ""."""
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value)


def _resolve_preview_document(
    project_id: str, document_id: str, user_id: str,
) -> Tuple[Dict[str, Any], str, str]:
    """Access-check and resolve a document for preview.

    Returns ``(doc, ext, file_path)``. Read-only access — non-owners may
    preview shared / master-corpus documents like the read GET does.
    ``doc_limit=0`` keeps the project access check cheap (no document rows
    serialized — critical for the ~2,700-doc master corpus). The master-corpus
    alias resolves to its backing source id before the ownership match, because
    ``get_document`` returns the SOURCE project_id, not the alias.
    """
    _owned_or_404(project_id, user_id, read_only=True, doc_limit=0)
    resolved_id = store._master_corpus_source(project_id) or project_id
    doc = store.get_document(document_id)
    if not doc or doc.get("project_id") != resolved_id:
        raise HTTPException(
            404, f"Document '{document_id}' not found in project '{project_id}'"
        )
    fp = doc.get("file_path")
    if not fp or not os.path.exists(fp):
        raise HTTPException(404, "Document file is not available for preview")
    _, ext = os.path.splitext((doc.get("original_name") or "").lower())
    return doc, ext, fp


def _table_preview(file_path: str, ext: str) -> Dict[str, Any]:
    """Build a ``{"kind":"table", "sheets":[...]}`` payload from a spreadsheet.

    Caps each sheet at PREVIEW_MAX_ROWS x PREVIEW_MAX_COLS and flags truncation.
    Legacy ``.xls`` needs xlrd (not installed) — pandas raises and the caller
    converts that to a 422.
    """
    sheets: List[Dict[str, Any]] = []
    truncated = False

    if ext == ".csv":
        import csv as _csv

        raw = file_crypto.read_document(file_path)
        text = raw.decode("utf-8", errors="replace")
        rows: List[List[str]] = []
        for i, row in enumerate(_csv.reader(io.StringIO(text))):
            if i >= PREVIEW_MAX_ROWS:
                truncated = True
                break
            if len(row) > PREVIEW_MAX_COLS:
                truncated = True
            rows.append([_cell_str(c) for c in row[:PREVIEW_MAX_COLS]])
        sheets.append({"name": "Sheet1", "rows": rows})

    elif ext == ".xlsx":
        import openpyxl

        with file_crypto.open_plaintext(file_path) as readable_path:
            wb = openpyxl.load_workbook(
                readable_path, data_only=True, read_only=True
            )
            try:
                for name in wb.sheetnames:
                    ws = wb[name]
                    rows = []
                    for r, row in enumerate(ws.iter_rows(values_only=True)):
                        if r >= PREVIEW_MAX_ROWS:
                            truncated = True
                            break
                        if len(row) > PREVIEW_MAX_COLS:
                            truncated = True
                        rows.append([_cell_str(c) for c in row[:PREVIEW_MAX_COLS]])
                    sheets.append({"name": name, "rows": rows})
            finally:
                wb.close()

    else:  # .xls — best effort via pandas; raises without xlrd → 422 upstream.
        import pandas as pd

        with file_crypto.open_plaintext(file_path) as readable_path:
            frames = pd.read_excel(readable_path, sheet_name=None, header=None)
        for name, df in frames.items():
            if df.shape[0] > PREVIEW_MAX_ROWS or df.shape[1] > PREVIEW_MAX_COLS:
                truncated = True
            clipped = df.iloc[:PREVIEW_MAX_ROWS, :PREVIEW_MAX_COLS]
            rows = [[_cell_str(c) for c in record] for record in clipped.values.tolist()]
            sheets.append({"name": str(name), "rows": rows})

    return {"kind": "table", "sheets": sheets, "truncated": truncated}


def _text_preview(file_path: str, ext: str) -> Dict[str, Any]:
    """Build a ``{"kind":"text", "text":...}`` payload (first PREVIEW_TEXT_CHARS).

    Legacy binary ``.doc`` is not readable by python-docx and raises → 422.
    """
    if ext in {".txt", ".md"}:
        raw = file_crypto.read_document(file_path)
        text = raw.decode("utf-8", errors="replace")
    else:  # .docx / .doc
        import docx

        with file_crypto.open_plaintext(file_path) as readable_path:
            document = docx.Document(readable_path)
            text = "\n".join(p.text for p in document.paragraphs)
    truncated = len(text) > PREVIEW_TEXT_CHARS
    return {"kind": "text", "text": text[:PREVIEW_TEXT_CHARS], "truncated": truncated}


@router.get("/v1/projects/{project_id}/documents/{document_id}/preview")
async def preview_document(
    project_id: str, document_id: str, auth: dict = Depends(require_user)
):
    """Render-friendly JSON preview of a stored document, by extension.

    * .xlsx/.xls/.csv → {"kind":"table", "sheets":[...]}
    * .pdf            → {"kind":"pdf"} (raw bytes at the sibling /preview/raw)
    * .docx/.doc/.txt/.md → {"kind":"text", "text": ...}
    * anything else   → {"kind":"unsupported", "ext": ext}

    A malformed / unreadable file returns 422 (never 500).
    """
    _doc, ext, fp = _resolve_preview_document(project_id, document_id, auth["user_id"])
    try:
        if ext in _TABLE_EXTS:
            return _table_preview(fp, ext)
        if ext == ".pdf":
            return {"kind": "pdf"}
        if ext in _TEXT_EXTS:
            return _text_preview(fp, ext)
        return {"kind": "unsupported", "ext": ext}
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001 — bad file must 422, never 500.
        raise HTTPException(
            422, f"Could not render a preview for this document: {exc}"
        )


@router.get("/v1/projects/{project_id}/documents/{document_id}/preview/raw")
async def preview_document_raw(
    project_id: str, document_id: str, auth: dict = Depends(require_user)
):
    """Raw decrypted PDF bytes so the frontend can embed the document.

    Returns the bytes in-memory (NOT a FileResponse over an open_plaintext temp
    path — that temp file is deleted when the context manager exits, before
    Starlette can stream it, which fails whenever encryption-at-rest is on). The
    upload size cap keeps buffering the whole file in memory safe for v1.
    """
    _doc, ext, fp = _resolve_preview_document(project_id, document_id, auth["user_id"])
    if ext != ".pdf":
        raise HTTPException(400, "Raw preview is available only for PDF documents")
    try:
        raw = file_crypto.read_document(fp)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(422, f"Could not read the document: {exc}")
    return Response(content=raw, media_type="application/pdf")


@router.delete("/v1/projects/{project_id}")
async def delete_project(project_id: str, auth: dict = Depends(require_user)):
    """Delete a project: its document records, facts, AND files on disk."""
    # Resolve pilot master-corpus alias before the ownership check so admins
    # can delete the shared corpus project from the UI.
    resolved_id = store._master_corpus_source(project_id) or project_id
    # Admins see every project in the list, so they must be able to load (and
    # then delete) any of them — look up unscoped for admins. Without this an
    # admin gets a 404 for a project they don't own, BEFORE the admin-bypass
    # below ever runs (the "frozen projects / not found" bug). Non-admins keep
    # the per-user scoped lookup.
    is_admin = auth.get("role") == "admin"
    proj = store.get_project(
        project_id,
        user_id=None if is_admin else auth["user_id"],
        include_admin_approved=True,
    )
    if not proj:
        raise HTTPException(404, f"Project '{project_id}' not found")
    if proj.get("user_id") != auth["user_id"] and not is_admin:
        raise HTTPException(403, "Admin or project owner required")
    # SOFT delete: archive the project — hide it from listings, detail,
    # ownership gates and retrieval — WITHOUT removing the row. `chunks` is
    # ON DELETE CASCADE on project_id, so a hard delete would destroy the
    # project's RAG chunks. The operator principle is "delete the UI, never
    # the RAG; build on it only", so documents, files on disk, and chunks all
    # stay and the project is restorable (set status back to 'active').
    store.archive_project(resolved_id)
    audit.record("project.archived", project_id=resolved_id,
                 user_id=auth["user_id"])
    return {
        "status": "archived",
        "project_id": resolved_id,
    }


@router.post("/v1/projects/{project_id}/conversations/{conversation_id}/clear")
async def clear_project_conversation(
    project_id: str,
    conversation_id: str,
    auth: dict = Depends(require_user),
):
    """Wipe one conversation's messages + facts without deleting the
    conversation row. Lets the operator escape a thread poisoned by
    prior hallucinated tool-skip / fabricated-table turns without
    creating a new project.

    Owner-only. Cross-project conversation IDs are rejected with 404 to
    avoid info-leak via timing.
    """
    # Resolve pilot master-corpus alias so the shared corpus can be cleared
    # from the workspace UI.
    resolved_id = store._master_corpus_source(project_id) or project_id
    proj = store.get_project(
        project_id, user_id=auth["user_id"], include_admin_approved=True
    )
    if not proj:
        raise HTTPException(404, f"Project '{project_id}' not found")
    if proj.get("user_id") != auth["user_id"] and auth.get("role") != "admin":
        raise HTTPException(403, "Admin or project owner required")
    from app.core import agent_memory

    # Workspace conversation IDs are deterministic (ws-{project_id}).
    # Accept both the alias and the backing project id for the master corpus.
    # Reject any other workspace prefix that doesn't match this project
    # before we let the call near agent_memory.
    expected_ids = {f"ws-{project_id}", f"ws-{resolved_id}"}
    if conversation_id.startswith("ws-") and conversation_id not in expected_ids:
        raise HTTPException(404, "Conversation not found")

    # For non-workspace conversation IDs, confirm the stored row (if any)
    # belongs to this project (alias or source id).
    conv = agent_memory.get_conversation(conversation_id)
    if conv is not None and conv.get("project_id") not in (None, "", project_id, resolved_id):
        raise HTTPException(404, "Conversation not found")

    cleared = agent_memory.clear_conversation(conversation_id)
    audit.record(
        "conversation.cleared",
        project_id=project_id,
        conversation_id=conversation_id,
        user_id=auth["user_id"],
        messages=cleared["messages"],
        facts=cleared["facts"],
    )
    return {
        "status": "cleared",
        "conversation_id": conversation_id,
        **cleared,
    }


@router.post("/v1/projects/{project_id}/connectors/aconex")
async def connect_aconex(
    project_id: str, req: ConnectorRequest, auth: dict = Depends(require_user)
):
    """Set the Aconex connection flag for a project.

    Stub for the full Oracle Aconex connector (Roadmap V2 cross-cutting item).
    Until the real OAuth/import client lands, this lets the readiness gate be
    satisfied explicitly.
    """
    _owned_or_404(project_id, auth["user_id"])
    if not store.set_aconex(project_id, req.connected):
        raise HTTPException(404, f"Project '{project_id}' not found")
    audit.record("connector.aconex", project_id=project_id,
                 connected=req.connected, user_id=auth["user_id"])
    return {
        "status": "ok",
        "project_id": project_id,
        "aconex_connected": req.connected,
        "readiness": store.compute_readiness(project_id),
    }


@router.get("/v1/projects/{project_id}/connectors")
async def list_connectors(project_id: str, auth: dict = Depends(require_user)):
    """Connector status for a project.

    Aconex is currently a connection *flag* — the full Oracle Aconex OAuth
    client is pending API credentials (Roadmap V2 open question). Setting the
    flag lets the readiness gate be satisfied for projects whose live Aconex
    feed is managed outside the platform.
    """
    proj = _owned_or_404(project_id, auth["user_id"])
    return {
        "project_id": project_id,
        "connectors": [
            {
                "name": "aconex",
                "connected": proj["aconex_connected"],
                "mode": "flag",
                "note": "Full OAuth client pending Aconex API credentials.",
            }
        ],
    }


# ── documents — store only, no analysis (Roadmap V2 · 0.3) ──────────────────

@router.post("/v1/projects/{project_id}/documents", status_code=201)
async def add_document(
    project_id: str,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    role: Optional[str] = Form(None),
    auth: dict = Depends(require_user),
):
    """Attach a document to a project.

    This STORES and CLASSIFIES the file — it runs no pipeline, no blocks, no
    analysis. Attaching a file is not the same as asking for analysis; run a
    block explicitly (via /v1/execute) when you actually want results.
    """
    proj = _owned_or_404(project_id, auth["user_id"])

    original_name = (file.filename or "unknown").strip()
    if not original_name or original_name in (".", ".."):
        raise HTTPException(400, "Invalid filename")
    original_name = os.path.basename(original_name.replace("\\", "/"))
    _, ext = os.path.splitext(original_name.lower())
    if ext not in ALLOWED_DOC_EXTENSIONS:
        raise HTTPException(400, f"File type '{ext}' not allowed")

    # Reject oversize uploads BEFORE reading the file into memory — this path
    # buffers the whole file (and copies it to encrypt), so an unbounded upload
    # of a large BIM model OOMs the single shared worker and drops every
    # concurrent user, not just the uploader.
    file.file.seek(0, 2)
    upload_size = file.file.tell()
    file.file.seek(0)
    if upload_size > MAX_DOC_UPLOAD_SIZE:
        raise HTTPException(
            413,
            f"File too large ({upload_size} bytes). Max is {MAX_DOC_UPLOAD_SIZE} bytes.",
        )

    file_id = str(uuid.uuid4())[:8]
    stored_as = f"{file_id}_{original_name}"
    filepath = os.path.join(DATA_DIR, stored_as)
    # Persist the document — encrypted at rest iff DATA_ENCRYPTION_KEY is set
    # (opt-in; plaintext otherwise — see app/core/file_crypto.py). The recorded
    # `size` is the original plaintext size, not the (larger) ciphertext size.
    file.file.seek(0)
    raw_bytes = file.file.read()
    file_crypto.write_document(filepath, raw_bytes)
    size = len(raw_bytes)

    if role is not None and role not in store.VALID_ROLES:
        raise HTTPException(
            400, f"Invalid role '{role}'. Allowed: {sorted(store.VALID_ROLES)}"
        )

    doc = store.add_document(
        project_id, original_name, stored_as, filepath, size, role=role
    )
    audit.record("document.added", project_id=project_id,
                 document_id=doc["id"], name=original_name, size=size, user_id=auth["user_id"])
    background_tasks.add_task(doc_index.maybe_eager_index, project_id, doc["id"])

    # V2 inline safety + QA/QC detection for image uploads — runs PIL +
    # COCO YOLO + the fine-tuned safety_qaqc detector and surfaces a
    # compact summary the frontend can show to the user without a
    # separate /v1/execute round-trip. Failures are non-fatal: the upload
    # still succeeds, the result just won't carry safety_qaqc data.
    safety_summary = None
    if ext in {".jpg", ".jpeg", ".png", ".webp", ".tif", ".tiff", ".bmp"}:
        try:
            from app.dependencies import get_block_instance
            image_block = get_block_instance("image")
            if image_block is not None:
                analysis = await image_block.execute(
                    {"file_path": filepath},
                    {"mode": "safety_qaqc", "prompt": "construction safety + QA/QC scan"},
                )
                body = analysis.get("result", {}) or {}
                detections = body.get("safety_qaqc") or []
                if detections:
                    safety_summary = {
                        "count": len(detections),
                        "top": [
                            {
                                "class": d.get("class"),
                                "confidence": round(float(d.get("confidence") or 0.0), 3),
                            }
                            for d in detections[:8]
                        ],
                    }
        except Exception:
            # Detection is best-effort; never fail the upload over it.
            pass

    response: Dict[str, Any] = {
        "status": "stored",
        "message": (
            f"Added '{original_name}' — classified as {doc['doc_type']} "
            f"(role: {doc['doc_role']})."
            + (f" Detected {safety_summary['count']} safety/QA-QC issue(s)."
               if safety_summary else " No analysis was run; ask in chat to analyze it.")
        ),
        "document": doc,
        "readiness": store.compute_readiness(project_id),
    }
    if safety_summary:
        response["safety_qaqc"] = safety_summary
    return response


# ── gated progress tracking (Roadmap V2 · 0.2) ──────────────────────────────

@router.post("/v1/projects/{project_id}/progress")
async def project_progress(
    project_id: str, req: ProgressRequest, auth: dict = Depends(require_user)
):
    """Run the progress tracker for a project — but only if the project is ready.

    Until the project has a baseline schedule, daily and weekly reports, and
    Aconex connected, this returns a structured 'not_ready' response naming
    exactly what is missing — never fabricated all-zero numbers.
    """
    proj = _owned_or_404(project_id, auth["user_id"])

    readiness = proj["readiness"]
    if not readiness["ready"]:
        return {
            "status": "not_ready",
            "project_id": project_id,
            "message": (
                "Project is not ready for progress tracking. "
                "Load the missing items, then try again."
            ),
            "missing": readiness["missing"],
            "readiness": readiness,
        }

    container = block_instances.get("construction")
    if container is None:
        construction_cls = get_block("construction")
        if construction_cls is None:
            raise HTTPException(
                status_code=503,
                detail="construction kit not enabled — set CEREBRUM_DOMAIN_KITS",
            )
        container = _create_block_instance(construction_cls)
        block_instances["construction"] = container

    params = req.model_dump()
    if not params.get("reporting_period"):
        params.pop("reporting_period", None)

    tracker = await container.progress_tracker({}, params)
    return {
        "status": "success",
        "project_id": project_id,
        "readiness": readiness,
        "tracker": tracker,
    }


# ── project memory (Roadmap V2 · Epic 3) ────────────────────────────────────

class FactRequest(BaseModel):
    key: str
    value: str
    source_document: Optional[str] = None
    confidence: Optional[float] = None


@router.get("/v1/projects/{project_id}/memory")
async def get_memory(
    project_id: str, q: Optional[str] = None, auth: dict = Depends(require_user)
):
    """List the durable facts known about a project (optionally keyword-filtered)."""
    _owned_or_404(project_id, auth["user_id"])
    facts = store.search_facts(project_id, q) if q else store.list_facts(project_id)
    return {"project_id": project_id, "facts": facts, "count": len(facts)}


@router.post("/v1/projects/{project_id}/memory", status_code=201)
async def add_memory(
    project_id: str, req: FactRequest, auth: dict = Depends(require_user)
):
    """Add or correct a project fact (manual entry / correction)."""
    _owned_or_404(project_id, auth["user_id"])
    key = (req.key or "").strip()
    if not key:
        raise HTTPException(400, "Fact key is required")
    return store.set_fact(
        project_id, key, req.value,
        source_document=req.source_document, confidence=req.confidence,
    )


@router.delete("/v1/projects/{project_id}/memory/{key}")
async def delete_memory(
    project_id: str, key: str, auth: dict = Depends(require_user)
):
    """Forget a single project fact."""
    _owned_or_404(project_id, auth["user_id"])
    if not store.delete_fact(project_id, key):
        raise HTTPException(404, f"Fact '{key}' not found for project '{project_id}'")
    return {"status": "deleted", "project_id": project_id, "key": key}


# ── data governance (Roadmap V2 · Epic 6) ───────────────────────────────────

@router.delete("/v1/projects/{project_id}/documents/{document_id}")
async def delete_document(
    project_id: str, document_id: str, auth: dict = Depends(require_user)
):
    """Delete a single document — its record and its file on disk."""
    _owned_or_404(project_id, auth["user_id"])
    doc = store.get_document(document_id)
    if not doc or doc.get("project_id") != project_id:
        raise HTTPException(
            404, f"Document '{document_id}' not found in project '{project_id}'"
        )
    fp = doc.get("file_path")
    file_removed = False
    if fp and os.path.exists(fp):
        try:
            os.remove(fp)
            file_removed = True
        except OSError:
            pass
    store.delete_document(document_id)
    # Drop the deleted doc from the project's doc_index too — otherwise its
    # stale chunks keep surfacing in RAG retrieval (verified failure mode on
    # the Diriyah project where a deleted duplicate kept appearing as a
    # Sources-footer entry).
    index_pruned = False
    try:
        from app.core import doc_index as _doc_index

        def _drop(current):
            current = current or {"project_id": project_id, "documents": [], "skipped": []}
            current["documents"] = [
                d for d in (current.get("documents") or [])
                if d.get("document_id") != document_id
            ]
            current["skipped"] = [
                s for s in (current.get("skipped") or [])
                if s.get("document_id") != document_id
            ]
            return current

        _doc_index._update_index(project_id, _drop)  # noqa: SLF001
        index_pruned = True
    except Exception:  # noqa: BLE001 — never block delete on index cleanup
        pass
    audit.record("document.deleted", project_id=project_id,
                 document_id=document_id, file_removed=file_removed,
                 index_pruned=index_pruned, user_id=auth["user_id"])
    return {
        "status": "deleted",
        "document_id": document_id,
        "file_removed": file_removed,
        "index_pruned": index_pruned,
    }


@router.get("/v1/projects/{project_id}/audit")
async def project_audit(
    project_id: str, limit: int = 100, auth: dict = Depends(require_user)
):
    """The audit trail for a project — uploads, deletions, purges."""
    _owned_or_404(project_id, auth["user_id"])
    return {
        "project_id": project_id,
        "entries": audit.read_audit(limit, project_id),
    }


@router.get("/v1/governance")
async def governance_status(auth: dict = Depends(require_user)):
    """Where client data lives and how long it is kept (Roadmap V2 · Epic 6)."""
    if auth["role"] != "admin":
        raise HTTPException(403, "Admin only")
    retention = int(os.getenv("DATA_RETENTION_DAYS", "0") or "0")
    return {
        "data_directory": os.getenv("DATA_DIR", "./data"),
        "retention_days": retention if retention > 0 else "indefinite",
        "audit_logging": True,
        "delete_on_request": True,
        "policy_document": "DATA_GOVERNANCE.md",
    }


@router.post("/v1/governance/purge")
async def governance_purge(auth: dict = Depends(require_user)):
    """Purge documents older than DATA_RETENTION_DAYS (no-op if unset)."""
    if auth["role"] != "admin":
        raise HTTPException(403, "Admin only")
    days = int(os.getenv("DATA_RETENTION_DAYS", "0") or "0")
    if days <= 0:
        return {"status": "skipped",
                "reason": "DATA_RETENTION_DAYS is not set", "purged": 0}
    purged = store.purge_documents_older_than(days)
    files_removed = 0
    # Audit each purged document individually — the summary "governance.purge"
    # entry below records only a count, not the IDs. Without per-row entries
    # there's no forensic trail of which specific documents the bulk purge
    # removed (the "BOQ disappeared with no explanation" failure mode).
    for doc in purged:
        audit.record(
            "document.deleted",
            project_id=doc.get("project_id"),
            document_id=doc.get("id"),
            name=doc.get("original_name"),
            reason="governance_purge",
            retention_days=days,
            user_id=auth["user_id"],
        )
        fp = doc.get("file_path")
        if fp and os.path.exists(fp):
            try:
                os.remove(fp)
                files_removed += 1
            except OSError:
                pass
    audit.record("governance.purge",
                 documents_purged=len(purged), files_removed=files_removed, user_id=auth["user_id"])
    return {
        "status": "purged",
        "documents_purged": len(purged),
        "files_removed": files_removed,
    }
