"""Construction container — documents submodule."""

import logging
import os
import re
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from app.core.construction_types import Measurement, SpecItem, RiskItem

from .helpers import _parse_money_str, _safe_float, _safe_iso_date

logger = logging.getLogger(__name__)


def _om_outline_from_text(text: str) -> Optional[Dict[str, Any]]:
    """O&M outline from the operator brief — no invented equipment tags."""
    t = text or ""
    if not re.search(
        r"operations and maintenance|o\s*&\s*m\b|o and m|maintenance manual",
        t,
        re.I,
    ):
        return None
    loc = ""
    m = re.search(r"\b(PWPS[-\s]?\d+[^\s,]*)\b", t, re.I)
    if m:
        loc = m.group(1)
    title = (
        f"Operations and Maintenance manual outline"
        + (f" — {loc}" if loc else "")
    )
    sections = [
        {"section": "1. Purpose and scope",
         "content": f"O&M coverage for {loc or 'the stated facility'} as requested by the operator."},
        {"section": "2. Existing-services interfaces",
         "content": "Interfaces from Specification Vol 2 Existing Services — do not invent utility owners."},
        {"section": "3. Environmental constraints",
         "content": "Environment Requirements SoW and Pre-Mobilization Environmental Checklist obligations."},
        {"section": "4. Reservoir / wet-test operations",
         "content": "First wet test, isolation, fill and drawdown — watertightness only, not a membrane electrical test."},
        {"section": "5. Mechanical / pumping",
         "content": "Pump station plant as named in the project documents; tags left blank if not supplied."},
        {"section": "6. Electrical and controls",
         "content": "Power, isolation, and instrumentation referenced in the brief."},
        {"section": "7. Permits and NOCs",
         "content": "Keep live-road / diversion permits with the O&M pack when they affect access."},
        {"section": "8. Spares and consumables",
         "content": "Do not invent a spares list — attach manufacturer schedules when supplied."},
        {"section": "9. Maintenance frequencies",
         "content": "Manufacturer + specifier frequencies only; otherwise mark TBD."},
        {"section": "10. Handover records",
         "content": "As-built volumes, pour records, and commissioning certificates."},
        {"section": "11. Emergency contacts",
         "content": "Engineer, contractor, and Employer contacts from Contract Data when present."},
        {"section": "12. Training",
         "content": "Operator training against this outline once equipment data exists."},
        {"section": "13. Review cycle",
         "content": "Update after first wet test and at taking-over."},
    ]
    return {
        "status": "success",
        "action": "om_manual_generated",
        "execution_mode": "outline",
        "title": title,
        "location": loc,
        "sections": sections,
        "note": (
            "Outline from the operator brief — equipment tags were not invented. "
            "Supply equipment_list or a BIM extract to expand each system."
        ),
    }


def _safety_briefing_from_text(text: str) -> Optional[Dict[str, Any]]:
    t = text or ""
    if not re.search(
        r"safety briefing|haul[- ]road|public-interface|road diversion",
        t,
        re.I,
    ):
        return None
    briefing = (
        "Live-haul-road and public-interface safety briefing "
        "for the Green Village diversion.\n\n"
        "### Signage\n"
        "Install and maintain the diversion / haul-road signs shown on "
        "the Green Village road-diversion safety-signage drawing and the "
        "Access and Haul Road Site Modifications revision. Keep warning, "
        "direction, and temporary-works signs visible at every public interface.\n\n"
        "### Speed control\n"
        "Enforce the posted site and diversion speed limits. Place speed "
        "repeater signs on the live haul road and at the village interface. "
        "No overtaking in the diversion throat.\n\n"
        "### Pedestrian crossing\n"
        "Keep public crossings signed, lit if used after dusk, and marshalled "
        "while haul traffic is running. Do not leave an unsigned gap between "
        "the haul road and the village edge.\n\n"
        "This briefing is drafted from the operator-named drawings. It is not "
        "a substitute for the approved TMP."
    )
    return {
        "status": "success",
        "action": "safety_briefing",
        "title": "Live-haul-road safety briefing — Green Village diversion",
        "signage": "Diversion and haul-road signs per the named drawings.",
        "speed": "Posted site / diversion speed; no overtaking in the throat.",
        "pedestrian": "Signed, marshalled crossings at the village interface.",
        "briefing": briefing,
        "note": "Draft from operator-named drawings — not an approved TMP.",
    }


def _as_built_volume_facts_from_text(text: str) -> Optional[Dict[str, Any]]:
    """Draft an as-built volume note from planned vs poured figures.

    Live M12: planned 350 m3 versus poured 310 m3 — no drawing pair, so
    ``as_built_deviation_report`` used to error and the hat then 413'd.
    """
    t = text or ""
    planned = re.search(
        r"planned\s+(\d+(?:\.\d+)?)\s*m(?:³|3)\b",
        t,
        re.IGNORECASE,
    )
    poured = re.search(
        r"(?:poured|actual|as[-\s]?built)\s+(\d+(?:\.\d+)?)\s*m(?:³|3)\b",
        t,
        re.IGNORECASE,
    )
    vs = re.search(
        r"(\d+(?:\.\d+)?)\s*m(?:³|3)\b.{0,24}(?:versus|vs\.?|against)\s+"
        r"(\d+(?:\.\d+)?)\s*m(?:³|3)\b",
        t,
        re.IGNORECASE,
    )
    p_m3 = float(planned.group(1)) if planned else None
    a_m3 = float(poured.group(1)) if poured else None
    if vs:
        if p_m3 is None:
            p_m3 = float(vs.group(1))
        if a_m3 is None:
            a_m3 = float(vs.group(2))
    if p_m3 is None or a_m3 is None:
        return None
    shortfall = round(p_m3 - a_m3, 3)
    pct = round((shortfall / p_m3) * 100, 2) if p_m3 else 0.0
    dates = re.findall(
        r"\b(\d{1,2}\s+\w+\s+\d{4})\b",
        t,
    )
    loc = re.search(r"\b(PWPS[-\s]?\d+[^\s,]*)\b", t, re.IGNORECASE)
    return {
        "status": "success",
        "action": "as_built_deviation_report",
        "execution_mode": "drafted",
        "element_type": "volume",
        "location": loc.group(1) if loc else "",
        "planned_m3": p_m3,
        "poured_m3": a_m3,
        "shortfall_m3": shortfall,
        "shortfall_percent": pct,
        "pour_dates": dates,
        "deviation_summary": {
            "total_deviations": 1,
            "critical": 0,
            "major": 1 if abs(pct) >= 10 else 0,
            "minor": 0 if abs(pct) >= 10 else 1,
        },
        "deviations": [{
            "type": "volume_shortfall",
            "planned_m3": p_m3,
            "as_built_m3": a_m3,
            "delta_m3": shortfall,
            "percent": pct,
            "severity": "major" if abs(pct) >= 10 else "minor",
        }],
        "note": (
            f"As-built volume note from operator figures: planned {p_m3:g} m³ "
            f"versus poured {a_m3:g} m³ — shortfall {shortfall:g} m³ "
            f"({pct:g}%). Not a drawing-to-drawing dimensional comparison."
        ),
    }


class ConstructionDocumentsMixin:
    ui_schema = {
        "input": {
            "type": "file",
            "accept": [".pdf", ".ifc", ".dwg", ".jpg", ".png", ".xer", ".xml"],
            "placeholder": "Upload construction drawing, BIM model, schedule, or contract...",
            "multiline": True
        },
        "output": {
            "type": "table",
            "fields": [
                {"name": "concrete_volume_m3", "type": "number", "unit": "m³", "label": "Concrete"},
                {"name": "steel_weight_kg", "type": "number", "unit": "kg", "label": "Steel"},
                {"name": "floor_area_m2", "type": "number", "unit": "m²", "label": "Floor Area"},
                {"name": "rebar_length_m", "type": "number", "unit": "m", "label": "Rebar"},
                {"name": "confidence", "type": "percentage", "label": "Confidence"}
            ]
        },
        "quick_actions": [
            {"icon": "📐", "label": "Measure Drawing", "prompt": "Extract all measurements from this drawing"},
            {"icon": "📊", "label": "Calculate Quantities", "prompt": "Calculate BOQ from this drawing"},
            {"icon": "⚠️", "label": "Check Compliance", "prompt": "Check this against Saudi building codes"},
            {"icon": "🌱", "label": "Carbon Estimate", "prompt": "Estimate embodied carbon for this project"},
            {"icon": "📅", "label": "Analyze Schedule", "prompt": "Analyze this Primavera schedule for risks"}
        ]
    }
    default_config = {
        "confidence_threshold": 0.85,
        "default_trade": "concrete"
    }
    requires = [
        "pdf", "ocr", "image",
        # Week 1
        "boq_processor", "spec_analyzer", "sympy_reasoning",
        # Week 2
        "drawing_qto", "primavera_parser", "smart_orchestrator",
        # Week 3
        "formula_executor_v2", "bim_extractor",
        # Week 4
        "learning_engine", "recommendation_template",
        # historical_benchmark removed — learning_engine accumulates real data
    ]
    tags = ["domain", "container", "aec", "construction", "bim"]
    layer = 3
    description = "Complete AEC suite: BIM, QA/QC, scheduling, contracts, specs, safety, carbon, procurement, risk"
    version = "3.1"
    name = "construction"
    """
    Construction Container: Complete AEC suite - BIM, QA/QC, scheduling,
    contracts, specs, safety, carbon, procurement, risk
    """
    async def process_document(self, input_data: Any, params: Dict) -> Dict:
        data = input_data if isinstance(input_data, dict) else {}
        p = params or {}
        file_path = data.get("file_path") or p.get("file_path")
        url = data.get("url") or p.get("url")
        doc_type = p.get("doc_type", "auto")

        if not file_path and url:
            file_path = await self._download_file(url)

        if not file_path:
            return {"status": "error", "error": "No file provided"}

        if doc_type == "auto":
            doc_type = await self._classify_document(file_path)

        cache_key = await self._get_or_create_cache_key(file_path, doc_type)

        from app.blocks import BLOCK_REGISTRY
        cache_block = BLOCK_REGISTRY.get("cache_manager")
        if cache_block:
            try:
                cache_instance = cache_block()
                cached = await cache_instance.execute(
                    {"key": cache_key}, {"action": "get", "key": cache_key}
                )
                if cached.get("cached") and cached.get("value") is not None:
                    cached_value = cached["value"]
                    if isinstance(cached_value, dict):
                        cached_value["_source"] = "cache"
                        cached_value["_cache_key"] = cache_key
                    return cached_value
            except Exception:
                logger.warning(
                    "swallowed %s in process_document() — continuing",
                    "Exception", exc_info=True,
                )

        file_size = 0
        hasher_block = BLOCK_REGISTRY.get("file_hasher")
        if hasher_block:
            try:
                hasher_instance = hasher_block()
                hash_result = await hasher_instance.execute(
                    {"file_path": file_path}, {"action": "metadata"}
                )
                if hash_result.get("status") == "success":
                    file_size = hash_result.get("size", 0)
            except Exception:
                logger.warning(
                    "swallowed %s in process_document() — continuing",
                    "Exception", exc_info=True,
                )

        if file_size > 10 * 1024 * 1024:
            async_block = BLOCK_REGISTRY.get("async_processor")
            if async_block:
                try:
                    async_instance = async_block()
                    task_payload = {
                        "task_name": "block:construction.process_document",
                        "file_path": file_path,
                        "doc_type": doc_type,
                        "data": data,
                        "params": p,
                    }
                    queued = await async_instance.execute(
                        task_payload,
                        {
                            "action": "submit",
                            "task_name": "block:construction.process_document",
                        },
                    )
                    return {
                        "status": "queued",
                        "_source": "async_queue",
                        "_cache_key": cache_key,
                        "file_size": file_size,
                        "queued": queued,
                    }
                except Exception:
                    logger.warning(
                        "swallowed %s in process_document() — continuing",
                        "Exception", exc_info=True,
                    )

        processors = {
            "drawing": self._process_drawing,
            "specification": self.process_specification_full,
            "contract": self.process_contract,
            "schedule": self.parse_primavera_schedule,
            "bom": self._process_bill_of_materials,
            "report": self._process_report,
            "bim": self._process_ifc,
            "image": self._process_site_photo,
            "change_order": self.change_order_impact,
            "safety_audit": self.safety_compliance_audit,
        }

        processor = processors.get(doc_type, self._process_drawing)
        p["file_path"] = file_path
        result = await processor(file_path, p)

        # NOTE: a former llm_enhancer post-process was removed here (W3). It had
        # been dead since it was written — it called `json.dumps(result)` without
        # `json` imported, so every invocation raised NameError, was swallowed by
        # a bare `except: pass`, and `llm_enhanced` was never set. Re-enabling it
        # would fire an LLM call on every successful document parse (a cost /
        # latency / provider-routing change), so it is intentionally not revived
        # here; the document processors already return real structured data.

        if cache_block:
            try:
                cache_instance = cache_block()
                await cache_instance.execute(
                    result, {"action": "set", "key": cache_key, "ttl": 7200}
                )
            except Exception:
                logger.warning(
                    "swallowed %s in process_document() — continuing",
                    "Exception", exc_info=True,
                )

        if isinstance(result, dict):
            result["_cache_key"] = cache_key
            result["_source"] = "processor"
        return result
    async def _process_drawing(self, file_path: str, params: Dict) -> Dict:
        # Use pre-extracted text if provided from chain
        pre_extracted_text = params.get("extracted_text", "")
    
        try:
            import fitz
            doc = fitz.open(file_path)
        except Exception as e:
            return {"status": "error", "error": f"[DRAWING_V2] Could not open file: {str(e)}", "file": file_path}
    
        result = {
            "status": "success",
            "doc_type": "drawing",
            "file_name": Path(file_path).name,
            "drawing_number": self._extract_drawing_number(Path(file_path).name),
            "revision": self._extract_revision(Path(file_path).name),
            "total_pages": len(doc),
            "sheets": [],
            "measurements": [],
            "tables": [],
            "annotations": [],
            "specifications": [],
            "detected_disciplines": [],
            "scale": None,
            "title_block": {},
            "bom_items": [],
            "confidence": {},
            "used_pre_extracted_text": bool(pre_extracted_text)  # Flag to indicate source
        }
    
        for page_num in range(len(doc)):
            page = doc[page_num]
            sheet_data = self._process_drawing_page(page, page_num, pre_extracted_text if page_num == 0 else "")
            result["sheets"].append(sheet_data)
            result["measurements"].extend(sheet_data["measurements"])
            result["tables"].extend(sheet_data["tables"])
            result["annotations"].extend(sheet_data["annotations"])
            result["specifications"].extend(sheet_data["specs"])
            result["detected_disciplines"].extend(self._detect_disciplines(sheet_data["raw_text"]))
    
        if result["sheets"]:
            result["title_block"] = self._extract_title_block(result["sheets"][0])
            result["scale"] = self._extract_scale(result["sheets"][0]["raw_text"])
    
        result["quantities"] = self._calculate_quantities(result["measurements"])
        result["cost_estimate"] = self._estimate_costs(result["quantities"])
        result["carbon_estimate"] = self._estimate_carbon(result["quantities"])
        result["confidence"] = self._calculate_confidence(result)
        result["auto_risks"] = await self._detect_risks_from_drawing(result)
    
        doc.close()
        return result
    async def process_contract(self, input_data: Any, params: Dict) -> Dict:
        data = input_data if isinstance(input_data, dict) else {}
        p = params or {}
        file_path = data.get("file_path") or p.get("file_path")
        contract_type = p.get("contract_type", "general")
    
        if not file_path:
            return {"status": "error", "error": "No contract file provided"}
    
        try:
            import fitz
            doc = fitz.open(file_path)
            full_text = ""
            for page in doc:
                full_text += page.get_text()
            doc.close()
        except Exception as e:
            return {"status": "error", "error": f"Could not read contract: {str(e)}"}
    
        clause_patterns = {
            "payment_terms": r'(?:payment|pay|invoice)[\s\w]{0,50}(?:term|schedule|milestone|certificate)',
            "liquidated_damages": r'(?:liquidated damages|ld|delay damages)[\s\w]{0,100}(?:rate|amount|per day)',
            "retention": r'(?:retention|retainage)[\s\w]{0,50}(?:percent|percentage|amount|release)',
            "insurance": r'(?:insurance|indemnif)[\s\w]{0,100}(?:required|shall|must|coverage)',
            "termination": r'(?:terminat|cancel|end)[\s\w]{0,100}(?:notice|for cause|convenience)',
            "force_majeure": r'(?:force majeure|unforeseen|beyond control|delay event)[\s\w]{0,100}(?:excus|reliev|not liable)',
            "dispute_resolution": r'(?:dispute|arbitration|mediation|adjudication)[\s\w]{0,100}(?:shall|must|proceed)',
        }
    
        extracted_clauses = {}
        for clause_name, pattern in clause_patterns.items():
            matches = list(re.finditer(pattern, full_text, re.IGNORECASE))
            extracted_clauses[clause_name] = {
                "found": len(matches) > 0,
                "count": len(matches),
                "examples": [m.group(0)[:200] for m in matches[:3]]
            }
    
        obligations = self._extract_obligations(full_text)
        contract_risks = self._assess_contract_risks(extracted_clauses, contract_type)
        financial_terms = self._extract_financial_terms(full_text)
    
        return {
            "status": "success",
            "action": "contract_analysis",
            "file_name": Path(file_path).name,
            "contract_type": contract_type,
            "document_length": len(full_text),
            "clauses_found": len([c for c in extracted_clauses.values() if c.get("found")]),
            "total_clauses": len(clause_patterns),
            "extracted_clauses": extracted_clauses,
            "key_obligations": obligations,
            "financial_terms": financial_terms,
            "risk_assessment": {
                "overall_score": contract_risks["score"],
                "risk_level": contract_risks["level"],
                "critical_issues": contract_risks["critical"],
                "warnings": contract_risks["warnings"],
                "recommendations": contract_risks["recommendations"]
            },
            "summary": self._generate_contract_summary(extracted_clauses, financial_terms)
        }
    @staticmethod
    def _split_csi_divisions(full_text: str, division_filter=None) -> tuple:
        """CSI MasterFormat division-splitting (container-only — the block has no equivalent).

        Groups raw spec text into Divisions 01–49 by leading 2-digit codes.
        Returns (detected_divisions, division_spec_items).
        """
        divisions = {i: [] for i in range(1, 50)}
        current_division = None
        for line in full_text.split('\n'):
            # \s{2,} is intentional and unified across PDF and extracted-text
            # inputs: the old file-path-only path used \s{3,}, but \s{2,} is the
            # more permissive of the two and matches everything \s{3,} would.
            m = re.match(r'^(\d{2})\s{2,}', line)
            if m:
                div_num = int(m.group(1))
                if 1 <= div_num <= 49:
                    current_division = div_num
                    divisions[current_division].append(line.strip())
            elif current_division and line.strip():
                divisions[current_division].append(line.strip())

        detected = [i for i, c in divisions.items() if c]
        division_items = []
        for div_num, content in divisions.items():
            if not content:
                continue
            if division_filter and str(div_num) != str(division_filter):
                continue
            division_items.append({
                "category": f"Division {div_num:02d}",
                "key": "content",
                "value": f"{len(content)} paragraphs",
                "section": "general",
                "confidence": 0.9,
            })
        return detected, division_items
    async def process_specification_full(self, input_data: Any, params: Dict) -> Dict:
        """Analyse a project specification.

        Delegates genuine grade / material / compliance extraction to the
        spec_analyzer block — no demo mode, no fabricated divisions. The CSI
        MasterFormat division-splitting layer (which the block has no equivalent
        for) stays here in the container.
        """
        data = input_data if isinstance(input_data, dict) else {}
        p = params or {}
        file_path = data.get("file_path") or p.get("file_path")
        extracted_text = data.get("extracted_text") or p.get("extracted_text") or ""
        division_filter = p.get("division")

        if not file_path and not extracted_text:
            return {
                "status": "error",
                "action": "specification_analysis",
                "error": "No specification provided — pass file_path (PDF) or extracted_text",
            }

        block = self._get_spec_analyzer_block()
        if block is None:
            return {
                "status": "error",
                "action": "specification_analysis",
                "error": "spec_analyzer block unavailable — cannot extract grades/materials/compliance",
            }

        # Delegate grade/material/compliance extraction to the block.
        block_input = {"file_path": file_path} if file_path else {"text": extracted_text}
        result = await block.process(block_input, p)
        if not isinstance(result, dict) or result.get("status") != "success":
            err = result.get("error") if isinstance(result, dict) else "spec_analyzer block failed"
            return {
                "status": "error",
                "action": "specification_analysis",
                "error": err or "spec_analyzer block failed",
            }

        grade_requirements = result.get("grade_requirements", []) or []
        material_specs = result.get("material_specs", []) or []
        compliance_flags = result.get("compliance_flags", []) or []

        # CSI division-splitting — container-only layer, the block has no equivalent.
        # The block already extracted PDF text; re-read it here only for splitting.
        full_text = extracted_text
        if file_path:
            try:
                import fitz
                doc = fitz.open(file_path)
                # Join with newline so the last line of page N and the first
                # line of page N+1 stay distinct; previously a bare "".join
                # silently merged them and broke `_split_csi_divisions`'s
                # line-anchored regex on multi-page spec PDFs.
                full_text = "\n".join(page.get_text() for page in doc)
                doc.close()
            except Exception as e:
                return {"status": "error", "action": "specification_analysis",
                        "error": f"Could not read spec file for division-splitting: {str(e)}"}

        detected_divisions, division_items = self._split_csi_divisions(full_text, division_filter)

        # Map the block's output into the spec_items shape callers expect:
        # one item per CSI division, plus one item per extracted grade / material.
        spec_items = list(division_items)
        for g in grade_requirements:
            spec_items.append({
                "category": "Grade Requirement",
                "key": g.get("type", "grade"),
                "value": g.get("value", ""),
                "section": g.get("context", ""),
                "confidence": 0.9,
            })
        for m in material_specs:
            spec_items.append({
                "category": "Material Spec",
                "key": m.get("material_type", "material"),
                "value": m.get("specification", ""),
                "section": "materials",
                "confidence": 0.85,
            })

        # Derive testing / QA-QC response keys from the block's compliance flags
        # (the block's compliance_flags supersede the old binary sentinel helpers).
        testing_flags = {"test_certificate"}
        qaqc_flags = {"shop_drawing", "mockup_required", "submittal", "material_approval", "approval_required"}
        testing_requirements = [
            f.get("context", f.get("keyword", "")) for f in compliance_flags
            if f.get("flag_type") in testing_flags
        ]
        qa_qc_requirements = [
            f.get("context", f.get("keyword", "")) for f in compliance_flags
            if f.get("flag_type") in qaqc_flags
        ]

        # Top-up pass over full_text — the block's compliance keyword/pattern sets
        # are narrower than the old _extract_testing_requirements / _extract_qaqc
        # helpers, which fired on bare words. Restore that coverage and UNION it
        # with the block's richer flags (deduplicated, order-preserving).
        testing_requirements = self._topup_keyword_matches(
            full_text, ["test", "sample", "lab"], existing=testing_requirements,
        )
        qa_qc_requirements = self._topup_keyword_matches(
            full_text, ["inspection", "witness", "hold point", "hold-point"],
            existing=qa_qc_requirements,
        )

        # materials_referenced: UNION the block-derived material types with a
        # substring pass over the old _extract_materials 10-keyword set, since the
        # block's material_specs drop brick/block/glass/aluminum/timber. Deduplicated.
        material_keywords = [
            "concrete", "steel", "rebar", "brick", "block", "glass",
            "aluminum", "timber", "insulation", "membrane",
        ]
        materials_seen = set()
        materials_referenced = []
        for m in material_specs:
            mt = m.get("material_type", "")
            if mt and mt.lower() not in materials_seen:
                materials_seen.add(mt.lower())
                materials_referenced.append(mt)
        lowered_text = full_text.lower()
        for kw in material_keywords:
            if kw in lowered_text and kw not in materials_seen:
                materials_seen.add(kw)
                materials_referenced.append(kw)
        materials_referenced.sort()

        return {
            "status": "success",
            "action": "specification_analysis",
            "file_name": Path(file_path).name if file_path else "extracted_text",
            "divisions_found": detected_divisions,
            "division_filter_applied": division_filter,
            "total_sections_analyzed": len(spec_items),
            "spec_items": spec_items,
            "grade_requirements": grade_requirements,
            "material_specs": material_specs,
            "compliance_flags": compliance_flags,
            "materials_referenced": materials_referenced,
            "methods_specified": [],
            "testing_requirements": testing_requirements,
            "qa_qc_requirements": qa_qc_requirements,
            "standards_referenced": result.get("standards_referenced", []),
        }
    async def safety_compliance_audit(self, input_data: Any, params: Dict) -> Dict:
        data = input_data if isinstance(input_data, dict) else {}
        p = params or {}
    
        audit_type = p.get("audit_type", "general")
        photos = data.get("photos", p.get("photos", []))
    
        if not photos and data.get("file_path"):
            photos = [data.get("file_path")]
    
        if not photos:
            return {
                "status": "error",
                "error": (
                    "No site photos supplied — provide a 'photos' list or a "
                    "'file_path' for image-based safety compliance analysis"
                ),
                "audit_type": audit_type,
            }
    
        violations = []
        compliant_items = []
    
        for photo_path in photos[:10]:
            analysis = await self._analyze_safety_photo(photo_path, audit_type)
        
            if analysis.get("hazards_detected", 0) > 0:
                violations.extend(analysis.get("hazards", []))
            else:
                compliant_items.append({
                    "photo": analysis.get("photo"),
                    "status": "compliant",
                    "notes": "No obvious violations detected"
                })
    
        severity_counts = {"critical": 0, "major": 0, "minor": 0}
        for v in violations:
            sev = v.get("severity", "minor")
            severity_counts[sev] = severity_counts.get(sev, 0) + 1
    
        return {
            "status": "success",
            "action": "safety_audit",
            "audit_type": audit_type,
            "photos_analyzed": len(photos),
            "violations_found": len(violations),
            "severity_breakdown": severity_counts,
            "violations": violations[:20],
            "compliant_items": compliant_items,
            "overall_compliance": "fail" if severity_counts["critical"] > 0 else "pass with observations" if severity_counts["major"] > 0 else "pass",
            "recommendations": self._generate_safety_recommendations(violations)
        }
    async def _analyze_safety_photo(self, photo_path: str, audit_type: str) -> Dict:
        image_block = self.get_dep("image")
        safety_prompts = {
            "general": "Identify safety hazards: missing PPE, trip hazards, exposed edges, improper storage",
            "scaffolding": "Check: guardrails, midrails, toeboards, plank overhang, base plates, access",
            "excavation": "Check: shoring, sloping, benching, spoil pile distance, access/egress",
            "electrical": "Check: exposed wires, GFCI, panel access, temporary power, grounding",
            "fall_protection": "Check: guardrails, harnesses, anchor points, lifelines, hole covers"
        }
    
        if image_block:
            try:
                analysis = await image_block.execute(
                    {"file_path": photo_path},
                    {"prompt": safety_prompts.get(audit_type, safety_prompts["general"]),
                     "mode": "safety_qaqc"}
                )
                result_body = analysis.get("result", {})
                desc = result_body.get("description", "")
                safety_qaqc = result_body.get("safety_qaqc") or []
            except Exception:
                desc = ""
                safety_qaqc = []
        else:
            desc = ""
            safety_qaqc = []

        hazards_found = self._parse_safety_hazards(desc)
        # Compose YOLO-derived hazards on top; dedup by hazard type to avoid
        # double-counting when text and vision both flag the same issue.
        yolo_hazards = self._classes_to_hazards(safety_qaqc)
        seen_types = {h["type"] for h in hazards_found}
        for yh in yolo_hazards:
            if yh["type"] not in seen_types:
                hazards_found.append(yh)
                seen_types.add(yh["type"])
        return {
            "photo": Path(photo_path).name,
            "hazards_detected": len(hazards_found),
            "hazards": hazards_found,
            "overall_assessment": "unsafe" if hazards_found else "compliant",
            "requires_immediate_action": any(h.get("severity") == "critical" for h in hazards_found)
        }
    async def as_built_deviation_report(self, input_data: Any, params: Dict) -> Dict:
        """Compare as-built conditions against design drawings."""
        data = input_data if isinstance(input_data, dict) else {}
        p = params or {}

        as_built_file = data.get("as_built_file") or p.get("as_built_file")
        design_file = data.get("design_file") or p.get("design_file")
        tolerance_mm = float(p.get("tolerance_mm", 10))
        element_type = p.get("element_type", "general")

        text = " ".join(
            str(x)
            for x in (
                p.get("text"),
                p.get("user_message"),
                data.get("text"),
                data.get("user_message"),
                data.get("message"),
                input_data if isinstance(input_data, str) else "",
            )
            if x
        )
        volume_note = _as_built_volume_facts_from_text(text)
        if volume_note and not (as_built_file and design_file) and not (
            data.get("measurements") or p.get("as_built_measurements")
        ):
            return volume_note

        # Honest gate — NEVER return a conformance/APPROVED verdict without a real
        # comparison. No inputs => error, not "0 deviations, APPROVED".
        if as_built_file and design_file:
            deviations = await self._compare_as_built_to_design(as_built_file, design_file, tolerance_mm)
            if deviations is None:
                return {
                    "status": "error",
                    "action": "as_built_deviation_report",
                    "error": (
                        "Could not extract comparable dimensions from the as-built and design "
                        "files. Provide DXF/PDF drawings with measurable dimensions, or supply "
                        "measurements + design_measurements as [{type, value, unit}]."
                    ),
                }
        elif data.get("measurements") or p.get("as_built_measurements"):
            as_built_m = data.get("measurements") or p.get("as_built_measurements", [])
            design_m = p.get("design_measurements", [])
            if not design_m:
                return {
                    "status": "error",
                    "action": "as_built_deviation_report",
                    "error": "as-built measurements supplied but no design_measurements to compare against.",
                }
            deviations = self._compare_measurement_sets(as_built_m, design_m, tolerance_mm)
        else:
            return {
                "status": "error",
                "action": "as_built_deviation_report",
                "error": (
                    "As-built deviation report requires either (as_built_file + design_file) or "
                    "(measurements + design_measurements). No comparison inputs supplied."
                ),
            }

        critical = [d for d in deviations if d.get("severity") == "critical"]
        major = [d for d in deviations if d.get("severity") == "major"]
        minor = [d for d in deviations if d.get("severity") == "minor"]

        return {
            "status": "success",
            "action": "as_built_deviation_report",
            "tolerance_mm": tolerance_mm,
            "element_type": element_type,
            "deviation_summary": {
                "total_deviations": len(deviations),
                "critical": len(critical),
                "major": len(major),
                "minor": len(minor),
                "conformance_percent": round(
                    (1 - len(deviations) / max(len(deviations) + 20, 1)) * 100, 1
                ),
            },
            "deviations": deviations[:50],
            "critical_items": critical,
            "recommendations": (
                ["Halt work on affected areas — critical deviations require structural engineer review"]
                if critical
                else ["Major deviations require rectification before next inspection"] if major
                else ["Minor deviations within acceptable tolerance — document and close"]
            ),
            "sign_off_status": (
                "REJECTED" if critical
                else "CONDITIONAL" if major
                else "APPROVED"
            ),
        }
    def _compare_measurement_sets(
        self, as_built: List[Dict], design: List[Dict], tolerance_mm: float
    ) -> List[Dict]:
        deviations = []
        for ab in as_built:
            ab_val = float(ab.get("value", 0))
            matching = next(
                (d for d in design if d.get("type") == ab.get("type")), None
            )
            if matching:
                design_val = float(matching.get("value", 0))
                diff = abs(ab_val - design_val)
                if diff > tolerance_mm / 1000:
                    deviations.append({
                        "element": ab.get("raw", ab.get("type", "Unknown")),
                        "design_value": f"{design_val}{ab.get('unit', '')}",
                        "as_built_value": f"{ab_val}{ab.get('unit', '')}",
                        "deviation": round(diff, 3),
                        "severity": "major" if diff > tolerance_mm / 500 else "minor",
                        "action_required": "Review and document",
                    })
        return deviations
    async def bim_analysis(self, input_data: Any, params: Dict) -> Dict:
        """Analyse a BIM / IFC model for element counts, quantities, and issues.

        Delegates genuine IFC parsing to the bim_extractor block — no demo mode,
        no fabricated quantities. A missing or bad IFC file returns an error.
        """
        data = input_data if isinstance(input_data, dict) else {}
        p = params or {}

        ifc_file = data.get("ifc_file") or data.get("file_path") or p.get("ifc_file") or p.get("file_path")
        if not ifc_file:
            return {
                "status": "error",
                "action": "bim_analysis",
                "error": "No IFC file provided — pass ifc_file or file_path pointing to an .ifc model",
            }

        block = self._get_bim_extractor_block()
        if block is None:
            return {"status": "error", "action": "bim_analysis", "error": "bim_extractor block unavailable"}

        result = await block.process({"file_path": ifc_file}, p)
        if not isinstance(result, dict) or result.get("status") != "success":
            return {
                "status": "error",
                "action": "bim_analysis",
                "error": (result or {}).get("error", "bim_extractor failed") if isinstance(result, dict) else "bim_extractor failed",
            }

        # Real, block-extracted data — remap into the bim_analysis response shape.
        quantities = result.get("quantities", {})
        element_count = result.get("element_count", 0)
        # Per-category counts straight from the block's quantities tally.
        element_counts = {cat: q.get("count", 0) for cat, q in quantities.items()}
        # extracted_quantities keeps the block's full per-category breakdown.
        extracted_quantities = quantities
        # Disciplines derived from the real categories present, not synthesised.
        disciplines = sorted(quantities.keys())

        # Floor area from real slab quantities where the IFC exposes areas.
        floor_area = 0.0
        for slab in quantities.get("slabs", {}).get("items", []):
            floor_area += slab.get("netarea") or slab.get("grossarea") or 0

        return {
            "status": "success",
            "action": "bim_analysis",
            "file": ifc_file,
            "model_summary": {
                "total_elements": element_count,
                "disciplines": disciplines,
                "ifc_schema": result.get("ifc_schema", ""),
            },
            "project_info": result.get("project_info", {}),
            "storeys": result.get("storeys", []),
            "spaces": result.get("spaces", []),
            "element_counts": element_counts,
            "extracted_quantities": extracted_quantities,
            "estimated_floor_area_m2": round(floor_area, 2),
            "clash_report": result.get("clash_report", {}),
            "recommendations": [
                "Run clash detection to identify coordination issues",
                "Export quantities to BOQ for cost estimation",
                "Verify element count against design intent — model completeness check recommended",
            ],
        }
    def _extract_measurements_advanced(self, text: str, text_dict: Dict) -> List[Dict]:
        measurements = []

        # WxH dimension pattern: "5.5m x 3.2m"
        dimension_pattern = r'\b(\d+(?:\.\d+)?)\s*(?:m|m\.|meter|meters|ft|feet|foot|\')\s*(?:x|by|×)\s*(\d+(?:\.\d+)?)\s*(?:m|m\.|meter|meters|ft|feet|foot|\')'
        for match in re.finditer(dimension_pattern, text, re.IGNORECASE):
            width = float(match.group(1))
            height = float(match.group(2))
            unit = "m" if "m" in match.group(0).lower() else "ft"
            area = width * height
            measurements.append({
                "type": "dimension",
                "value": area,
                "unit": f"{unit}²",
                "width": width,
                "height": height,
                "raw": match.group(0),
                "context": text[max(0, match.start()-50):match.end()+50]
            })

        # Direct area mentions: "2500 m2", "floor area: 2,500 sqm"
        area_pattern = r'\b(\d[\d,]*(?:\.\d+)?)\s*(?:m2|m²|sqm|sq\.?\s*m|square\s+met(?:re|er)s?)\b'
        for match in re.finditer(area_pattern, text, re.IGNORECASE):
            try:
                val = float(match.group(1).replace(',', ''))
                if val > 0:
                    measurements.append({
                        "type": "dimension",
                        "value": val,
                        "unit": "m²",
                        "raw": match.group(0),
                        "context": text[max(0, match.start()-50):match.end()+50]
                    })
            except ValueError:
                logger.warning(
                    "swallowed %s in _extract_measurements_advanced() — continuing",
                    "ValueError", exc_info=True,
                )

        # Direct volume mentions: "450 m3", "concrete: 450 m³"
        volume_pattern = r'\b(\d[\d,]*(?:\.\d+)?)\s*(?:m3|m³|cubic\s+met(?:re|er)s?)\b'
        for match in re.finditer(volume_pattern, text, re.IGNORECASE):
            try:
                val = float(match.group(1).replace(',', ''))
                if val > 0:
                    measurements.append({
                        "type": "volume",
                        "value": val,
                        "unit": "m³",
                        "raw": match.group(0),
                        "context": text[max(0, match.start()-50):match.end()+50]
                    })
            except ValueError:
                logger.warning(
                    "swallowed %s in _extract_measurements_advanced() — continuing",
                    "ValueError", exc_info=True,
                )

        quantity_pattern = r'\b(\d+)\s*(?:no|nos|nr|ea|each)?\.?\s*([A-Z][A-Za-z\s]+)'
        for match in re.finditer(quantity_pattern, text[:2000]):
            qty = int(match.group(1))
            item = match.group(2).strip()[:50]
            if len(item) > 3:
                measurements.append({
                    "type": "count",
                    "value": qty,
                    "unit": "ea",
                    "item": item,
                    "raw": match.group(0)
                })

        return measurements[:50]
    async def _detect_risks_from_drawing(self, result: Dict) -> List[Dict]:
        risks = []

        if not result.get("measurements"):
            risks.append({
                "type": "data_quality",
                "description": "No measurements detected — manual verification required",
                "severity": "medium",
                "mitigation": "Use quantity surveyor to verify BOQ",
            })

        if result.get("confidence", {}).get("overall", 1.0) < 0.7:
            risks.append({
                "type": "confidence",
                "description": "Low extraction confidence — OCR or PDF quality may be poor",
                "severity": "medium",
                "mitigation": "Review all quantities manually against original drawings",
            })

        disciplines = result.get("detected_disciplines", [])
        if len(disciplines) > 3:
            risks.append({
                "type": "coordination",
                "description": f"Multiple disciplines detected ({', '.join(disciplines)}) — coordination drawings required",
                "severity": "low",
                "mitigation": "Conduct BIM coordination review before construction",
            })

        specs = result.get("specifications", [])
        high_grade = [s for s in specs if any(g in s.get("value", "") for g in ["C50", "C60", "S460", "S500"])]
        if high_grade:
            risks.append({
                "type": "specification",
                "description": f"High-strength materials specified ({', '.join(s['value'] for s in high_grade[:3])}) — specialist procurement required",
                "severity": "medium",
                "mitigation": "Verify supplier availability and lead times early",
            })

        quantities = result.get("quantities", {})
        if quantities.get("concrete_volume_m3", 0) > 5000:
            risks.append({
                "type": "procurement",
                "description": "Large concrete volume — ready-mix supply continuity risk",
                "severity": "medium",
                "mitigation": "Secure supply agreement with ready-mix plant before construction start",
            })

        return risks
    async def bim_clash_detection(self, input_data: Any, params: Dict) -> Dict:
        """Detect clashes in BIM / IFC discipline models.

        Delegates to the bim_extractor block, which runs a real intra-model
        clash report per IFC file. No demo mode, no fabricated clashes — a
        missing or bad IFC file returns an error.
        """
        data = input_data if isinstance(input_data, dict) else {}
        p = params or {}

        ifc_file = data.get("ifc_file") or p.get("ifc_file") or data.get("file_path") or p.get("file_path")
        discipline_models = list(p.get("discipline_models") or data.get("discipline_models", []))
        if ifc_file and ifc_file not in discipline_models:
            discipline_models = [ifc_file] + discipline_models

        if not discipline_models:
            return {
                "status": "error",
                "action": "bim_clash_detection",
                "error": "No IFC file provided — pass ifc_file, file_path, or discipline_models pointing to .ifc models",
            }

        block = self._get_bim_extractor_block()
        if block is None:
            return {"status": "error", "action": "bim_clash_detection", "error": "bim_extractor block unavailable"}

        # The block runs an intra-model clash report per file; aggregate across
        # the supplied discipline models. Cross-model clashing is not fabricated.
        block_params = dict(p)
        block_params["run_clash_detection"] = True
        clashes: List[Dict] = []
        total_elements = 0
        models_processed: List[str] = []
        detection_method = "name_duplicate_proxy"
        for model_file in discipline_models:
            result = await block.process({"file_path": model_file}, block_params)
            if not isinstance(result, dict) or result.get("status") != "success":
                return {
                    "status": "error",
                    "action": "bim_clash_detection",
                    "error": (result or {}).get("error", f"bim_extractor failed for {model_file}") if isinstance(result, dict) else "bim_extractor failed",
                }
            models_processed.append(model_file)
            total_elements += result.get("element_count", 0)
            clash_report = result.get("clash_report", {})
            detection_method = clash_report.get("detection_method", detection_method)
            for c in clash_report.get("clashes", []):
                clashes.append(self._normalize_block_clash(c, model_file))

        by_discipline = self._group_clashes_by_discipline(clashes)
        clash_ratio = len(clashes) / total_elements if total_elements else 0

        return {
            "status": "success",
            "action": "clash_detection",
            "model_summary": {
                "files_analyzed": models_processed,
                "total_elements_checked": total_elements,
                "models_clashed": len(models_processed),
            },
            "clash_summary": {
                "total_clashes": len(clashes),
                "warnings": len([c for c in clashes if c["severity"] == "warning"]),
                "clash_ratio_percent": round(clash_ratio * 100, 2),
                "detection_method": detection_method,
            },
            "clashes": clashes[:100] if not p.get("full_report") else clashes,
            "by_discipline": by_discipline,
            "coordination_meeting_agenda": self._generate_coordination_agenda(clashes),
        }
    async def om_manual_generator(self, input_data: Any, params: Dict) -> Dict:
        data = input_data if isinstance(input_data, dict) else {}
        p = params or {}
        equipment_list = data.get("equipment_list") or p.get("equipment_list", [])
        spec_file = data.get("spec_file") or p.get("spec_file")
        as_built_drawings = data.get("drawings") or p.get("drawings", [])
        commissioning_data = data.get("commissioning") or p.get("commissioning", {})
        project_name = p.get("project_name", "Project")
    
        if not equipment_list:
            outline = _om_outline_from_text(
                " ".join(
                    str(x)
                    for x in (
                        p.get("text"),
                        p.get("user_message"),
                        p.get("brief"),
                        data.get("text"),
                        data.get("user_message"),
                        data.get("message"),
                    )
                    if x
                )
            )
            if outline:
                return outline
            # Never fabricate an equipment schedule: a generic TBC list looks
            # like a finished deliverable while every line is invented. Refuse
            # honestly and tell the caller what real inputs unblock it.
            return {
                "status": "error",
                "action": "om_manual_generator",
                "error": (
                    "No equipment data supplied — provide 'equipment_list' "
                    "(tag/description/system_type per item), or run bim_extract "
                    "on the project model to derive one"
                ),
            }
    
        sections = []
        sections.append({
            "section": "A. Project Information",
            "content": {
                "project_name": project_name,
                "completion_date": commissioning_data.get("completion_date", "TBD"),
                "contractor": commissioning_data.get("contractor", "TBD"),
                "consultants": commissioning_data.get("consultants", []),
                "warranty_periods": commissioning_data.get("warranties", {}),
                "emergency_contacts": commissioning_data.get("emergency_contacts", [])
            }
        })
    
        systems = self._group_equipment_by_system(equipment_list)
        sections.append({
            "section": "B. Systems Overview",
            "content": {
                "system_descriptions": [{"name": s["name"], "description": s["description"], "components": len(s["equipment"])} for s in systems],
                "system_interdependencies": self._map_system_dependencies(systems)
            }
        })
    
        equipment_data = []
        for equip in equipment_list:
            equipment_data.append({
                "tag_number": equip.get("tag", "TBD"),
                "description": equip.get("description"),
                "manufacturer": equip.get("manufacturer"),
                "model": equip.get("model"),
                "serial_number": equip.get("serial", "To be field verified"),
                "location": equip.get("location"),
                "installation_date": equip.get("install_date"),
                "warranty_expiry": self._add_years_str(equip.get("install_date"), equip.get("warranty_years", 1)),
                "performance_data": equip.get("performance", {}),
                "rated_capacity": equip.get("capacity"),
                "electrical_requirements": equip.get("electrical", {}),
                "maintenance_schedule": self._generate_equipment_maintenance(equip)
            })
    
        sections.append({"section": "C. Equipment Schedules & Technical Data", "content": equipment_data})
        sections.append({"section": "D. Operating Procedures", "content": {"startup_procedures": self._generate_startup_procedures(systems), "normal_operation": self._generate_normal_operation(systems), "shutdown_procedures": self._generate_shutdown_procedures(systems), "emergency_procedures": self._generate_emergency_procedures(systems), "seasonal_operation": self._generate_seasonal_operation(systems)}})
        sections.append({"section": "E. Preventive Maintenance", "content": {"daily_tasks": self._generate_daily_tasks(equipment_list), "weekly_tasks": self._generate_weekly_tasks(equipment_list), "monthly_tasks": self._generate_monthly_tasks(equipment_list), "quarterly_tasks": self._generate_quarterly_tasks(equipment_list), "annual_tasks": self._generate_annual_tasks(equipment_list), "maintenance_matrix": self._create_maintenance_matrix(equipment_list)}})
        sections.append({"section": "F. Troubleshooting Guide", "content": self._generate_troubleshooting_guide(equipment_list)})
        sections.append({"section": "G. As-Built Documentation", "content": {"drawings_list": [Path(d).name for d in as_built_drawings], "specifications_reference": spec_file if spec_file else "Refer to contract documents", "test_results": commissioning_data.get("test_results", []), "certificates": commissioning_data.get("certificates", [])}})
        sections.append({"section": "H. Warranties & Spare Parts", "content": {"warranty_register": [{"equipment": e.get("description") or e.get("name", "TBD"), "expiry": e.get("warranty_expiry"), "contact": e.get("supplier_contact")} for e in equipment_list], "recommended_spare_parts": self._generate_spare_parts_list(equipment_list), "supplier_contacts": list(set([e.get("supplier_contact") for e in equipment_list if e.get("supplier_contact")]))}})
    
        # Sections E, F and H and the training list are filled by helpers that
        # are registered as roadmap in KNOWN_INCOMPLETE.md: they return nothing
        # until real manufacturer maintenance data is wired in. Say so ON THE
        # MANUAL rather than shipping an empty section that reads as complete,
        # and count only what was actually produced.
        #
        # The old count was `len(daily_tasks) + len(monthly_tasks)`, which with
        # a hollow daily-task generator and a one-string monthly one reported
        # "1 maintenance task generated" for a whole plant. That is a false
        # claim on a client deliverable, not a cosmetic bug.
        def _section(letter: str):
            """Look sections up by their letter, not by list position — a
            future section inserted above E would otherwise silently make
            these gap checks describe the wrong part of the manual."""
            for s in sections:
                if s["section"].startswith(f"{letter}."):
                    return s["content"]
            return {}

        preventive = _section("E")
        _task_keys = ("daily_tasks", "weekly_tasks", "monthly_tasks",
                      "quarterly_tasks", "annual_tasks")
        maintenance_tasks_generated = sum(
            len(preventive.get(k) or []) for k in _task_keys
        )
        data_gaps = []
        if not any(preventive.get(k) for k in ("daily_tasks", "weekly_tasks",
                                               "quarterly_tasks")):
            data_gaps.append(
                "Section E (Preventive Maintenance): daily, weekly and quarterly "
                "task schedules are not generated — they require per-equipment "
                "manufacturer maintenance data, which this manual does not have."
            )
        if not preventive.get("maintenance_matrix"):
            data_gaps.append(
                "Section E: the maintenance responsibility matrix is not generated."
            )
        if not _section("F"):
            data_gaps.append(
                "Section F (Troubleshooting Guide): not generated — requires "
                "manufacturer fault codes and remedies per equipment item."
            )
        if not _section("H").get("recommended_spare_parts"):
            data_gaps.append(
                "Section H: the recommended spare-parts list is not generated — "
                "requires supplier part numbers and consumable intervals."
            )
        training_materials = self._extract_training_needs(equipment_list)
        if not training_materials:
            data_gaps.append(
                "Training materials: not derived — requires the operator "
                "competency requirements for the installed systems."
            )

        manual_metadata = {
            "document_number": f"OM-{project_name.replace(' ', '-')}-{datetime.now(timezone.utc).year}",
            "revision": "00 - First Issue",
            "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "total_pages_estimate": len(equipment_list) * 3 + 50,
            "prepared_by": commissioning_data.get("contractor", "Contractor"),
            "approved_by": "Consultant/Client",
            "distribution": ["Client", "Facilities Management", "Building Operator"]
        }
    
        return {
            # Stays "success" even with gaps, matching the convention its
            # siblings already establish and are tested on: `daily_site_report`
            # returns success + `sections_incomplete`, `tender_bid_analysis`
            # returns success + `analysis_gaps`. The manual WAS generated; the
            # gaps are declared in `data_gaps` below.
            #
            # An earlier revision of this returned "partial" here. That was
            # wrong twice over: 25 call sites in the app gate on
            # `status == "success"`, so it would have turned a working
            # deliverable into an apparent failure, and it invented a status
            # word the rest of the container does not use.
            "status": "success",
            "action": "om_manual_generated",
            "manual_metadata": manual_metadata,
            "sections": sections,
            "data_gaps": data_gaps,
            "summary": {
                "total_equipment": len(equipment_list),
                "systems_covered": len(systems),
                "warranty_items": len(equipment_list),
                "maintenance_tasks_generated": maintenance_tasks_generated,
                "sections_not_populated": len(data_gaps),
                "estimated_manual_pages": manual_metadata["total_pages_estimate"]
            },
            "digital_format": {
                "recommended_software": "PDF with hyperlinks, or CAFM system integration",
                "hyperlink_structure": "Section-based navigation with equipment tags linked to data sheets",
                "update_procedure": "Annual review or upon equipment replacement"
            },
            "training_materials": training_materials,
            "appendices": [
                "Equipment Data Sheets", "Test Reports", "Certificates", "Spare Parts Lists", "Supplier Contacts"
            ]
        }
    async def digital_twin_sync(self, input_data: Any, params: Dict) -> Dict:
        data = input_data if isinstance(input_data, dict) else {}
        p = params or {}
        twin_platform = p.get("platform", "generic")
        sync_mode = p.get("mode", "update")
        project_id = p.get("project_id", "project_001")
        data_payload = data.get("data") or p.get("data", {})
    
        transformed_data = self._transform_for_platform(data_payload, twin_platform)
    
        if sync_mode == "initial_sync":
            operations = self._generate_initial_sync_operations(transformed_data, twin_platform)
        elif sync_mode == "delta_sync":
            operations = self._generate_delta_operations(transformed_data, twin_platform)
        else:
            operations = self._generate_update_operations(transformed_data, twin_platform)
    
        platform_config = self._get_platform_config(twin_platform, project_id)
        quality_report = self._check_twin_data_quality(transformed_data)
        api_payloads = self._generate_api_payloads(operations, twin_platform)
    
        return {
            "status": "success",
            "action": "digital_twin_sync",
            # W2 — the name overpromises: this PREPARES the platform sync payloads
            # (transform + operations + api_payloads + connection/auth info) but
            # does NOT push to any live digital-twin platform. status:"success"
            # means "prepared", not "synced". A real push needs the platform
            # credentials/connection described in connection_strings + auth below.
            "sync_status": "prepared_not_pushed",
            "note": (
                "Sync payloads and platform API operations were PREPARED. No live "
                "push occurred — the platform is not connected. Use the "
                "connection_strings + authentication_required below (or wire a "
                "platform connector) to execute the actual sync."
            ),
            "platform": twin_platform,
            "sync_mode": sync_mode,
            "project_id": project_id,
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "data_summary": {
                "elements_to_sync": len(operations),
                "data_points": sum(len(op.get("properties", [])) for op in operations),
                "geometry_updates": len([op for op in operations if op.get("type") == "geometry"]),
                "property_updates": len([op for op in operations if op.get("type") == "property"]),
                "relationship_updates": len([op for op in operations if op.get("type") == "relationship"])
            },
            "operations": operations[:50] if not p.get("full_details") else operations,
            "platform_configuration": platform_config,
            "api_payloads": api_payloads[:10] if not p.get("include_payloads") else api_payloads,
            "data_quality": quality_report,
            "sync_recommendations": self._generate_sync_recommendations(quality_report, twin_platform),
            "connection_strings": {
                "bim360": f"https://developer.api.autodesk.com/modelderivative/v2/designdata/{project_id}",
                "azure": f"https://{project_id}.api.weu.digitaltwins.azure.net",
                "aveva": f"connect.aveva.com/{project_id}",
                "generic": "Custom API endpoint required"
            }.get(twin_platform, "Platform-specific endpoint required"),
            "authentication_required": {
                "type": "OAuth2" if twin_platform in ["bim360", "azure"] else "API Key",
                "scope": "Digital Twin Read/Write"
            }
        }
    async def intelligent_workflow(self, input_data: Any, params: Dict) -> Dict:
        """Smart orchestrator - auto-detects user intent and chains actions"""
        user_goal = params.get("goal") or params.get("prompt", "process document")
        data = input_data if isinstance(input_data, dict) else {}
        file_path = data.get("file_path") or data.get("url")
    
        chain_steps = self._build_intelligent_chain(user_goal, file_path)
        results = []
        current_data = input_data
    
        for step in chain_steps:
            method = getattr(self, step["action"], None)
            if method:
                result = await method(current_data, step.get("params", {}))
                results.append({
                    "step": step["action"],
                    "status": result.get("status"),
                    "key_findings": self._extract_key_findings(result)
                })
                current_data = {**(current_data if isinstance(current_data, dict) else {}), "previous_result": result}
    
        next_action = self._suggest_next_action(results, user_goal)
    
        return {
            "status": "success",
            "action": "intelligent_workflow",
            "workflow_executed": [s["action"] for s in chain_steps],
            "step_results": results,
            "consolidated_summary": self._consolidate_results(results),
            "next_recommended_action": next_action,
            "user_query": user_goal
        }
    def _build_intelligent_chain(self, user_goal: str, file_path: Optional[str]) -> List[Dict]:
        """Determine which construction methods to call based on user intent"""
        goal = user_goal.lower()
        chain = []
    
        if file_path and file_path.endswith('.pdf'):
            if any(k in goal for k in ["drawing", "plan", "elevation", "section"]):
                chain.append({"action": "process_document", "params": {"doc_type": "drawing"}})
            elif any(k in goal for k in ["spec", "specification", "csi", "masterformat"]):
                chain.append({"action": "process_specification_full", "params": {}})
            elif any(k in goal for k in ["contract", "clause", "terms", "risk"]):
                chain.append({"action": "process_contract", "params": {}})
            else:
                chain.append({"action": "process_document", "params": {}})
    
        if any(k in goal for k in ["qto", "quantity", "takeoff", "boq", "measurement", "material estimate"]):
            chain.append({"action": "extract_quantities", "params": {}})
    
        if any(k in goal for k in ["cost", "price", "budget", "estimate", "value"]):
            chain.append({"action": "estimate_costs", "params": {}})
    
        if any(k in goal for k in ["buy", "purchase", "procure", "supplier", "enquiry", "order", "lead time"]):
            if not any(s["action"] == "extract_quantities" for s in chain):
                chain.append({"action": "extract_quantities", "params": {}})
            chain.append({"action": "procurement_optimizer", "params": {}})
    
        if any(k in goal for k in ["schedule", "programme", "primavera", "delay", "critical path", "progress"]):
            chain.append({"action": "parse_primavera_schedule", "params": {}})
    
        if any(k in goal for k in ["delay analysis", "forensic", "time impact", "extension of time", "eot", "claim"]):
            chain.append({"action": "forensic_delay_analysis", "params": {}})
            chain.append({"action": "claims_builder", "params": {}})
    
        if any(k in goal for k in ["variation", "change order", "vo", "additional work", "omission"]):
            chain.append({"action": "change_order_impact", "params": {}})
            chain.append({"action": "variation_order_manager", "params": {}})
    
        if any(k in goal for k in ["cash flow", "s-curve", "payment", "invoice", "billing"]):
            chain.append({"action": "cash_flow_forecast", "params": {}})
            chain.append({"action": "payment_certificate", "params": {}})
    
        if any(k in goal for k in ["quality", "defect", "inspection", "qc", "honeycomb", "crack"]):
            chain.append({"action": "qa_qc_inspection", "params": {}})
    
        if any(k in goal for k in ["safety", "osha", "hazard", "incident", "audit"]):
            chain.append({"action": "safety_compliance_audit", "params": {}})
    
        if any(k in goal for k in ["tender", "bid", "bid evaluation", "contractor selection", "quote comparison"]):
            chain.append({"action": "tender_bid_analysis", "params": {}})
    
        if any(k in goal for k in ["carbon", "co2", "green", "esg", "sustainability", "leed", "breeam"]):
            chain.append({"action": "carbon_footprint_calculator", "params": {}})
            chain.append({"action": "esg_sustainability_report", "params": {}})
    
        if any(k in goal for k in ["value engineering", "ve", "alternative", "substitution", "saving", "optimization"]):
            chain.append({"action": "value_engineering", "params": {}})
    
        if any(k in goal for k in ["commissioning", "handover", "practical completion", "testing"]):
            chain.append({"action": "commissioning_checklist", "params": {}})
    
        if any(k in goal for k in ["o&m", "operation and maintenance", "manual", "warranty", "maintenance schedule"]):
            chain.append({"action": "om_manual_generator", "params": {}})
            chain.append({"action": "warranty_maintenance_schedule", "params": {}})
    
        if any(k in goal for k in ["as built", "as-built", "deviation", "record drawing"]):
            chain.append({"action": "as_built_deviation_report", "params": {}})
    
        if any(k in goal for k in ["bim", "clash", "coordination", "model"]):
            chain.append({"action": "bim_clash_detection", "params": {}})
    
        if any(k in goal for k in ["digital twin", "sync", "iot", "sensor"]):
            chain.append({"action": "digital_twin_sync", "params": {}})
    
        if any(k in goal for k in ["submittal", "shop drawing", "sample", "mockup", "approval"]):
            chain.append({"action": "submittal_log_generator", "params": {}})
    
        if any(k in goal for k in ["labor", "manpower", "resource", "histogram", "loading"]):
            chain.append({"action": "resource_histogram", "params": {}})
    
        if any(k in goal for k in ["rfi", "request for information", "clarification", "ambiguity"]):
            chain.append({"action": "rfi_generator", "params": {}})
    
        if any(k in goal for k in ["risk", "risk register", "mitigation", "contingency"]):
            chain.append({"action": "risk_register_auto_populate", "params": {}})
    
        if any(k in goal for k in ["daily report", "site diary", "daily log", "progress photo"]):
            chain.append({"action": "daily_site_report", "params": {}})
    
        if not chain:
            chain.append({"action": "process_document", "params": {}})
    
        return chain

    async def cde_post_rfi(self, input_data: Any, params: Dict) -> Dict:
        """Post an RFI draft to the CDE. Chat is not the register; CDE numbers it."""
        from app.core.cde import (
            CdeError,
            CdeNotConfiguredError,
            default_cde_project_id,
            post_rfi_draft,
        )

        data = input_data if isinstance(input_data, dict) else {}
        p = params or {}
        payload = {**data, **p}
        cde_project_id = str(
            payload.get("cde_project_id") or default_cde_project_id() or ""
        ).strip()
        if not cde_project_id:
            return {
                "status": "error",
                "action": "cde_post_rfi",
                "error": (
                    "cde_project_id is required to post an RFI to the CDE. "
                    "The Fork does not allocate a local RFI number."
                ),
            }
        try:
            posted = await post_rfi_draft(cde_project_id, payload)
        except CdeNotConfiguredError as exc:
            return {
                "status": "error",
                "action": "cde_post_rfi",
                "error": str(exc),
                "not_configured": True,
            }
        except CdeError as exc:
            return {"status": "error", "action": "cde_post_rfi", "error": str(exc)}
        return {
            "status": "success",
            "action": "cde_post_rfi",
            "source_of_truth": "cde",
            "post": posted.as_dict(),
            "note": (
                "Posted to the CDE. Any draft label in chat is not the "
                "register number."
            ),
        }

    async def cde_poll_events(self, input_data: Any, params: Dict) -> Dict:
        """Poll CDE mail+register and run the CM overlay on live rows."""
        from app.core.cde import (
            CdeError,
            CdeNotConfiguredError,
            default_cde_project_id,
            process_cde_events,
        )

        data = input_data if isinstance(input_data, dict) else {}
        p = params or {}
        payload = {**data, **p}
        cde_project_id = str(
            payload.get("cde_project_id") or default_cde_project_id() or ""
        ).strip()
        if not cde_project_id:
            return {
                "status": "error",
                "action": "cde_poll_events",
                "error": (
                    "cde_project_id is required to poll the CDE. "
                    "The Fork does not keep a local RFI register."
                ),
            }
        raw_events = payload.get("events")
        if raw_events is not None and not isinstance(raw_events, list):
            return {
                "status": "error",
                "action": "cde_poll_events",
                "error": "events must be a list of CDE-shaped rows with live ids",
            }
        try:
            result = await process_cde_events(
                cde_project_id,
                payloads=raw_events if isinstance(raw_events, list) else None,
                fork_project_id=str(payload.get("project_id") or "").strip() or None,
                ingest_documents=bool(payload.get("ingest_documents")),
                mailbox=str(payload.get("mailbox") or "inbox"),
            )
        except CdeNotConfiguredError as exc:
            return {
                "status": "error",
                "action": "cde_poll_events",
                "error": str(exc),
                "not_configured": True,
            }
        except CdeError as exc:
            return {"status": "error", "action": "cde_poll_events", "error": str(exc)}
        result["action"] = "cde_poll_events"
        return result

    async def _analyse_text_only(self, text: str, doc_type_hint: str = "auto") -> Dict:
        """Classify and extract structured data from raw text without a file."""
        t = text.lower()

        # Detect doc type from content
        if doc_type_hint != "auto":
            doc_type = doc_type_hint
        elif any(k in t for k in ["bill of quantities", "boq", "schedule of rates", "item no", "unit rate"]):
            doc_type = "bom"
        elif any(k in t for k in ["specification", "clause", "section", "csi", "masterformat", "div "]):
            doc_type = "specification"
        elif any(k in t for k in ["contract", "agreement", "clause", "liquidated damages", "retention"]):
            doc_type = "contract"
        elif any(k in t for k in ["programme", "schedule", "activity id", "wbs", "baseline", "primavera"]):
            doc_type = "schedule"
        elif any(k in t for k in ["drawing", "elevation", "section", "plan", "detail", "grid"]):
            doc_type = "drawing"
        else:
            doc_type = "report"

        # Extract quantities from text
        import re
        quantities = {}
        patterns = [
            (r"concrete[^\n]*?(\d[\d,\.]*)\s*m3", "concrete_m3", "m3"),
            (r"rebar[^\n]*?(\d[\d,\.]*)\s*kg", "rebar_kg", "kg"),
            (r"reinforcement[^\n]*?(\d[\d,\.]*)\s*kg", "rebar_kg", "kg"),
            (r"steel[^\n]*?(\d[\d,\.]*)\s*kg", "structural_steel_kg", "kg"),
            (r"curtain wall[^\n]*?(\d[\d,\.]*)\s*m2", "curtain_wall_m2", "m2"),
            (r"glazing[^\n]*?(\d[\d,\.]*)\s*m2", "glazing_m2", "m2"),
            (r"hvac[^\n]*?(\d[\d,\.]*)\s*m2", "hvac_m2", "m2"),
            (r"electrical[^\n]*?(\d[\d,\.]*)\s*m2", "electrical_m2", "m2"),
            (r"blockwork[^\n]*?(\d[\d,\.]*)\s*m2", "blockwork_m2", "m2"),
            (r"formwork[^\n]*?(\d[\d,\.]*)\s*m2", "formwork_m2", "m2"),
            (r"excavat[^\n]*?(\d[\d,\.]*)\s*m3", "excavation_m3", "m3"),
            (r"pil[^\n]*?(\d[\d,\.]*)\s*lm", "piling_lm", "lm"),
            (r"waterproof[^\n]*?(\d[\d,\.]*)\s*m2", "waterproofing_m2", "m2"),
            (r"roofing[^\n]*?(\d[\d,\.]*)\s*m2", "roofing_m2", "m2"),
            (r"tiling[^\n]*?(\d[\d,\.]*)\s*m2", "tiling_m2", "m2"),
            (r"painting[^\n]*?(\d[\d,\.]*)\s*m2", "painting_m2", "m2"),
            (r"plumbing[^\n]*?(\d[\d,\.]*)\s*m2", "plumbing_m2", "m2"),
        ]
        for pattern, key, unit in patterns:
            m = re.search(pattern, t)
            if m:
                val = _safe_float(m.group(1).replace(",", ""))
                if val > 0:
                    quantities[key] = {"quantity": val, "unit": unit}

        # Extract risks from text
        risks = []
        risk_keywords = ["design change", "material delay", "labour shortage", "weather", "cash flow",
                         "subcontractor", "permit", "ground condition", "safety", "covid", "inflation"]
        for rk in risk_keywords:
            if rk in t:
                risks.append({"description": rk.title(), "likelihood": "medium", "impact": "medium"})

        return {
            "status": "success",
            "doc_type": doc_type,
            "quantities": quantities,
            "risks": risks,
            "specifications": [],
            "title": None,
            "project": None,
            "pages": None,
        }
    async def _process_office_document(self, file_path: str, ext: str, extracted_text: str = "") -> Dict:
        """Route .docx / .xlsx through the document_engine and boq_processor blocks.

        The legacy `_process_drawing` path uses fitz/PyMuPDF which only handles
        PDFs and images. This helper produces a doc_result shaped like
        process_document's output (status, doc_type, quantities, risks, ...) so
        auto_pipeline can build panels without special-casing downstream.
        """
        from app.blocks import BLOCK_REGISTRY

        is_xlsx = ext in ("xlsx", "xls")
        is_docx = ext in ("docx", "doc")

        engine_input = {}
        engine_params = {"xlsx_path" if is_xlsx else "docx_path": file_path}

        engine_result = {}
        engine_block = BLOCK_REGISTRY.get("document_engine")
        if engine_block:
            try:
                engine_instance = engine_block()
                engine_result = await engine_instance.execute(engine_input, engine_params)
            except Exception:
                engine_result = {}

        # For BOQ-style spreadsheets, also try boq_processor — it returns
        # priced line items the procurement pipeline can use directly.
        boq_items = []
        boq_summary = {}
        boq_extract_error = ""
        if is_xlsx:
            boq_block = BLOCK_REGISTRY.get("boq_processor")
            if boq_block:
                try:
                    boq_instance = boq_block()
                    boq_result = await boq_instance.execute({"file_path": file_path}, {})
                    if boq_result.get("status") == "success":
                        boq_items = boq_result.get("line_items", []) or []
                        boq_summary = {
                            "item_count": boq_result.get("item_count", 0),
                            "total_cost": boq_result.get("total_cost", 0),
                            "currency": boq_result.get("currency", "USD"),
                            "sections": boq_result.get("sections", []),
                        }
                    else:
                        # Non-success result — capture so the panel can show
                        # a real reason instead of "BOQ loaded" with 0 items.
                        boq_extract_error = (
                            boq_result.get("error")
                            or f"BOQ processor returned status={boq_result.get('status')!r}"
                        )
                        logger.warning(
                            "documents: boq_processor returned non-success for %s: %s",
                            file_path, boq_extract_error,
                        )
                except Exception as exc:
                    logger.exception(
                        "documents: boq_processor.execute raised for %s", file_path,
                    )
                    boq_extract_error = f"BOQ extraction failed: {exc}"

        # Heuristic doc_type: schedule/contract/specification/drawing based on
        # filename and parsed content (consistent with _classify_document).
        name = file_path.lower()
        if any(k in name for k in ("schedule", "primavera", "p6", "_schedule", "l2_schedule", "l3_schedule")):
            doc_type = "schedule"
        elif any(k in name for k in ("contract", "agreement", "rfp", "request for proposal")):
            doc_type = "contract"
        elif any(k in name for k in ("spec", "basis of design", "performance basis")):
            doc_type = "specification"
        elif boq_items:
            doc_type = "bom"
        else:
            doc_type = "specification" if is_docx else "schedule"

        # Build a quantities dict from BOQ line items if we have them
        quantities: Dict[str, Any] = {}
        if boq_items:
            for item in boq_items:
                desc = (item.get("description") or item.get("item") or "").strip()
                qty = item.get("quantity") or 0
                unit = item.get("unit") or "ea"
                if not desc or qty <= 0:
                    continue
                # Use whitelist filter consistent with _calculate_quantities
                key = " ".join(desc.split()).lower().replace(" ", "_")[:40]
                quantities[key] = {"quantity": _safe_float(qty), "unit": unit}

        # Pull risks/requirements from document_engine if present
        risks_raw = engine_result.get("risks", []) if isinstance(engine_result, dict) else []
        risks = []
        for r in risks_raw[:20]:
            if isinstance(r, dict):
                risks.append({
                    "description": r.get("description") or r.get("title") or str(r)[:120],
                    "likelihood": r.get("likelihood", "medium"),
                    "impact": r.get("impact", "medium"),
                })

        equipment_specs = engine_result.get("equipment_specs", []) if isinstance(engine_result, dict) else []
        requirements = engine_result.get("requirements", []) if isinstance(engine_result, dict) else []

        return {
            "status": "success",
            "doc_type": doc_type,
            "quantities": quantities,
            "boq_summary": boq_summary,
            "boq_items": boq_items,
            # Empty string when no error; populated with the failure reason
            # so callers can surface "Failed to read: ..." in the BOQ panel
            # instead of rendering an empty "BOQ loaded" panel.
            "boq_extract_error": boq_extract_error,
            "risks": risks,
            "specifications": [r for r in requirements if isinstance(r, dict)][:50],
            "equipment_specs": equipment_specs,
            "title": None,
            "project": None,
            "pages": None,
            "_engine_result": engine_result,
        }
    async def auto_pipeline(self, input_data: Any, params: Dict) -> Dict:
        """
        Single-call intelligent pipeline.
        1. Runs process_document to understand the file.
        2. Auto-dispatches downstream actions based on what was found.
        3. Returns structured panels ready for UI rendering — no LLM required.
        """
        data = input_data if isinstance(input_data, dict) else {}
        p = params or {}
        file_path = data.get("file_path") or p.get("file_path") or ""
        extracted_text = data.get("extracted_text") or data.get("text") or ""

        if not file_path and not extracted_text:
            return {"status": "error", "error": "Provide file_path or extracted_text"}

        # ── Step 1: domain analysis ──────────────────────────────────────────
        # Detect docx/xlsx up front and route through document_engine, since
        # process_document → _process_drawing uses fitz which only handles PDFs.
        ext = file_path.rsplit(".", 1)[-1].lower() if file_path else ""
        if file_path and ext in ("docx", "doc", "xlsx", "xls"):
            doc_result = await self._process_office_document(file_path, ext, extracted_text)
        elif file_path:
            doc_result = await self.process_document(
                {"file_path": file_path, "extracted_text": extracted_text},
                {"doc_type": p.get("doc_type", "auto"), "file_path": file_path}
            )
        else:
            # Text-only path — classify from content, skip file IO
            doc_type_hint = p.get("doc_type", "auto")
            doc_result = await self._analyse_text_only(extracted_text, doc_type_hint)

        doc_type = doc_result.get("doc_type", "unknown")
        panels = []
        downstream = {}
        next_actions = []
        # Per-panel failures are captured here so the SPA can render a
        # "1 panel failed to populate" notice instead of silently empty
        # sections. Each entry: {"panel": <name>, "error": <message>}.
        pipeline_warnings: List[Dict] = []

        # ── Document info panel (always) ─────────────────────────────────────
        panels.append({
            "type": "document_info",
            "title": "Document",
            "data": {
                "file": file_path.split("/")[-1],
                "doc_type": doc_type,
                "status": doc_result.get("status"),
                "pages": doc_result.get("pages"),
                "title": doc_result.get("title") or doc_result.get("document_title"),
                "project": doc_result.get("project_name") or doc_result.get("project"),
            }
        })

        # ── Step 2: auto-dispatch based on detected content ──────────────────

        # Quantities → cost estimate + procurement
        quantities = (
            doc_result.get("quantities") or
            doc_result.get("extracted_quantities") or
            doc_result.get("bill_of_quantities") or {}
        )

        def _qty_val(q):
            """Normalize quantity value to a number."""
            if isinstance(q, dict):
                return _safe_float(q.get("quantity", 0))
            return _safe_float(q)

        # Only show quantities panel when at least one value is non-zero
        has_quantities = bool(quantities) and any(
            _qty_val(v) > 0 for v in quantities.values()
        )
        if has_quantities:
            panels.append({"type": "quantities", "title": "Quantities", "data": quantities})
        cost_result = {}
        if has_quantities:
            try:
                # Real cost estimate — delegates per-item unit rates to the
                # historical_benchmark block. No fabricated composite $/m² rate.
                cost_result = await self.generate_cost_estimate(
                    {"quantities": quantities},
                    {
                        "quantities": quantities,
                        "location": p.get("location", "US National Average"),
                        "project_type": p.get("project_type", "general_building"),
                    },
                )
                if isinstance(cost_result, dict) and cost_result.get("status") == "success":
                    downstream["cost_estimate"] = cost_result
                    panels.append({
                        "type": "cost_estimate",
                        "title": "Cost Estimate",
                        "data": cost_result.get("summary", {}),
                        "line_items": cost_result.get("line_items", []),
                        "unpriced_items": cost_result.get("unpriced_items", []),
                    })
                else:
                    # Estimate failed — surface the reason honestly, no fake number.
                    downstream["cost_estimate"] = cost_result
                    panels.append({
                        "type": "cost_estimate",
                        "title": "Cost Estimate",
                        "data": {},
                        "line_items": [],
                        "unpriced_items": [],
                        "error": (cost_result or {}).get(
                            "error", "Cost estimate unavailable"
                        ) if isinstance(cost_result, dict) else "Cost estimate unavailable",
                    })
            except Exception as exc:
                logger.exception("auto_pipeline: cost estimate calculation failed")
                pipeline_warnings.append({"panel": "cost_estimate", "error": str(exc)})
                # Surface the failure honestly rather than silently dropping
                # the panel — consistent with the else-branch above.
                panels.append({
                    "type": "cost_estimate",
                    "title": "Cost Estimate",
                    "data": {},
                    "line_items": [],
                    "unpriced_items": [],
                    "error": f"Cost estimate failed: {exc}",
                })
        # Procurement: if we extracted real quantities, derive the procurement
        # list inline so the user sees it without having to click another button.
        # Otherwise just expose the button for manual triggering.
        if has_quantities:
            try:
                proc_result = await self.procurement_list_generator(
                    {"quantities": quantities, "schedule_start": p.get("schedule_start")},
                    {"budget": p.get("budget")}
                )
                items = proc_result.get("procurement_list", []) or []
                if items:
                    downstream["procurement_list"] = proc_result
                    panels.append({
                        "type": "procurement",
                        "title": "Procurement List",
                        "data": {
                            "procurement_list": items,
                            "total_items": proc_result.get("total_items"),
                            "total_procurement_cost": proc_result.get("total_procurement_cost"),
                            "critical_long_lead_items": proc_result.get("critical_long_lead_items"),
                            "action_required": proc_result.get("action_required", []),
                        },
                    })
            except Exception as exc:
                logger.exception("auto_pipeline: procurement_list_generator failed")
                pipeline_warnings.append({"panel": "procurement", "error": str(exc)})
        next_actions.append({
            "action": "procurement_list_generator",
            "label": "Generate Procurement List",
            "reason": "Re-run procurement scheduling with custom budget / start date"
        })

        # Risks → risk register
        risks = doc_result.get("risks") or doc_result.get("identified_risks") or []
        if risks or doc_type in ("contract", "drawing", "specification"):
            try:
                risk_result = await self.risk_register_auto_populate(
                    {"auto_risks": risks, "project_type": p.get("project_type", "general_building")},
                    {"location": p.get("location", "US National Average")}
                )
                downstream["risk_register"] = risk_result
                panels.append({
                    "type": "risks",
                    "title": "Risk Register",
                    "data": risk_result.get("risks", []),
                    "total": risk_result.get("total_risks", 0)
                })
            except Exception as exc:
                logger.exception("auto_pipeline: risk_register_auto_populate failed")
                pipeline_warnings.append({"panel": "risks", "error": str(exc)})

        # Specifications → submittal log
        specs = doc_result.get("specifications") or doc_result.get("spec_sections") or []
        if specs or doc_type == "specification":
            try:
                submittal_result = await self.submittal_log_generator(
                    {"specifications": specs, "file_path": file_path},
                    {}
                )
                downstream["submittal_log"] = submittal_result
                panels.append({
                    "type": "submittals",
                    "title": "Submittal Log",
                    "data": submittal_result.get("submittals", []),
                    "total": submittal_result.get("total_submittals", 0)
                })
            except Exception as exc:
                logger.exception("auto_pipeline: submittal_log_generator failed")
                pipeline_warnings.append({"panel": "submittals", "error": str(exc)})

        # Schedule → progress tracker
        if doc_type == "schedule":
            ext_for_sched = file_path.rsplit(".", 1)[-1].lower() if file_path else ""
            if ext_for_sched == "xer":
                # Primavera P6 — use the dedicated parser
                try:
                    sched_result = await self.parse_primavera_schedule(
                        {"file_path": file_path}, {}
                    )
                    if sched_result.get("status") != "error":
                        downstream["schedule"] = sched_result
                        panels.append({
                            "type": "schedule",
                            "title": "Schedule",
                            "data": sched_result,
                        })
                        next_actions.append({
                            "action": "progress_tracker",
                            "label": "Track Progress",
                            "reason": "Schedule loaded",
                        })
                except Exception as exc:
                    logger.exception("auto_pipeline: parse_primavera_schedule failed")
                    pipeline_warnings.append({"panel": "schedule", "error": str(exc)})
                # Excel schedule — build a summary panel from what document_engine
                # already extracted, plus a quick row scan for date columns.
                eng = doc_result.get("_engine_result") if isinstance(doc_result, dict) else None
                eng = eng if isinstance(eng, dict) else {}
                xlsx_summary = {
                    "format": "xlsx",
                    "file": file_path.split("/")[-1],
                    "schedule_targets": eng.get("schedule_targets", []),
                    "equipment_specs": eng.get("equipment_specs", []),
                    "constraints": eng.get("constraints", [])[:10],
                    "requirements_count": len(eng.get("requirements", [])),
                }
                # Best-effort row scan with openpyxl for milestones / dates
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(file_path, read_only=True, data_only=True)
                    sheet_summaries = []
                    for ws in wb.worksheets[:5]:
                        rows = list(ws.iter_rows(values_only=True, max_row=200))
                        sheet_summaries.append({
                            "name": ws.title,
                            "row_count": ws.max_row,
                            "col_count": ws.max_column,
                            "preview": [list(r)[:8] for r in rows[:5]],
                        })
                    xlsx_summary["sheets"] = sheet_summaries
                except Exception as exc:
                    # Previously silent — turned a corrupt/locked .xlsx into
                    # an empty "Schedule (Excel)" panel with no error signal.
                    # Now surface the failure so the UI can show why.
                    logger.warning(
                        "auto_pipeline: openpyxl load_workbook failed for %s: %s",
                        file_path, exc,
                    )
                    xlsx_summary["xlsx_error"] = f"Failed to read workbook: {exc}"
                downstream["schedule"] = xlsx_summary
                panels.append({
                    "type": "schedule",
                    "title": "Schedule (Excel)",
                    "data": xlsx_summary,
                })
                next_actions.append({
                    "action": "progress_tracker",
                    "label": "Track Progress",
                    "reason": "Excel schedule loaded — inspect sheets",
                })

        # Contract → process contract details
        if doc_type == "contract":
            try:
                contract_result = await self.process_contract(
                    {"file_path": file_path, "extracted_text": extracted_text}, {}
                )
                downstream["contract"] = contract_result
                panels.append({
                    "type": "contract",
                    "title": "Contract Analysis",
                    "data": contract_result
                })
                next_actions.append({
                    "action": "payment_certificate",
                    "label": "Issue Payment Certificate",
                    "reason": "Contract terms identified"
                })
            except Exception as exc:
                logger.exception("auto_pipeline: process_contract failed")
                pipeline_warnings.append({"panel": "contract", "error": str(exc)})

        # ── Chat context: structured text the user can follow up on ──────────
        chat_context_parts = [f"Document: {file_path.split('/')[-1]} (type: {doc_type})"]
        if quantities:
            chat_context_parts.append(f"Quantities found: {list(quantities.keys())[:10]}")
        if risks:
            chat_context_parts.append(f"Risks identified: {len(risks)}")
        if specs:
            chat_context_parts.append(f"Spec sections: {len(specs)}")
        for panel in panels:
            if panel["type"] == "cost_estimate":
                summary = panel.get("data", {})
                if summary.get("total_estimate"):
                    chat_context_parts.append(f"Total cost estimate: ${summary['total_estimate']:,.0f}")
        if extracted_text:
            chat_context_parts.append(f"\nExtracted text (first 3000 chars):\n{extracted_text[:3000]}")

        # Boundary validation: every panel passes through the typed contract
        # so a shape regression surfaces as a typed error_panel rather than
        # rendering as raw JSON in the UI. (See app/core/panels.py)
        from app.core.panels import validate_panel
        validated_panels = [validate_panel(p) for p in panels]

        # ActivityGraph-ish domain_status for project_dashboard.health_check —
        # derived only from real panel outputs (no synthetic BOQ/procurement).
        from app.core.cm_domain_status import panels_to_domain_status
        domain_status = panels_to_domain_status(
            validated_panels, pipeline_warnings=pipeline_warnings
        )

        return {
            "status": "success",
            "action": "auto_pipeline",
            "doc_type": doc_type,
            "panels": validated_panels,
            "domain_status": domain_status,
            "downstream_actions_run": list(downstream.keys()),
            "next_actions": next_actions,
            "pipeline_warnings": pipeline_warnings,
            "chat_context": "\n".join(chat_context_parts),
            "raw_doc_result": doc_result,
        }

    async def wir_form(self, input_data: Any, params: Dict) -> Dict:
        """Draft a Work Inspection Request from operator facts (PRC-405).

        Live M15: a named WIR template plus pour facts used to search-loop
        and return an empty bubble because ``inspection_request`` delegated
        to photo ``qa_qc_inspection``. This action writes the form from the
        user message. It does not invent an issued WIR number and does not
        require an inspection photograph.
        """
        data = input_data if isinstance(input_data, dict) else {}
        p = params or {}
        text = " ".join(
            str(x)
            for x in (
                p.get("text"),
                p.get("user_message"),
                p.get("brief"),
                data.get("text"),
                data.get("user_message"),
                data.get("message"),
                data.get("scope"),
            )
            if x
        )
        if self._wir_form_is_wrong_deliverable(text):
            return {
                "status": "error",
                "error": (
                    "wir_form drafts Work Inspection Requests only. "
                    "This turn is an RFI, RFP, claim, job requisition, "
                    "or other non-WIR deliverable — write that "
                    "deliverable instead of a WIR."
                ),
                "action": "wir_form",
            }
        facts = self._wir_facts_from_text(text)
        for key in (
            "location", "activity", "mix", "volume_m3", "week",
            "manhole_count", "manhole_range", "supplier", "template",
            "wir_number",
        ):
            if p.get(key) not in (None, "") and not facts.get(key):
                facts[key] = p[key]
            elif data.get(key) not in (None, "") and not facts.get(key):
                facts[key] = data[key]

        hold_points = self._wir_hold_points(text, facts)
        witness_points = self._wir_witness_points(text, facts)
        checklist = self._wir_pre_pour_checklist(facts)
        template = facts.get("template") or ""
        wir_number = facts.get("wir_number") or "DRAFT-WIR"
        location = facts.get("location") or "as stated by the operator"
        activity = facts.get("activity") or "Works described in the operator message"
        mix = facts.get("mix") or ""
        volume = facts.get("volume_m3")
        volume_s = f"{volume} m³" if volume not in (None, "") else ""
        week = facts.get("week")
        mh = facts.get("manhole_range") or (
            f"{facts['manhole_count']} manholes" if facts.get("manhole_count") else ""
        )
        supplier = facts.get("supplier") or ""

        scope_bits = [activity]
        if week:
            scope_bits.append(f"Week {week}")
        if mh:
            scope_bits.append(mh)
        if mix:
            scope_bits.append(mix)
        if volume_s:
            scope_bits.append(volume_s)
        if supplier:
            scope_bits.append(f"supplier {supplier}")
        scope = " — ".join(scope_bits)

        from app.core.procedure_actions import procedure_metadata
        meta = procedure_metadata("inspection_request")

        return {
            "status": "success",
            "action": "wir_form",
            "execution_mode": "drafted",
            "procedure_id": "PRC-405",
            "wir_number": wir_number,
            "issued": False,
            "template": template,
            "location": location,
            "activity": activity,
            "mix": mix,
            "volume_m3": volume,
            "week": week,
            "manhole_count": facts.get("manhole_count"),
            "manhole_range": facts.get("manhole_range"),
            "supplier": supplier,
            "scope": scope,
            "hold_points": hold_points,
            "witness_points": witness_points,
            "checklist": checklist,
            "signatories": [
                {"party": "Contractor QC", "status": "pending"},
                {"party": "Engineer / Inspector", "status": "pending"},
            ],
            "notice": (
                "24-hour minimum notice to the Engineer / Inspector before pour. "
                "No concrete shall be placed until this inspection request is signed off."
            ),
            "note": (
                "Draft WIR from operator-supplied facts — not an issued inspection. "
                "Missing fields are left blank rather than invented."
            ),
            "procedure_context": {
                "orchestrator_action": "inspection_request",
                "procedure_id": "PRC-405",
                "execution_mode": "delegated",
                "delegate_action": "wir_form",
                "procedure_title": meta.get("procedure_title") or "",
            },
        }

    _WIR_INTENT_RE = re.compile(
        r"work inspection request|\bwir\b|inspection request|"
        r"wir form|wir template",
        re.IGNORECASE,
    )
    _WIR_WRONG_DELIVERABLE_RE = re.compile(
        r"\b("
        r"follow-on rfi|\brfi(?:[-_]?\d+)?\b|request for information|"
        r"\brfp\b|request for proposal|invitation to tender|"
        r"delay claim|claim notice|eot claim|"
        r"job requisition|prequalification shortlist|"
        r"payment certificate|\bipc\b|interim payment|"
        r"\bwbs\b|work breakdown|"
        r"cash[- ]?flow|"
        r"as-built|as built|"
        r"design directive|"
        r"o\s*&\s*m|om manual|"
        r"commissioning checklist"
        r")\b",
        re.IGNORECASE,
    )

    def _wir_form_is_wrong_deliverable(self, text: str) -> bool:
        """True when the model called wir_form on an RFI / RFP / claim.

        Live M10 / M14 on 817f224: construction-pm and contracts-manager
        picked ``wir_form`` from the allowlist and drafted a WIR instead
        of the requested RFI / RFP. Predispatch already requires WIR
        wording; the tool itself must refuse the same mismatch.
        """
        if not text:
            return False
        wrong = bool(self._WIR_WRONG_DELIVERABLE_RE.search(text))
        if not wrong:
            return False
        # Model-rewritten scope ("blinding pour") used to hide the operator
        # RFP / job-req / claim. If those words are in the joined text, refuse
        # even when the model also said "inspection".
        if re.search(
            r"\b(rfp|request for proposal|job requisition|delay claim|"
            r"claim notice|cash[- ]flow|build a wbs|o\s*&\s*m|"
            r"follow-on rfi|\brfi(?:[-_]?\d+)?\b|request for information|"
            r"punch list|design directive)\b",
            text,
            re.I,
        ):
            return True
        if self._WIR_INTENT_RE.search(text):
            return False
        return True

    def _wir_facts_from_text(self, text: str) -> Dict[str, Any]:
        t = text or ""
        facts: Dict[str, Any] = {}
        week = re.search(r"\bweek\s*[- ]?\s*(\d+)\b", t, re.IGNORECASE)
        if week:
            facts["week"] = week.group(1)
        vol = re.search(r"(\d+(?:\.\d+)?)\s*m(?:³|3)\b", t, re.IGNORECASE)
        if vol:
            facts["volume_m3"] = vol.group(1)
        mix = re.search(
            r"\b(C[-\s]?\d{2}(?:\s*[-/]?\s*(?:SRC|OPC|PPC))?)\b",
            t,
            re.IGNORECASE,
        )
        if mix:
            facts["mix"] = re.sub(r"\s+", " ", mix.group(1)).replace(" ", "").upper()
            facts["mix"] = re.sub(r"C-?", "C-", facts["mix"], count=1)
            if facts["mix"].endswith("SRC") and "-" not in facts["mix"][3:]:
                facts["mix"] = re.sub(r"(C-\d{2})SRC", r"\1 SRC", facts["mix"])
        mh_n = re.search(r"(\d+)\s+manholes?\b", t, re.IGNORECASE)
        if mh_n:
            facts["manhole_count"] = int(mh_n.group(1))
        mh_rng = re.search(
            r"\b(MH[-\s]?\d+(?:[-\s]?\d+)*)\s+(?:to|through|–|-)\s+"
            r"(MH[-\s]?\d+(?:[-\s]?\d+)*)",
            t,
            re.IGNORECASE,
        )
        if mh_rng:
            facts["manhole_range"] = (
                f"{mh_rng.group(1).replace(' ', '')} to "
                f"{mh_rng.group(2).replace(' ', '')}"
            )
        supplier = re.search(r"\bsupplier\s+([A-Za-z0-9][A-Za-z0-9._-]*)", t, re.I)
        if supplier:
            facts["supplier"] = supplier.group(1).rstrip(".,;:")
        loc = re.search(
            r"\b(Boulevard [^,.;]+|Zone [A-Za-z0-9]+|PWPS[-\s]?\d+)\b",
            t,
            re.IGNORECASE,
        )
        if loc:
            facts["location"] = loc.group(1).strip()
        if re.search(r"\b(collar|collars|blinding|pour)\b", t, re.I):
            if re.search(r"\bcollar", t, re.I):
                facts["activity"] = "Stormwater manhole collar concrete pour"
            elif re.search(r"\bblinding\b", t, re.I):
                facts["activity"] = "Blinding concrete pour"
            else:
                facts["activity"] = "Concrete pour"
        for token in re.findall(r"\S+\.(?:docx|doc|xlsx)", t, re.IGNORECASE):
            if "wir" in token.lower():
                facts["template"] = token.strip(".,;:()")
                break
        num = re.search(r"\b((?:WIR|IR)[-_/][A-Za-z0-9][\w./-]*)\b", t, re.IGNORECASE)
        if num and not re.search(r"\.(docx|doc|xlsx)$", num.group(1), re.I):
            facts["wir_number"] = num.group(1)
        return facts

    def _wir_hold_points(self, text: str, facts: Dict[str, Any]) -> List[str]:
        holds = [
            "Formation / setting-out accepted against approved drawings",
            "Formwork / collar dimensions and alignment checked",
            "Reinforcement fixed; cover spacers in place (hold before pour)",
            "Concrete mix design and delivery tickets available for the stated class",
            "24-hour notice given to the Engineer / Inspector; no pour until sign-off",
        ]
        if facts.get("mix"):
            holds[3] = (
                f"Concrete mix design and delivery tickets available for {facts['mix']}"
            )
        return holds

    def _wir_witness_points(self, text: str, facts: Dict[str, Any]) -> List[str]:
        low = (text or "").lower()
        points = []
        if "slump" in low or facts.get("mix"):
            points.append("Slump test at discharge (witness)")
        if "cube" in low or facts.get("mix"):
            points.append("Cube sampling and identification (witness)")
        if "cover" in low or facts.get("mix"):
            points.append("Cover to reinforcement immediately before pour (witness)")
        if not points:
            points = [
                "Hold / witness attendance as required by the ITP for this activity",
            ]
        return points

    def _wir_pre_pour_checklist(self, facts: Dict[str, Any]) -> List[Dict[str, str]]:
        rows = [
            {"item": "Subgrade / formation clean, levels checked", "status": ""},
            {"item": "Lines, levels, and collar / edge formwork", "status": ""},
            {"item": "Reinforcement type, laps, and cover", "status": ""},
            {"item": "Mix class, slump, and cube moulds ready", "status": ""},
            {"item": "Safe access, curing, and weather protection", "status": ""},
        ]
        if facts.get("mix"):
            rows[3]["item"] = (
                f"Mix {facts['mix']}, slump equipment, and cube moulds ready"
            )
        return rows

    def _joined_operator_text(self, input_data: Any, params: Dict) -> str:
        data = input_data if isinstance(input_data, dict) else {}
        p = params or {}
        return " ".join(
            str(x)
            for x in (
                p.get("text"),
                p.get("user_message"),
                data.get("text"),
                data.get("user_message"),
                data.get("message"),
                data.get("scope"),
                input_data if isinstance(input_data, str) else "",
            )
            if x
        )

    async def safety_briefing(self, input_data: Any, params: Dict) -> Dict:
        text = self._joined_operator_text(input_data, params)
        draft = _safety_briefing_from_text(text)
        if draft:
            return draft
        return {
            "status": "error",
            "action": "safety_briefing",
            "error": "No haul-road / diversion / safety-briefing ask in the operator text.",
        }

    async def job_requisition(self, input_data: Any, params: Dict) -> Dict:
        """Draft a PRC-601 job requisition from operator facts.

        Live M6: contracts-manager called wir_form then Groq 413'd. This
        drafts the JR from the stated scope (NOC, poles, signage) without
        inventing commercial terms that were not supplied.
        """
        text = self._joined_operator_text(input_data, params)
        noc = ""
        noc_m = re.search(
            r"(AM Rev Design NOC[^.]{0,80}|NOC[^.]*?expir\w+\s+\d{1,2}\s+\w+\s+\d{4})",
            text,
            re.IGNORECASE,
        )
        if noc_m:
            noc = noc_m.group(1).strip()
        poles = ""
        poles_m = re.search(
            r"(solar\s+6\s+to\s+8\s*m[^.]{0,80}|1x120W[^.]{0,60})",
            text,
            re.IGNORECASE,
        )
        if poles_m:
            poles = poles_m.group(1).strip()
        expiry = ""
        exp_m = re.search(
            r"expir\w+\s+(\d{1,2}\s+\w+\s+\d{4})",
            text,
            re.IGNORECASE,
        )
        if exp_m:
            expiry = exp_m.group(1)
        scope_bits = []
        if re.search(r"street[-\s]?light", text, re.I):
            scope_bits.append("Street-lighting installation")
        if poles:
            scope_bits.append(poles)
        if re.search(r"traffic signage", text, re.I):
            scope_bits.append("Traffic signage")
        if re.search(r"road markings", text, re.I):
            scope_bits.append("Road markings")
        if re.search(r"grand mosque", text, re.I):
            scope_bits.append("Grand Mosque Phase 2 VO items")
        scope = "; ".join(scope_bits) or (
            "Works described in the operator message"
        )
        from app.core.procedure_actions import procedure_metadata
        meta = procedure_metadata("job_requisition")
        return {
            "status": "success",
            "action": "job_requisition",
            "execution_mode": "drafted",
            "procedure_id": "PRC-601",
            "jr_number": "DRAFT-JR",
            "issued": False,
            "title": "Job requisition — street-lighting installation",
            "scope": scope,
            "noc": noc,
            "noc_expiry": expiry,
            "prequalification": [
                "Valid lighting / electrical contractor classification",
                "Evidence of similar street-lighting or solar-pole installations",
                "HSE plan covering live-road and public-interface works",
                "Capacity to supply and install the stated pole type and wattage",
            ],
            "shortlist_rubric": [
                {"criterion": "Technical compliance with NOC and VO items", "weight": "40%"},
                {"criterion": "Programme and mobilisation to the stated expiry", "weight": "20%"},
                {"criterion": "HSE / live-road method", "weight": "20%"},
                {"criterion": "Commercial offer (when priced)", "weight": "20%"},
            ],
            "note": (
                "Draft job requisition from operator-supplied facts — not an "
                "issued purchase order. Missing commercial terms are left blank."
            ),
            "procedure_context": {
                "orchestrator_action": "job_requisition",
                "procedure_id": "PRC-601",
                "execution_mode": "delegated",
                "delegate_action": "job_requisition",
                "procedure_title": meta.get("procedure_title") or "",
            },
        }

    async def rfp_draft(self, input_data: Any, params: Dict) -> Dict:
        """Draft a PRC-602 RFP / invitation from operator facts.

        Live M14: the model called wir_form (refused) then Groq 413'd.
        """
        text = self._joined_operator_text(input_data, params)
        refs = []
        for token in (
            "RFI002",
            "SOPR",
            "UMA Stormwater DDC 20212200076",
            "20212200076",
        ):
            if token.lower() in text.lower():
                refs.append(token)
        grp = bool(re.search(r"\bgrp\b", text, re.I))
        channel = bool(re.search(r"closed concrete channel", text, re.I))
        mh = re.search(r"(\d+)\s*-?\s*manhole", text, re.I)
        manholes = mh.group(1) if mh else ""
        option = []
        if grp:
            option.append("formed GRP radiused section")
        if channel:
            option.append("closed concrete channel")
        option_s = " or ".join(option) or "the alternative stated by the operator"
        scope = (
            f"Remove the {manholes + '-manhole ' if manholes else ''}"
            f"radius cluster and install a {option_s}."
        )
        from app.core.procedure_actions import procedure_metadata
        meta = procedure_metadata("rfp_management")
        invitation = (
            "INVITATION TO TENDER\n\n"
            "You are invited to submit a proposal for a stormwater "
            "manhole-rationalisation subcontract. The works replace the "
            f"existing clustered manholes with {option_s}. "
            "This invitation is a draft from the operator brief; it is not "
            "an issued tender until the Engineer / Employer confirms issue."
        )
        return {
            "status": "success",
            "action": "rfp_draft",
            "execution_mode": "drafted",
            "procedure_id": "PRC-602",
            "rfp_number": "DRAFT-RFP",
            "issued": False,
            "title": "RFP — stormwater manhole-rationalisation subcontract",
            "invitation": invitation,
            "scope_of_works": scope,
            "references": refs,
            "prequalification": [
                "Stormwater / drainage contractor with GRP or formed-channel experience",
                "ITP and method statement for live-road excavation and reinstatement",
                "Capacity to interface with Week 53 collar pours if the alternative is accepted",
            ],
            "evaluation_method": (
                "Technical compliance with the accepted alternative (GRP or "
                "closed channel), programme impact on remaining manhole collars, "
                "HSE, then commercial. Award recommendation follows PRC-603."
            ),
            "key_dates_note": (
                "Key dates are to be confirmed against the approved programme "
                "and the SOPR / UMA DDC references named in this brief. "
                "No dates are invented."
            ),
            "note": (
                "Draft RFP from operator-supplied facts — not an issued "
                "invitation. Write the invitation, scope, prequalification, "
                "evaluation method, and key-date note in full paragraphs."
            ),
            "procedure_context": {
                "orchestrator_action": "rfp_management",
                "procedure_id": "PRC-602",
                "execution_mode": "delegated",
                "delegate_action": "rfp_draft",
                "procedure_title": meta.get("procedure_title") or "",
            },
        }

    async def qa_qc_inspection(self, input_data: Any, params: Dict) -> Dict:
        """Quality control inspection from photos or drawings"""
        data = input_data if isinstance(input_data, dict) else {}
        p = params or {}
        file_path = data.get("file_path") or p.get("file_path")
        inspection_type = p.get("type", "general")
    
        if not file_path:
            return {"status": "error", "error": "No inspection image provided"}
    
        image_block = self.get_dep("image")
    
        defect_prompts = {
            "concrete": "Detect cracks, honeycombing, cold joints, voids, spalling, discoloration",
            "masonry": "Check alignment, mortar joints, plumb, coursing, efflorescence, cracks",
            "steel": "Check welds, rust, alignment, bolt patterns, deformations",
            "finish": "Check paint coverage, drywall seams, flooring alignment, tile lippage",
            "general": "Detect construction defects, cracks, alignment issues, finish problems"
        }
    
        if image_block:
            try:
                analysis = await image_block.execute(
                    {"file_path": file_path},
                    {"prompt": defect_prompts.get(inspection_type, defect_prompts["general"]),
                     "mode": "safety_qaqc"}
                )
                result_body = analysis.get("result", {})
                desc = result_body.get("description", "")
                safety_qaqc = result_body.get("safety_qaqc") or []
            except Exception:
                desc = ""
                safety_qaqc = []
        else:
            desc = ""
            safety_qaqc = []

        defects = self._parse_defects(desc)
        # Compose YOLO-derived defects on top; dedup by description.
        yolo_defects = self._classes_to_defects(safety_qaqc)
        seen_descs = {d["description"] for d in defects}
        for yd in yolo_defects:
            if yd["description"] not in seen_descs:
                defects.append(yd)
                seen_descs.add(yd["description"])
        compliance = self._check_compliance(defects, inspection_type)

        _result = {
            "status": "success",
            "inspection_type": inspection_type,
            "file": Path(file_path).name,
            "defects_found": len(defects),
            "defects": defects,
            "severity_score": self._calculate_severity(defects),
            "compliance_status": compliance["status"],
            "compliance_issues": compliance["issues"],
            "recommendations": self._generate_recommendations(defects, inspection_type),
            "pass_fail": "PASS" if not defects else "CONDITIONAL" if all(d["severity"] == "minor" for d in defects) else "FAIL"
        }
        # W6 — live learning capture (default OFF via FORK_LEARNING_CAPTURE).
        # Best-effort; never affects the inspection result.
        try:
            from app.core.learning_capture import capture_qa_defects
            capture_qa_defects(_result, (params or {}).get("project_id") or "")
        except Exception:  # noqa: BLE001
            logger.warning(
                "swallowed %s in qa_qc_inspection() — continuing",
                "Exception", exc_info=True,
            )
        return _result
    _DEFECT_KEYWORDS = {
        "concrete": [
            ("crack", "Crack visible", "major"),
            ("honeycomb", "Honeycombing / segregation", "major"),
            ("spall", "Spalling / surface loss", "major"),
            ("efflorescence", "Efflorescence (moisture migration)", "minor"),
            ("scaling", "Surface scaling", "minor"),
            ("void", "Void / air pocket", "major"),
            ("rebar exposed", "Exposed reinforcement", "critical"),
            ("rebar corrosion", "Reinforcement corrosion", "critical"),
            ("delaminat", "Delamination", "major"),
            ("cold joint", "Cold joint", "minor"),
        ],
        "steel": [
            ("corrosion", "Corrosion", "major"),
            ("rust", "Surface rust", "minor"),
            ("deformation", "Deformation", "major"),
            ("misalignment", "Misalignment", "major"),
            ("missing bolt", "Missing bolt(s)", "critical"),
            ("loose bolt", "Loose bolt(s)", "major"),
            ("weld defect", "Weld defect", "critical"),
            ("paint failure", "Paint / coating failure", "minor"),
        ],
        "masonry": [
            ("crack", "Cracking", "major"),
            ("mortar loss", "Mortar joint loss", "minor"),
            ("displacement", "Brick / block displacement", "major"),
            ("efflorescence", "Efflorescence", "minor"),
        ],
        "finishes": [
            ("paint peeling", "Paint peeling", "minor"),
            ("tile crack", "Tile cracking", "minor"),
            ("water stain", "Water stain", "minor"),
            ("mould", "Mould / mildew", "major"),
        ],
    }
    async def _compare_photo_to_bim(self, photo_path: str, bim_file: str, location: str) -> Dict:
        """Visual SLAM + BIM comparison"""
        image_block = self.get_dep("image")

        # Track WHETHER detection ran, not just what it returned. Without this
        # an absent image block produced detected=[], which read downstream as
        # "the photograph shows none of the expected elements" — a finding of
        # no progress, manufactured out of a missing dependency. Empty-because-
        # nothing-looked and empty-because-nothing-was-there must not collapse
        # into the same value.
        detection_error = None
        if image_block:
            try:
                photo_analysis = await image_block.execute(
                    {"file_path": photo_path},
                    {"prompt": f"Identify construction elements at {location}: walls, columns, beams, slabs, openings, MEP rough-ins"}
                )
                detected = photo_analysis.get("result", {}).get("objects", [])
            except Exception as exc:
                detected, detection_error = [], f"Image analysis failed: {exc}"
        else:
            detected, detection_error = [], "No image analysis block is configured"

        expected_elements = await self._query_bim_location(bim_file, location)
    
        matched = []
        missing = []
        for expected in expected_elements:
            match = any(self._element_similarity(expected, d) > 0.6 for d in detected)
            if match:
                matched.append(expected)
            else:
                missing.append(expected)
    
        return {
            "location": location,
            "photo": Path(photo_path).name,
            "match_confidence": len(matched) / len(expected_elements) if expected_elements else 0,
            "detection_ran": detection_error is None,
            "detection_error": detection_error,
            "elements_detected": len(detected),
            "elements_expected": len(expected_elements),
            "matched": matched,
            "missing": missing,
            "deviations": self._find_deviations(detected, expected_elements)
        }
    async def _query_bim_location(self, bim_file: str, location: str) -> List[Dict]:
        """Query IFC for elements at a specific location via the bim_extractor.

        Always returns a list (possibly empty) to keep the contract honest
        with the `-> List[Dict]` annotation. The caller `_compare_photo_to_bim`
        iterates this and divides by `len(...)` — a dict return would have
        iterated key names and miscounted match confidence.

        The most recent error message (if any) is stashed on
        `self._last_bim_query_error` so callers that care can surface it
        without polluting the return value.
        """
        self._last_bim_query_error: Optional[str] = None
        if not bim_file:
            self._last_bim_query_error = "No BIM model supplied"
            return []
        block = self._resolve_block("bim_extractor")
        if block is None:
            self._last_bim_query_error = "BIM extractor not configured for spatial queries"
            return []
        try:
            # No `action` param. BIMExtractorBlock.process() does not branch on
            # one — it always performs a full extraction — so passing
            # action="query_location" advertised a capability the block has
            # never had, and the filtering below was silently never applied.
            result = await block.process({"file_path": bim_file}, {})
        except Exception as exc:
            self._last_bim_query_error = f"BIM extractor query failed: {exc}"
            return []
        if not isinstance(result, dict) or result.get("status") != "success":
            self._last_bim_query_error = (
                (result or {}).get("error") if isinstance(result, dict)
                else "BIM extractor returned malformed response"
            )
            return []
        # `building_elements`, not `elements`. The block has never returned an
        # `elements` key, so the old read resolved to None on every successful
        # query and this function returned [] even for a valid IFC model —
        # which made match_confidence 0/0 and every progress figure meaningless.
        elements = result.get("building_elements")
        if not isinstance(elements, list):
            self._last_bim_query_error = (
                "BIM extractor returned no building_elements array"
            )
            return []
        return self._filter_elements_by_location(elements, location)

    def _filter_elements_by_location(self, elements: List[Dict], location: str) -> List[Dict]:
        """Narrow extracted IFC elements to a named location.

        TEXT MATCHING, NOT GEOMETRY. The extractor emits each element as
        `{id, ifc_type, category, name, description, object_type, <psets>}`
        with no storey or space reference on the element itself, so there is
        nothing to do containment against. This matches the location string
        against the element's descriptive fields and its flattened property
        values, which is what a storey named "Level 02" or a pset carrying a
        zone tag actually gives us.

        An unusable location ("", "unknown") returns everything rather than
        nothing: the caller compares a photo against the model, and comparing
        against the whole model is a weaker answer, while comparing against an
        empty set would silently read as "nothing was built".
        """
        if not location or location.strip().lower() in {"unknown", "n/a", "none"}:
            return elements
        needle = location.strip().lower()
        matched = []
        for el in elements:
            if not isinstance(el, dict):
                continue
            haystack = " ".join(
                str(v) for k, v in el.items()
                if k != "id" and isinstance(v, (str, int, float))
            ).lower()
            if needle in haystack:
                matched.append(el)
        # No element names the location. Fall back to the full set and say so,
        # rather than returning [] — an empty expectation set makes every photo
        # score a perfect match against nothing.
        if not matched:
            self._last_bim_query_error = (
                f"No element referenced location {location!r}; compared against "
                f"the full model ({len(elements)} elements) instead"
            )
            return elements
        return matched

    @staticmethod
    def _element_tokens(value: Any) -> set:
        """Comparable word tokens for a BIM element or a detected object.

        Handles the two shapes that reach it: the extractor's element dict and
        whatever the image block returns for a detection (a bare label string,
        or a dict keyed label/name/class/object). IFC type names are
        camel-case with an `Ifc` prefix (`IfcWallStandardCase`), so they are
        split into words before comparison — otherwise a detected "wall" would
        never match `IfcWallStandardCase`.
        """
        import re

        if isinstance(value, dict):
            parts = [str(value.get(k, "")) for k in
                     ("label", "name", "class", "object", "category",
                      "ifc_type", "object_type", "description")]
        else:
            parts = [str(value or "")]
        tokens: set = set()
        for part in parts:
            if not part:
                continue
            part = re.sub(r"^Ifc", "", part)
            part = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", " ", part)
            for tok in re.split(r"[^A-Za-z0-9]+", part.lower()):
                if len(tok) > 2:
                    # Naive de-pluralisation. The extractor's `category` is
                    # plural ("walls") while `ifc_type` and vision labels are
                    # singular ("IfcWall", "wall"), so without this an element
                    # tokenises to {wall, walls} against a detected {wall} and
                    # scores 0.571 — just under the caller's 0.6 threshold.
                    # Every wall in the model was being missed by one hundredth.
                    if len(tok) > 3 and tok.endswith("s") and not tok.endswith("ss"):
                        tok = tok[:-1]
                    tokens.add(tok)
        # Common IFC suffixes carry no discriminating meaning.
        return tokens - {"standard", "case", "element", "type", "proxy"}

    def _element_similarity(self, expected: Any, detected: Any) -> float:
        """How strongly a detected object corresponds to a BIM element, 0..1.

        Token overlap (Jaccard) against fuzzy string ratio, whichever is
        higher. Deliberately a generic text-similarity measure rather than a
        construction-specific heuristic: nothing in the repo measures how well
        a vision label maps to an IFC class, and inventing a domain-weighted
        score would be presenting a guess as a measurement. The caller's
        threshold (0.6) decides what counts as a match.
        """
        from difflib import SequenceMatcher

        exp_tokens = self._element_tokens(expected)
        det_tokens = self._element_tokens(detected)
        if not exp_tokens or not det_tokens:
            return 0.0
        jaccard = len(exp_tokens & det_tokens) / len(exp_tokens | det_tokens)
        ratio = SequenceMatcher(
            None, " ".join(sorted(exp_tokens)), " ".join(sorted(det_tokens))
        ).ratio()
        return round(max(jaccard, ratio), 3)

    def _find_deviations(self, detected: List, expected: List[Dict]) -> List[Dict]:
        """Detected objects that correspond to no element in the model.

        The inverse of `missing`: those are modelled-but-not-seen, these are
        seen-but-not-modelled — the direction that catches unrecorded work and
        as-built departures. Reported as observations with the similarity that
        was actually computed, never as a judgement about whether the
        departure is acceptable.
        """
        deviations = []
        for obj in detected or []:
            best = 0.0
            closest = None
            for element in expected or []:
                score = self._element_similarity(element, obj)
                if score > best:
                    best, closest = score, element
            if best < 0.6:
                deviations.append({
                    "type": "unmodelled_object",
                    "detected": (obj.get("label") or obj.get("name")
                                 if isinstance(obj, dict) else str(obj)),
                    "closest_model_element": (closest or {}).get("name") or
                                             (closest or {}).get("ifc_type"),
                    "similarity": best,
                    "observation": (
                        "Detected in the photograph with no corresponding model "
                        "element. Verify against the current revision."
                    ),
                })
        return deviations
