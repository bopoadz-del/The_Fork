"""Formula-linked cost-BOQ Excel generator.

The platform GENERATES auditable cost workbooks (not static-value dumps),
matching the gold-standard structure:

  Cover        - title, project info, cost summary linked to BOQ_Summary
  BOQ_Detail   - line items with Amount = =Qty*Rate, category subtotals =SUM()
  BOQ_Summary  - category Amount = =BOQ_Detail!G<subtotal> (cross-sheet),
                 % of total, cumulative, construction subtotal, grand total
  Cost_Charts  - Pie + Bar charts linked to the summary

Construction is cost: every derived number is a LIVE Excel formula so the
client can audit Rate x Qty = Amount, exactly like a real priced BOQ.

``evaluate_workbook_total`` re-reads a saved workbook and evaluates the
generated formulas (a minimal engine for the patterns we emit) so callers /
tests can verify the formulas actually compute the right totals.
"""
from __future__ import annotations

import re
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries

# Professional-finance palette (dark blue header, per the style spec).
_HDR_FILL = PatternFill("solid", fgColor="122B49")
_HDR_FONT = Font(bold=True, color="FFFFFF")
_TITLE = Font(bold=True, size=16)
_CAT = Font(bold=True, color="122B49")
_BOLD = Font(bold=True)
_INPUT = Font(color="0000FF")    # blue  = fixed inputs (qty, rate)
_XREF = Font(color="008000")     # green = cross-sheet references
_MONEY = "#,##0.00"


def generate_cost_boq(
    meta: Dict[str, Any],
    categories: List[Dict[str, Any]],
) -> "openpyxl.Workbook":
    """Build a formula-linked cost-BOQ workbook from structured data.

    ``meta``: {title, project, location, currency, date}
    ``categories``: [{"name": str, "items": [{item_no, description, unit, qty, rate}]}]
    """
    ccy = meta.get("currency", "SAR")
    wb = openpyxl.Workbook()

    # ── BOQ_Detail (default sheet) ──────────────────────────────────────────
    det = wb.active
    det.title = "BOQ_Detail"
    det["B2"] = "DETAILED BILL OF QUANTITIES"
    det["B2"].font = _TITLE
    det["B3"] = meta.get("project", "")
    hdr = 5
    for j, h in enumerate(
        ["Item No", "Description", "Unit", "Quantity", f"Rate ({ccy})", f"Amount ({ccy})"],
        start=2,
    ):
        c = det.cell(row=hdr, column=j, value=h)
        c.font, c.fill, c.alignment = _HDR_FONT, _HDR_FILL, Alignment(horizontal="center")

    r = hdr + 1
    subtotal_rows: List[tuple] = []  # (category_name, subtotal_row)
    for cat in categories:
        det.cell(row=r, column=2, value=cat["name"]).font = _CAT
        r += 1
        start = r
        for it in cat.get("items", []):
            det.cell(row=r, column=2, value=it.get("item_no", ""))
            det.cell(row=r, column=3, value=it.get("description", ""))
            det.cell(row=r, column=4, value=it.get("unit", ""))
            qc = det.cell(row=r, column=5, value=it.get("qty", 0)); qc.font = _INPUT
            rc = det.cell(row=r, column=6, value=it.get("rate", 0)); rc.font = _INPUT; rc.number_format = _MONEY
            ac = det.cell(row=r, column=7, value=f"=E{r}*F{r}"); ac.number_format = _MONEY  # LIVE
            r += 1
        end = r - 1
        det.cell(row=r, column=2, value=f"Subtotal - {cat['name']}").font = _BOLD
        st = det.cell(row=r, column=7, value=f"=SUM(G{start}:G{end})")
        st.font, st.number_format = _BOLD, _MONEY
        subtotal_rows.append((cat["name"], r))
        r += 2

    # ── BOQ_Summary ─────────────────────────────────────────────────────────
    summ = wb.create_sheet("BOQ_Summary")
    summ["B2"] = "BOQ SUMMARY"; summ["B2"].font = _TITLE
    sh = 5
    for j, h in enumerate(
        ["Sr. No", "Category", f"Amount ({ccy})", "% of Total", f"Cumulative ({ccy})"], start=2
    ):
        c = summ.cell(row=sh, column=j, value=h); c.font, c.fill = _HDR_FONT, _HDR_FILL
    first = sh + 1
    sr = first
    for i, (name, det_row) in enumerate(subtotal_rows, start=1):
        summ.cell(row=sr, column=2, value=i)
        summ.cell(row=sr, column=3, value=name)
        ac = summ.cell(row=sr, column=4, value=f"=BOQ_Detail!G{det_row}")
        ac.font, ac.number_format = _XREF, _MONEY  # green cross-sheet link
        sr += 1
    last = sr - 1
    total_row = sr
    summ.cell(row=total_row, column=3, value="Construction Cost Subtotal").font = _BOLD
    tc = summ.cell(row=total_row, column=4, value=f"=SUM(D{first}:D{last})")
    tc.font, tc.number_format = _BOLD, _MONEY
    for row in range(first, last + 1):
        pc = summ.cell(row=row, column=5, value=f"=IFERROR(D{row}/D${total_row},0)")
        pc.number_format = "0.0%"
        cum = f"=D{row}" if row == first else f"=F{row - 1}+D{row}"
        summ.cell(row=row, column=6, value=cum).number_format = _MONEY

    # ── Cover (first sheet) ─────────────────────────────────────────────────
    cov = wb.create_sheet("Cover", 0)
    cov["B2"] = meta.get("title", "Bill of Quantities"); cov["B2"].font = Font(bold=True, size=18)
    cov["B4"] = "Project:"; cov["C4"] = meta.get("project", "")
    cov["B5"] = "Location:"; cov["C5"] = meta.get("location", "")
    cov["B6"] = "Currency:"; cov["C6"] = ccy
    cov["B7"] = "Date:"; cov["C7"] = meta.get("date", "")
    cov["B9"] = "Contractor Cost Summary"; cov["B9"].font = _BOLD
    cov["B10"] = "Construction Cost"
    cc = cov["C10"]; cc.value = f"=IFERROR(BOQ_Summary!D{total_row},0)"; cc.font = _XREF; cc.number_format = _MONEY
    cov["B11"] = "TOTAL PROJECT COST"
    tp = cov["C11"]; tp.value = f"=IFERROR(BOQ_Summary!D{total_row},0)"
    tp.font = Font(bold=True, color="008000"); tp.number_format = _MONEY
    cov.sheet_view.showGridLines = False

    # ── Cost_Charts ─────────────────────────────────────────────────────────
    ch = wb.create_sheet("Cost_Charts")
    ch["B2"] = "COST VISUALIZATION"; ch["B2"].font = _TITLE
    ch.cell(row=5, column=2, value="Category"); ch.cell(row=5, column=3, value=f"Amount ({ccy})")
    cr = 6
    for i, (name, _det_row) in enumerate(subtotal_rows):
        ch.cell(row=cr, column=2, value=name)
        ch.cell(row=cr, column=3, value=f"=BOQ_Summary!D{first + i}").number_format = _MONEY
        cr += 1
    last_chart = cr - 1
    data = Reference(ch, min_col=3, min_row=5, max_row=last_chart)
    cats = Reference(ch, min_col=2, min_row=6, max_row=last_chart)
    pie = PieChart(); pie.title = "Cost Breakdown by Category"
    pie.add_data(data, titles_from_data=True); pie.set_categories(cats)
    ch.add_chart(pie, "E5")
    bar = BarChart(); bar.title = "Category Cost Comparison"
    bar.add_data(data, titles_from_data=True); bar.set_categories(cats)
    ch.add_chart(bar, "E22")

    return wb


# ── minimal formula evaluator (validation guard) ────────────────────────────
def evaluate_workbook_total(src) -> float:
    """Re-read a saved cost-BOQ workbook and EVALUATE its generated formulas to
    return the Construction Cost Subtotal. Validates that the live formulas
    actually compute (not just that the strings exist). Handles only the
    patterns this generator emits: =A*B, =A+B, =SUM(range), =Sheet!Cell,
    =IFERROR(x,0), bare cell refs."""
    wb = openpyxl.load_workbook(src, data_only=False)

    def cell(sheet: str, coord: str) -> float:
        v = wb[sheet][coord].value
        if isinstance(v, str) and v.startswith("="):
            return _eval(sheet, v[1:])
        return float(v) if isinstance(v, (int, float)) else 0.0

    def _eval(sheet: str, expr: str) -> float:
        expr = expr.strip()
        m = re.fullmatch(r"IFERROR\((.+),\s*0\)", expr)
        if m:
            try:
                return _eval(sheet, m.group(1))
            except Exception:
                return 0.0
        m = re.fullmatch(r"SUM\(([A-Z]+\d+):([A-Z]+\d+)\)", expr)
        if m:
            c1, r1, c2, r2 = range_boundaries(f"{m.group(1)}:{m.group(2)}")
            return sum(
                cell(sheet, f"{get_column_letter(col)}{row}")
                for col in range(c1, c2 + 1) for row in range(r1, r2 + 1)
            )
        m = re.fullmatch(r"([A-Za-z_]+)!([A-Z]+\$?\d+)", expr)
        if m:
            return cell(m.group(1), m.group(2).replace("$", ""))
        m = re.fullmatch(r"([A-Z]+\$?\d+)\*([A-Z]+\$?\d+)", expr)
        if m:
            return cell(sheet, m.group(1).replace("$", "")) * cell(sheet, m.group(2).replace("$", ""))
        m = re.fullmatch(r"([A-Z]+\$?\d+)\+([A-Z]+\$?\d+)", expr)
        if m:
            return cell(sheet, m.group(1).replace("$", "")) + cell(sheet, m.group(2).replace("$", ""))
        m = re.fullmatch(r"([A-Z]+\$?\d+)", expr)
        if m:
            return cell(sheet, expr.replace("$", ""))
        return 0.0

    summ = wb["BOQ_Summary"]
    for row in summ.iter_rows():
        for c in row:
            if isinstance(c.value, str) and "Construction Cost Subtotal" in c.value:
                return round(cell("BOQ_Summary", f"D{c.row}"))
    return 0.0
