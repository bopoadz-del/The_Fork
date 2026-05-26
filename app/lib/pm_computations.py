"""Reusable project-management computations — Reasoning Engine Plan 1.

Pure functions, no AI, no I/O. Generated code (Plan 4) and the reasoner
(Plan 5) import these instead of re-deriving the algorithms.

CPM math runs in working-day offsets (integers). See the plan header for the
offset conventions.
"""

import logging
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

from app.schemas.cpm import (
    Activity, CPMInput, CPMOutput, CPMResult, Dependency, DependencyType,
    GanttBar, HistogramPeriod, ResourceHistogram,
)

_logger = logging.getLogger(__name__)


class CircularDependencyError(ValueError):
    """Raised when the activity network contains a cycle."""


def topological_order(activities: List[Activity]) -> List[str]:
    """Activity ids in dependency order (Kahn's algorithm).

    Raises ValueError for an unknown predecessor, CircularDependencyError for
    a cycle. Ties broken by id, so the order is deterministic.
    """
    ids = {a.id for a in activities}
    indegree: Dict[str, int] = {a.id: 0 for a in activities}
    successors: Dict[str, List[str]] = {a.id: [] for a in activities}

    for a in activities:
        for dep in a.predecessors:
            if dep.predecessor_id not in ids:
                raise ValueError(
                    f"Activity '{a.id}' references unknown predecessor "
                    f"'{dep.predecessor_id}'"
                )
            indegree[a.id] += 1
            successors[dep.predecessor_id].append(a.id)

    queue = sorted(i for i, d in indegree.items() if d == 0)
    order: List[str] = []
    while queue:
        nid = queue.pop(0)
        order.append(nid)
        for succ in successors[nid]:
            indegree[succ] -= 1
            if indegree[succ] == 0:
                queue.append(succ)
        queue.sort()

    if len(order) != len(activities):
        cycle = sorted(set(indegree) - set(order))
        raise CircularDependencyError(
            f"Circular dependency among: {', '.join(cycle)}"
        )
    return order


def cpm_forward_pass(
    acts: Dict[str, Activity], order: List[str]
) -> Dict[str, Tuple[int, int]]:
    """Compute (ES, EF) working-day offsets. `order` must be topological."""
    es: Dict[str, int] = {}
    ef: Dict[str, int] = {}
    for nid in order:
        a = acts[nid]
        start = 0
        for dep in a.predecessors:
            p_es, p_ef = es[dep.predecessor_id], ef[dep.predecessor_id]
            if dep.type == DependencyType.FS:
                cand = p_ef + dep.lag
            elif dep.type == DependencyType.SS:
                cand = p_es + dep.lag
            elif dep.type == DependencyType.FF:
                cand = p_ef + dep.lag - a.duration
            else:  # SF
                cand = p_es + dep.lag - a.duration
            start = max(start, cand)
        es[nid] = start
        ef[nid] = start + a.duration
    return {nid: (es[nid], ef[nid]) for nid in order}


def _successor_map(acts: Dict[str, Activity]) -> Dict[str, List[Tuple[str, DependencyType, int]]]:
    """Map each activity id to a list of (successor_id, dep_type, lag)."""
    succ: Dict[str, list] = {nid: [] for nid in acts}
    for a in acts.values():
        for dep in a.predecessors:
            succ[dep.predecessor_id].append((a.id, dep.type, dep.lag))
    return succ


def cpm_backward_pass(
    acts: Dict[str, Activity], order: List[str], project_duration: int
) -> Dict[str, Tuple[int, int]]:
    """Compute (LS, LF) working-day offsets. `order` must be topological."""
    succ = _successor_map(acts)
    ls: Dict[str, int] = {}
    lf: Dict[str, int] = {}
    for nid in reversed(order):
        a = acts[nid]
        finish = project_duration
        for (s_id, s_type, lag) in succ[nid]:
            s_ls, s_lf = ls[s_id], lf[s_id]
            if s_type == DependencyType.FS:
                cand = s_ls - lag
            elif s_type == DependencyType.SS:
                cand = s_ls - lag + a.duration
            elif s_type == DependencyType.FF:
                cand = s_lf - lag
            else:  # SF
                cand = s_lf - lag + a.duration
            finish = min(finish, cand)
        lf[nid] = finish
        ls[nid] = finish - a.duration
    return {nid: (ls[nid], lf[nid]) for nid in acts}


def calculate_float(
    acts: Dict[str, Activity], fwd: Dict[str, Tuple[int, int]]
) -> Dict[str, int]:
    """Free float per activity: how far it can slip without delaying any
    successor's early dates. Returns -1 for activities with no successor
    (the caller substitutes total float)."""
    succ = _successor_map(acts)
    ff: Dict[str, int] = {}
    for nid, a in acts.items():
        es_j, ef_j = fwd[nid]
        slacks = []
        for (s_id, s_type, lag) in succ[nid]:
            es_k, ef_k = fwd[s_id]
            if s_type == DependencyType.FS:
                slacks.append(es_k - (ef_j + lag))
            elif s_type == DependencyType.SS:
                slacks.append(es_k - (es_j + lag))
            elif s_type == DependencyType.FF:
                slacks.append(ef_k - (ef_j + lag))
            else:  # SF
                slacks.append(ef_k - (es_j + lag))
        ff[nid] = max(0, min(slacks)) if slacks else -1
    return ff


def compute_cpm(data: CPMInput) -> CPMOutput:
    """Run the full Critical Path Method over an activity network."""
    activities = data.activities
    if not activities:
        return CPMOutput(results=[], project_duration=0, project_finish=None,
                         critical_path=[], critical_percentage=0.0,
                         near_critical=[])

    acts = {a.id: a for a in activities}
    if len(acts) != len(activities):
        raise ValueError("Duplicate activity ids in input")

    order = topological_order(activities)
    fwd = cpm_forward_pass(acts, order)
    project_duration = max(ef for (_es, ef) in fwd.values())
    bwd = cpm_backward_pass(acts, order, project_duration)
    ff = calculate_float(acts, fwd)
    cal, start = data.calendar, data.project_start

    def proj(offset: int):
        return cal.nth_working_day(start, offset) if (start and offset >= 0) else None

    results: List[CPMResult] = []
    for nid in order:
        a = acts[nid]
        es, ef = fwd[nid]
        ls, lf = bwd[nid]
        tf = ls - es
        results.append(CPMResult(
            id=a.id, name=a.name, duration=a.duration,
            early_start_day=es, early_finish_day=ef,
            late_start_day=ls, late_finish_day=lf,
            total_float=tf,
            free_float=tf if ff[nid] < 0 else ff[nid],
            is_critical=(tf <= 0),
            early_start=proj(es), early_finish=proj(ef),
            late_start=proj(ls), late_finish=proj(lf),
        ))

    critical = sorted((r for r in results if r.is_critical),
                      key=lambda r: (r.early_start_day, r.early_finish_day))
    near = [r.id for r in results if 0 < r.total_float <= 5]
    return CPMOutput(
        results=results,
        project_duration=project_duration,
        project_finish=proj(project_duration),
        critical_path=[r.id for r in critical],
        critical_percentage=round(len(critical) / len(results) * 100, 1),
        near_critical=near,
    )


_PERIOD_LENGTH = {"week": 5, "month": 21}
#: Fallback hours-per-day when the P6 CALENDAR table is unavailable. The real
#: value is read from each activity's CALENDAR row via :func:`parse_xer_full`;
#: this constant only survives as the safety net for synthetic networks.
_HOURS_PER_DAY = 8


def resource_histogram(
    results: List[CPMResult],
    activities: List[Activity],
    period_unit: str = "week",
    task_resources: Optional[List[Dict[str, Any]]] = None,
) -> ResourceHistogram:
    """Time-phased manpower. An activity contributes its crew to every period
    its early-date span overlaps (concurrent headcount, not man-days).

    When ``task_resources`` is provided (a list of TASKRSRC-shaped dicts from
    :func:`parse_xer_full`), prefer the real per-task assignments and skip the
    synthetic ``Activity.resources`` distribution. The TASKRSRC path is the
    *only* one that reflects what P6 actually exported; the synthetic path is
    a best-effort substitute for activities authored in code/tests without a
    resource model.
    """
    if task_resources:
        return histogram_from_taskrsrc(
            task_resources, period_unit=period_unit,
            results=results, activities=activities,
        )
    if period_unit not in _PERIOD_LENGTH:
        raise ValueError(f"period_unit must be one of {list(_PERIOD_LENGTH)}, got {period_unit!r}")
    length = _PERIOD_LENGTH[period_unit]
    res_by_id = {a.id: a.resources for a in activities}
    es_ef = {r.id: (r.early_start_day, r.early_finish_day) for r in results}

    if results:
        last_day = max(ef for (_es, ef) in es_ef.values())
        n_periods = max(1, -(-last_day // length))  # ceil division
    else:
        n_periods = 0

    periods: List[HistogramPeriod] = []
    by_trade_totals: Dict[str, float] = {}
    total_manhours = 0.0

    for p in range(n_periods):
        p_start, p_end = p * length, (p + 1) * length
        by_trade: Dict[str, float] = {}
        for rid, (es, ef) in es_ef.items():
            if ef <= p_start or es >= p_end:
                continue  # activity does not overlap this period
            for res in res_by_id.get(rid, []):
                by_trade[res.trade] = by_trade.get(res.trade, 0.0) + res.count
        periods.append(HistogramPeriod(
            index=p, label=f"{period_unit[0].upper()}{p + 1}",
            total=round(sum(by_trade.values()), 2), by_trade=by_trade,
        ))

    for a in activities:
        if a.id not in es_ef:
            raise ValueError(f"Activity '{a.id}' has no CPM result — "
                             "pass results from the same network")
        es, ef = es_ef[a.id]
        span = ef - es
        for res in a.resources:
            by_trade_totals[res.trade] = (
                by_trade_totals.get(res.trade, 0.0) + res.count
            )
            total_manhours += res.count * span * _HOURS_PER_DAY

    peak = max(periods, key=lambda hp: hp.total, default=None)
    return ResourceHistogram(
        period_unit=period_unit,
        periods=periods,
        peak_total=peak.total if peak else 0.0,
        peak_period=peak.label if peak else "",
        by_trade_totals=by_trade_totals,
        total_manhours=round(total_manhours, 2),
    )


def gantt_data(results: List[CPMResult]) -> List[GanttBar]:
    """One Gantt bar per activity, sorted by early start (then early finish)."""
    bars = [
        GanttBar(
            id=r.id, name=r.name,
            start_day=r.early_start_day, end_day=r.early_finish_day,
            is_critical=r.is_critical,
        )
        for r in results
    ]
    bars.sort(key=lambda b: (b.start_day, b.end_day))
    return bars


def compress_schedule(
    data: CPMInput, reductions: Dict[str, int]
) -> Tuple[CPMOutput, int]:
    """Apply working-day duration cuts to named activities and re-run CPM.

    `reductions` maps activity id -> working days to remove (floored at 0
    duration). Returns (revised CPMOutput, days saved vs the baseline).
    Raises ValueError if an id is not in the network.

    The returned delta is non-negative for FS-only networks and is 0 when
    the reductions do not touch the critical path.
    """
    ids = {a.id for a in data.activities}
    unknown = set(reductions) - ids
    if unknown:
        raise ValueError(f"Unknown activity ids: {', '.join(sorted(unknown))}")

    baseline = compute_cpm(data)
    revised_acts = []
    for a in data.activities:
        if a.id in reductions:
            new_dur = max(0, a.duration - reductions[a.id])
            revised_acts.append(a.model_copy(update={"duration": new_dur}))
        else:
            revised_acts.append(a)

    revised = compute_cpm(data.model_copy(update={"activities": revised_acts}))
    delta = baseline.project_duration - revised.project_duration
    return revised, delta


# ── I/O — Reasoning Engine Plan 6 ──────────────────────────────────────────

_XER_PRED_TYPE = {
    "PR_FS": DependencyType.FS, "PR_SS": DependencyType.SS,
    "PR_FF": DependencyType.FF, "PR_SF": DependencyType.SF,
}


def _xer_hours(value) -> float:
    """Parse an .xer hour-count cell; a malformed cell defaults to 0 rather
    than aborting the whole import."""
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _parse_xer_date(value: Any) -> Optional[date]:
    """Best-effort parse of an .xer date cell. Returns None on miss."""
    if not value:
        return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d", "%d-%b-%y %H:%M", "%d-%b-%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _tokenize_xer_tables(text: str) -> Dict[str, Dict[str, list]]:
    """Tokenise an .xer text blob into ``{table: {"fields": [...], "rows": [...]}}``."""
    tables: Dict[str, Dict[str, list]] = {}
    current: str = ""
    fields: List[str] = []
    for line in text.splitlines():
        if not line:
            continue
        cells = line.split("\t")
        tag = cells[0]
        if tag == "%T":
            current = cells[1] if len(cells) > 1 else ""
            tables[current] = {"fields": [], "rows": []}
        elif tag == "%F":
            fields = cells[1:]
            if current in tables:
                tables[current]["fields"] = fields
        elif tag == "%R":
            if current in tables:
                row = dict(zip(tables[current]["fields"], cells[1:]))
                tables[current]["rows"].append(row)
        elif tag == "%E":
            break
    return tables


def parse_xer_full(text: str) -> Dict[str, Any]:
    """Parse a Primavera P6 ``.xer`` text blob into the full result bundle.

    Returns a dict with:

    * ``activities`` — ``List[Activity]`` (same shape as :func:`parse_xer`).
    * ``calendars_parsed`` — ``{clndr_id: {"name": str, "hours_per_day": float}}``
      from the CALENDAR table. Standard US/EU calendars are 8 h/day; 10-h-shift
      sites export 10. Used to convert ``*_hr_cnt`` cells to working days
      *per activity* rather than via a single hardcoded constant.
    * ``task_resources`` — ``List[Dict]`` of TASKRSRC rows shaped as
      ``{task_id, rsrc_id, target_qty, remain_qty, start_date, end_date}``.
      Time-phased loading driven by these reflects what P6 actually planned;
      see :func:`histogram_from_taskrsrc`.

    Lag values still convert at 8 h/day (TASKPRED rows do not carry a calendar
    reference, so the per-activity calendar trick does not apply).
    """
    tables = _tokenize_xer_tables(text)

    # --- CALENDAR: clndr_id -> {name, hours_per_day} --------------------
    calendars: Dict[str, Dict[str, Any]] = {}
    for r in tables.get("CALENDAR", {}).get("rows", []):
        cid = r.get("clndr_id")
        if not cid:
            continue
        hpd = _xer_hours(r.get("day_hr_cnt"))
        calendars[cid] = {
            "name": r.get("clndr_name") or "",
            "hours_per_day": hpd if hpd > 0 else float(_HOURS_PER_DAY),
        }

    task_rows = tables.get("TASK", {}).get("rows", [])
    pred_rows = tables.get("TASKPRED", {}).get("rows", [])
    taskrsrc_rows = tables.get("TASKRSRC", {}).get("rows", [])

    # --- helper: hours-per-day for a given activity ---------------------
    fallback_logged = {"flag": False}

    def _hpd_for(clndr_id: Optional[str]) -> float:
        if clndr_id and clndr_id in calendars:
            return calendars[clndr_id]["hours_per_day"]
        if not fallback_logged["flag"]:
            _logger.warning(
                "parse_xer_full: missing/unknown calendar id; falling back to "
                "%s h/day for affected activities", _HOURS_PER_DAY,
            )
            fallback_logged["flag"] = True
        return float(_HOURS_PER_DAY)

    if not task_rows:
        return {
            "activities": [],
            "calendars_parsed": calendars,
            "task_resources": [],
        }

    # task_id -> task_code (human-readable id used as Activity.id)
    code_by_tid = {r.get("task_id"): r.get("task_code") or r.get("task_id")
                   for r in task_rows}
    clndr_by_tid = {r.get("task_id"): r.get("clndr_id") for r in task_rows}

    preds_by_tid: Dict[str, List[Dependency]] = {}
    for r in pred_rows:
        tid = r.get("task_id")
        pred_tid = r.get("pred_task_id")
        pred_code = code_by_tid.get(pred_tid)
        if not tid or not pred_code:
            continue
        ptype = _XER_PRED_TYPE.get(r.get("pred_type", "PR_FS"),
                                   DependencyType.FS)
        # Lag stays at 8 h/day: TASKPRED has no clndr_id, and lag is a logical
        # offset between activities (potentially on different calendars).
        lag_days = round(_xer_hours(r.get("lag_hr_cnt")) / _HOURS_PER_DAY)
        preds_by_tid.setdefault(tid, []).append(Dependency(
            predecessor_id=pred_code, type=ptype, lag=int(lag_days),
        ))

    activities: List[Activity] = []
    for r in task_rows:
        tid = r.get("task_id")
        hpd = _hpd_for(clndr_by_tid.get(tid))
        dur_days = round(_xer_hours(r.get("target_drtn_hr_cnt")) / hpd)
        activities.append(Activity(
            id=code_by_tid.get(tid) or tid,
            name=r.get("task_name") or "",
            duration=max(0, int(dur_days)),
            predecessors=preds_by_tid.get(tid, []),
        ))

    # --- TASKRSRC: per-task resource assignments ------------------------
    task_resources: List[Dict[str, Any]] = []
    for r in taskrsrc_rows:
        tid = r.get("task_id")
        if not tid:
            continue
        task_resources.append({
            # Expose the human task_code as well so callers can join against
            # CPM output (which keys by Activity.id == task_code).
            "task_id": tid,
            "task_code": code_by_tid.get(tid),
            "rsrc_id": r.get("rsrc_id") or "",
            "target_qty": _xer_hours(r.get("target_qty")),
            "remain_qty": _xer_hours(r.get("remain_qty")),
            "start_date": _parse_xer_date(
                r.get("target_start_date") or r.get("act_start_date")
            ),
            "end_date": _parse_xer_date(
                r.get("target_end_date") or r.get("act_end_date")
            ),
        })

    return {
        "activities": activities,
        "calendars_parsed": calendars,
        "task_resources": task_resources,
    }


def parse_xer(text: str) -> List[Activity]:
    """Parse Primavera P6 ``.xer`` text into Activity objects.

    Backwards-compatible thin wrapper around :func:`parse_xer_full`; existing
    callers that only need activities keep working. Durations are converted
    using the per-activity CALENDAR ``day_hr_cnt`` when available, falling
    back to 8 h/day. Lag from ``lag_hr_cnt`` still uses 8 h/day.
    """
    return parse_xer_full(text)["activities"]


def histogram_from_taskrsrc(
    task_resources: List[Dict[str, Any]],
    period_unit: str = "week",
    results: Optional[List[CPMResult]] = None,
    activities: Optional[List[Activity]] = None,
) -> ResourceHistogram:
    """Time-phase real P6 ``TASKRSRC`` rows into a :class:`ResourceHistogram`.

    Each row contributes ``target_qty`` man-hours spread uniformly across its
    [start_date, end_date] span, bucketed into periods. When a row has no
    dates (P6 sometimes omits them for unstarted activities), the helper
    falls back to the matching CPM ``early_start_day..early_finish_day`` span
    from ``results`` — this is the only reason ``results``/``activities`` are
    optional inputs. ``rsrc_id`` doubles as the trade label since TASKRSRC
    does not name a discipline directly.
    """
    if period_unit not in _PERIOD_LENGTH:
        raise ValueError(
            f"period_unit must be one of {list(_PERIOD_LENGTH)}, "
            f"got {period_unit!r}"
        )
    length_days = _PERIOD_LENGTH[period_unit]

    es_ef = {r.id: (r.early_start_day, r.early_finish_day)
             for r in (results or [])}
    # Cross-walk task_code (Activity.id) <-> task_id so TASKRSRC rows keyed
    # on the numeric task_id can be matched against CPM results keyed on the
    # code.
    code_by_tid: Dict[str, str] = {}
    for tr in task_resources:
        if tr.get("task_id") and tr.get("task_code"):
            code_by_tid[tr["task_id"]] = tr["task_code"]

    # Determine the calendar window (in working days) so we can build periods.
    span_max = 0
    for tr in task_resources:
        code = tr.get("task_code") or code_by_tid.get(tr.get("task_id"))
        if code and code in es_ef:
            span_max = max(span_max, es_ef[code][1])
    if span_max == 0 and es_ef:
        span_max = max(ef for (_es, ef) in es_ef.values())
    n_periods = max(1, -(-span_max // length_days)) if span_max else 0

    periods: List[HistogramPeriod] = []
    by_trade_totals: Dict[str, float] = {}
    total_manhours = 0.0

    # Build per-period buckets in working-day offsets.
    period_loading: List[Dict[str, float]] = [
        {} for _ in range(n_periods)
    ]
    for tr in task_resources:
        code = tr.get("task_code") or code_by_tid.get(tr.get("task_id"))
        qty = float(tr.get("target_qty") or 0.0)
        if qty <= 0:
            continue
        if code in es_ef:
            es, ef = es_ef[code]
        else:
            # No CPM cross-reference — skip (can't place on the timeline).
            continue
        span = max(1, ef - es)
        per_day = qty / span
        rsrc = tr.get("rsrc_id") or "unassigned"
        by_trade_totals[rsrc] = by_trade_totals.get(rsrc, 0.0) + qty
        total_manhours += qty
        for p in range(n_periods):
            p_start, p_end = p * length_days, (p + 1) * length_days
            overlap = max(0, min(ef, p_end) - max(es, p_start))
            if overlap <= 0:
                continue
            period_loading[p][rsrc] = (
                period_loading[p].get(rsrc, 0.0) + per_day * overlap
            )

    for p in range(n_periods):
        by_trade = period_loading[p]
        periods.append(HistogramPeriod(
            index=p, label=f"{period_unit[0].upper()}{p + 1}",
            total=round(sum(by_trade.values()), 2),
            by_trade={k: round(v, 2) for k, v in by_trade.items()},
        ))

    peak = max(periods, key=lambda hp: hp.total, default=None)
    return ResourceHistogram(
        period_unit=period_unit,
        periods=periods,
        peak_total=peak.total if peak else 0.0,
        peak_period=peak.label if peak else "",
        by_trade_totals={k: round(v, 2) for k, v in by_trade_totals.items()},
        total_manhours=round(total_manhours, 2),
    )


def write_schedule_excel(
    output: CPMOutput,
    path: str,
    histogram: "Optional[ResourceHistogram]" = None,
) -> str:
    """Write a CPMOutput to a formatted .xlsx and return the path.

    This is the one genuinely I/O function in the library. Produces a
    'Schedule' sheet (activity table + a Gantt grid) and, when `histogram` is
    given, a 'Manpower' sheet.

    DISPLAY NOTE: CPMResult finish offsets project one working day beyond the
    activity's actual last day (see Plan 1 header). This function is
    user-facing, so finish DATES shown here subtract one working day. The
    *_day integer columns keep the raw offsets.
    """
    from openpyxl import Workbook
    from app.lib.excel_templates import (
        header_row, paint_gantt_row, write_histogram_block,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    total_days = max((r.early_finish_day for r in output.results), default=0)
    gantt_first_col = 9  # day grid starts after the 8 table columns
    header_row(ws, 1, [
        "ID", "Name", "Duration",
        "Early Start", "Early Finish", "Total Float", "Critical",
        "",  # spacer before the day grid
    ] + [f"D{d}" for d in range(total_days)])

    def _finish_date(r):
        # r.early_finish projects EF+1; show the real last working day.
        if r.early_finish is None:
            return ""
        return str(r.early_finish - timedelta(days=1))

    for i, r in enumerate(output.results, start=2):
        ws.cell(row=i, column=1, value=r.id)
        ws.cell(row=i, column=2, value=r.name)
        ws.cell(row=i, column=3, value=r.duration)
        ws.cell(row=i, column=4,
                value=str(r.early_start) if r.early_start else "")
        ws.cell(row=i, column=5, value=_finish_date(r))
        ws.cell(row=i, column=6, value=r.total_float)
        ws.cell(row=i, column=7, value="YES" if r.is_critical else "")
        paint_gantt_row(
            ws, row=i, first_col=gantt_first_col,
            start_day=r.early_start_day, end_day=r.early_finish_day,
            total_days=total_days, is_critical=r.is_critical,
        )

    if histogram is not None:
        hs = wb.create_sheet("Manpower")
        write_histogram_block(
            hs, start_row=1,
            periods=[{"label": p.label, "total": p.total}
                     for p in histogram.periods],
        )

    wb.save(path)
    return path
