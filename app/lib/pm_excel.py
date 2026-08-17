"""Cost-loaded L2 schedule + EVM workbook generators (formula-linked).

Matches the Anthropic / Kenya 200MW / Canada 1GW L2 exemplars:
  L2 Schedule        - CPM (ES/EF/LS/LF/float/critical) + cost per activity
  Cost Loading       - cost baseline / S-curve with cumulative =prev+curr
  Manpower Histogram - man-days = =Dur*Manpower
  Summary            - links to total cost, duration, critical count

EVM workbook (from PV/EV/AC + BAC):
  CV=EV-AC, SV=EV-PV, CPI=EV/AC, SPI=EV/PV, EAC=BAC/CPI, ETC=EAC-AC, VAC=BAC-EAC
  — all LIVE Excel formulas, so the client can audit performance, not pasted.
"""
from __future__ import annotations

from collections import deque
from typing import Any, Dict, List, Tuple

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

_HDR_FILL = PatternFill("solid", fgColor="122B49")
_HDR_FONT = Font(bold=True, color="FFFFFF")
_TITLE = Font(bold=True, size=16)
_BOLD = Font(bold=True)
_INPUT = Font(color="0000FF")   # blue = inputs
_XREF = Font(color="008000")    # green = cross-sheet
_MONEY = "#,##0.00"


def _header(ws, row: int, labels: List[str]) -> None:
    for j, h in enumerate(labels, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font, c.fill, c.alignment = _HDR_FONT, _HDR_FILL, Alignment(horizontal="center")


def _cpm(activities: List[Dict[str, Any]]) -> Tuple[Dict[str, Dict[str, int]], int, List[str]]:
    """Forward/backward CPM pass (finish-to-start). Returns
    ({id: {es,ef,ls,lf,tf,critical}}, project_duration, topo_order)."""
    acts = {str(a["id"]): a for a in activities}
    ids = [str(a["id"]) for a in activities]
    preds = {i: [str(p) for p in acts[i].get("predecessors", [])] for i in ids}
    succ: Dict[str, List[str]] = {i: [] for i in ids}
    for i in ids:
        for p in preds[i]:
            if p in succ:
                succ[p].append(i)
    indeg = {i: sum(1 for p in preds[i] if p in acts) for i in ids}
    q = deque([i for i in ids if indeg[i] == 0])
    order: List[str] = []
    while q:
        n = q.popleft(); order.append(n)
        for s in succ[n]:
            indeg[s] -= 1
            if indeg[s] == 0:
                q.append(s)
    if len(order) != len(ids):          # cycle fallback — keep input order
        order = ids
    es: Dict[str, int] = {}; ef: Dict[str, int] = {}
    for i in order:
        es[i] = max([ef[p] for p in preds[i] if p in ef], default=0)
        ef[i] = es[i] + int(acts[i].get("duration", 0))
    proj = max(ef.values(), default=0)
    ls: Dict[str, int] = {}; lf: Dict[str, int] = {}
    for i in reversed(order):
        lf[i] = min([ls[s] for s in succ[i] if s in ls], default=proj)
        ls[i] = lf[i] - int(acts[i].get("duration", 0))
    out = {i: {"es": es[i], "ef": ef[i], "ls": ls[i], "lf": lf[i],
               "tf": ls[i] - es[i], "critical": (ls[i] - es[i]) <= 0} for i in ids}
    return out, proj, order


def _add_working_days(start, n: int):
    """Calendar date n working days (Mon-Fri) after ``start``. Honest CPM->date
    conversion (no holidays modelled) — used only when a start_date is given."""
    from datetime import timedelta
    d = start
    added = 0
    while added < n:
        d += timedelta(days=1)
        if d.weekday() < 5:
            added += 1
    return d


def generate_cost_loaded_schedule(meta: Dict[str, Any], activities: List[Dict[str, Any]]):
    cpm, proj, order = _cpm(activities)
    acts = {str(a["id"]): a for a in activities}
    ccy = meta.get("currency", "SAR")
    # Cost is only rendered when a real/indicative figure was supplied — a
    # schedule from a brief has no estimate, so we never show a fabricated
    # zero-cost column. Man-days (schedule-derived) carry the S-curve instead.
    has_cost = any(float(a.get("cost", 0) or 0) for a in activities)
    cost_label = meta.get("cost_basis") or "Cost"
    wb = openpyxl.Workbook()

    # ── L2 Schedule ─────────────────────────────────────────────────────────
    s = wb.active; s.title = "L2 Schedule"
    s["A1"] = f"L2 SCHEDULE — {meta.get('project', '')}"; s["A1"].font = _TITLE
    _header(s, 3, ["ID", "WBS", "Activity", "Dur", "Preds", "ES", "EF", "LS", "LF",
                   "Float", "Critical", f"Cost ({ccy})"])
    r = 4
    first = r
    for i in order:
        a, c = acts[i], cpm[i]
        s.cell(r, 1, i); s.cell(r, 2, a.get("wbs", "")); s.cell(r, 3, a.get("name", ""))
        s.cell(r, 4, a.get("duration", 0)); s.cell(r, 5, ",".join(str(p) for p in a.get("predecessors", [])))
        s.cell(r, 6, c["es"]); s.cell(r, 7, c["ef"]); s.cell(r, 8, c["ls"]); s.cell(r, 9, c["lf"])
        s.cell(r, 10, c["tf"]); s.cell(r, 11, "YES" if c["critical"] else "")
        cc = s.cell(r, 12, (a.get("cost", 0) if has_cost else None)); cc.font = _INPUT; cc.number_format = _MONEY
        r += 1
    s.cell(r, 3, "TOTAL").font = _BOLD
    tc = s.cell(r, 12, f"=SUM(L{first}:L{r - 1})"); tc.font = _BOLD; tc.number_format = _MONEY
    total_cost_row = r
    s.cell(r + 2, 3, "Project Duration (days)").font = _BOLD
    s.cell(r + 2, 4, proj).font = _BOLD

    # ── Progress baseline / S-curve (cumulative man-days, time-phased) ───────
    # The honest S-curve for a schedule is cumulative man-days over TIME, not
    # cost: man-days (crew x duration) are schedule-derived, whereas cost from a
    # brief would be fabricated. Cost is cumulated only when `has_cost`.
    cl = wb.create_sheet("Cost Loading")
    cl["A1"] = "PROGRESS BASELINE / S-CURVE (cumulative man-days)"; cl["A1"].font = _TITLE
    period = 7
    n_periods = max(1, (proj + period - 1) // period)
    _header(cl, 3, ["Week", "Days", "Man-days", "Cumulative Man-days", "% Complete"])
    sfirst = 4
    sr = sfirst
    for k in range(n_periods):
        w_start, w_end = k * period, k * period + period
        # man-days this week = sum over active activities of manpower x the
        # count of that activity's ES->EF days falling inside the week.
        md = 0
        for i in order:
            overlap = min(cpm[i]["ef"], w_end) - max(cpm[i]["es"], w_start)
            if overlap > 0:
                md += int(acts[i].get("manpower", 0) or 0) * overlap
        cl.cell(sr, 1, k + 1)
        cl.cell(sr, 2, f"{w_start + 1}-{min(w_end, proj)}")
        cl.cell(sr, 3, md).font = _INPUT
        cum = f"=C{sr}" if sr == sfirst else f"=D{sr - 1}+C{sr}"   # live cumulative
        cl.cell(sr, 4, cum).number_format = "#,##0"
        last_cum = f"$D${sfirst + n_periods - 1}"
        cl.cell(sr, 5, f"=IF({last_cum}=0,0,D{sr}/{last_cum})").number_format = "0%"
        sr += 1
    slast = sr - 1
    cl.cell(sr, 2, "TOTAL MAN-DAYS").font = _BOLD
    cl.cell(sr, 3, f"=SUM(C{sfirst}:C{slast})").font = _BOLD
    # S-curve: cumulative man-days vs week — a real time-phased curve.
    scurve = LineChart(); scurve.title = "Progress S-Curve (cumulative man-days)"
    scurve.y_axis.title = "Man-days"; scurve.x_axis.title = "Week"
    sdata = Reference(cl, min_col=4, min_row=3, max_row=slast)
    scats = Reference(cl, min_col=1, min_row=sfirst, max_row=slast)
    scurve.add_data(sdata, titles_from_data=True); scurve.set_categories(scats)
    cl.add_chart(scurve, "G3")

    # Cost baseline — only when a real/indicative cost was supplied.
    if has_cost:
        crow = sr + 3
        cl.cell(crow, 1, f"COST BASELINE ({cost_label}, {ccy}) — cumulative =prev+curr").font = _BOLD
        chrow = crow + 1
        _header(cl, chrow, ["ID", "Activity", f"{cost_label} ({ccy})", f"Cumulative ({ccy})"])
        cfirst = chrow + 1
        cr = cfirst
        for idx, i in enumerate(order):
            cl.cell(cr, 1, i); cl.cell(cr, 2, acts[i].get("name", ""))
            lk = cl.cell(cr, 3, f"='L2 Schedule'!L{first + idx}"); lk.font = _XREF; lk.number_format = _MONEY
            cum = f"=C{cfirst}" if cr == cfirst else f"=D{cr - 1}+C{cr}"   # live cumulative
            cl.cell(cr, 4, cum).number_format = _MONEY
            cr += 1

    # ── Manpower Histogram (man-days table + time-phased demand + chart) ────
    mp = wb.create_sheet("Manpower Histogram")
    mp["A1"] = "MANPOWER HISTOGRAM"; mp["A1"].font = _TITLE
    # Planned man-hours are shown beside the schedule-derived man-days whenever
    # the activities were planned in man-hours (BOQ-derived programmes are:
    # man-hours = quantity x norm). Both describe the same labour, so a reader
    # can reconcile them in the sheet -- Dur x Manpower x shift should equal the
    # planned man-hours, and a gap means the crew or duration was changed
    # without re-deriving. Column is omitted entirely when nothing carries
    # man-hours, so template-driven schedules look exactly as before.
    _has_mh = any(isinstance(a.get("manhours"), (int, float)) and a.get("manhours")
                  for a in acts.values())
    _hours_per_day = float(meta.get("hours_per_day") or 8)
    cols = ["ID", "Activity", "Dur", "Manpower", "Man-days"]
    if _has_mh:
        cols += ["Planned man-hours", f"Man-hours @ {_hours_per_day:g}h (=Man-days x shift)"]
    _header(mp, 3, cols)
    rr = 4
    for i in order:
        a = acts[i]
        mp.cell(rr, 1, i); mp.cell(rr, 2, a.get("name", ""))
        mp.cell(rr, 3, a.get("duration", 0)).font = _INPUT
        mp.cell(rr, 4, a.get("manpower", 0)).font = _INPUT
        mp.cell(rr, 5, f"=C{rr}*D{rr}").number_format = "#,##0"   # man-days = Dur*Manpower
        if _has_mh:
            mp.cell(rr, 6, a.get("manhours", 0)).number_format = "#,##0"
            mp.cell(rr, 7, f"=E{rr}*{_hours_per_day:g}").number_format = "#,##0"
        rr += 1
    mp.cell(rr, 2, "TOTAL").font = _BOLD
    mp.cell(rr, 5, f"=SUM(E4:E{rr - 1})").font = _BOLD
    if _has_mh:
        mp.cell(rr, 6, f"=SUM(F4:F{rr - 1})").font = _BOLD
        mp.cell(rr, 7, f"=SUM(G4:G{rr - 1})").font = _BOLD

    # Time-phased weekly staffing demand: each activity contributes its
    # manpower across its ES->EF window; sum per week = the histogram.
    period = 7
    n_periods = max(1, (proj + period - 1) // period)
    label_row = rr + 3
    mp.cell(label_row, 1, "TIME-PHASED MANPOWER DEMAND (per week)").font = _BOLD
    hrow = label_row + 1
    _header(mp, hrow, ["Week", "Days", "Manpower"])
    pfirst = hrow + 1
    pr = pfirst
    for k in range(n_periods):
        w_start, w_end = k * period, k * period + period
        demand = sum(
            int(acts[i].get("manpower", 0)) for i in order
            if cpm[i]["es"] < w_end and cpm[i]["ef"] > w_start
        )
        mp.cell(pr, 1, k + 1)
        mp.cell(pr, 2, f"{w_start + 1}-{min(w_end, proj)}")
        mp.cell(pr, 3, demand)
        pr += 1
    plast = pr - 1
    mp.cell(pr + 1, 1, "Peak Manpower").font = _BOLD
    mp.cell(pr + 1, 3, f"=MAX(C{pfirst}:C{plast})").font = _BOLD
    hist = BarChart(); hist.type = "col"; hist.title = "Manpower Histogram (weekly demand)"
    data = Reference(mp, min_col=3, min_row=hrow, max_row=plast)
    cats = Reference(mp, min_col=1, min_row=pfirst, max_row=plast)
    hist.add_data(data, titles_from_data=True); hist.set_categories(cats)
    mp.add_chart(hist, "G3")

    # ── Milestones (phase completions + project completion) ──────────────────
    ms = wb.create_sheet("Milestones")
    ms["A1"] = "MILESTONES"; ms["A1"].font = _TITLE
    start_date = None
    if meta.get("start_date"):
        try:
            from datetime import date as _date
            start_date = _date.fromisoformat(str(meta["start_date"])[:10])
        except (ValueError, TypeError):
            start_date = None
    hdr = ["Milestone", "WBS", "Finish (working day)"]
    if start_date:
        hdr.append("Finish Date (Mon-Fri)")
    hdr.append("Critical")
    _header(ms, 3, hdr)
    # A phase completion is the latest-finishing activity within each WBS group.
    phase_end: Dict[str, str] = {}
    for i in order:
        w = str(acts[i].get("wbs", "")) or "(unassigned)"
        if w not in phase_end or cpm[i]["ef"] > cpm[phase_end[w]]["ef"]:
            phase_end[w] = i
    milestones = [
        (f"Complete: {w} — {acts[i].get('name', '')}", w, cpm[i]["ef"], cpm[i]["critical"])
        for w, i in phase_end.items()
    ]
    milestones.append(("PROJECT COMPLETION", "", proj, True))
    milestones.sort(key=lambda m: m[2])
    mrow = 4
    for label, w, ef, crit in milestones:
        ms.cell(mrow, 1, label); ms.cell(mrow, 2, w); ms.cell(mrow, 3, ef)
        col = 4
        if start_date:
            ms.cell(mrow, 4, _add_working_days(start_date, ef).isoformat()); col = 5
        ms.cell(mrow, col, "YES" if crit else "")
        mrow += 1
    # Target milestones extracted from the source RFP/BOD (document_engine),
    # shown alongside the computed completions so the gap is visible.
    targets = meta.get("target_milestones") or []
    if targets:
        trow = mrow + 2
        ms.cell(trow, 1, "TARGET MILESTONES (from source document)").font = _BOLD
        thdr = trow + 1
        _header(ms, thdr, ["Milestone", "Target Date"])
        tr = thdr + 1
        for t in targets:
            nm = t.get("name") or t.get("context") or "Target"
            dt = t.get("target_date") or t.get("date") or t.get("year") or ""
            ms.cell(tr, 1, str(nm)[:80]); ms.cell(tr, 2, str(dt))
            tr += 1
    ms.sheet_view.showGridLines = False

    # ── Summary ─────────────────────────────────────────────────────────────
    summ = wb.create_sheet("Summary")
    summ["A1"] = f"SCHEDULE SUMMARY — {meta.get('project', '')}"; summ["A1"].font = _TITLE
    summ["B3"] = "Project Duration (days)"; summ["C3"] = f"='L2 Schedule'!D{total_cost_row + 2}"
    summ["B4"] = "Total Man-days"; summ["C4"] = f"='Manpower Histogram'!E{rr}"
    summ["B5"] = "Peak Manpower"; summ["C5"] = f"='Manpower Histogram'!C{plast + 2}"
    summ["B6"] = "Critical Activities"; summ["C6"] = sum(1 for i in order if cpm[i]["critical"])
    summ["B7"] = "Milestones"; summ["C7"] = len(milestones)
    if has_cost:
        summ["B8"] = f"Total {cost_label} ({ccy})"
        tcl = summ["C8"]; tcl.value = f"='L2 Schedule'!L{total_cost_row}"
        tcl.font = _XREF; tcl.number_format = _MONEY
    summ.sheet_view.showGridLines = False
    return wb


def generate_evm_workbook(meta: Dict[str, Any], periods: List[Dict[str, Any]]):
    """EVM workbook: PV/EV/AC inputs per period; CV/SV/CPI/SPI/EAC/ETC/VAC as
    live formulas. BAC drives EAC/VAC."""
    bac = meta.get("bac", 0)
    ccy = meta.get("currency", "SAR")
    wb = openpyxl.Workbook()
    e = wb.active; e.title = "EVM"
    e["A1"] = f"EARNED VALUE MANAGEMENT — {meta.get('project', '')}"; e["A1"].font = _TITLE
    e["A2"] = "BAC (Budget at Completion):"
    bc = e["B2"]; bc.value = bac; bc.font = _INPUT; bc.number_format = _MONEY
    _header(e, 4, ["Period", f"PV ({ccy})", f"EV ({ccy})", f"AC ({ccy})",
                   "CV (EV-AC)", "SV (EV-PV)", "CPI (EV/AC)", "SPI (EV/PV)",
                   f"EAC ({ccy})", f"ETC ({ccy})", f"VAC ({ccy})"])
    r = 5
    for p in periods:
        e.cell(r, 1, p.get("period", ""))
        e.cell(r, 2, p.get("pv", 0)).font = _INPUT
        e.cell(r, 3, p.get("ev", 0)).font = _INPUT
        e.cell(r, 4, p.get("ac", 0)).font = _INPUT
        e.cell(r, 5, f"=C{r}-D{r}").number_format = _MONEY          # CV = EV - AC
        e.cell(r, 6, f"=C{r}-B{r}").number_format = _MONEY          # SV = EV - PV
        e.cell(r, 7, f"=C{r}/D{r}").number_format = "0.000"        # CPI = EV / AC
        e.cell(r, 8, f"=C{r}/B{r}").number_format = "0.000"        # SPI = EV / PV
        e.cell(r, 9, f"=$B$2/G{r}").number_format = _MONEY          # EAC = BAC / CPI
        e.cell(r, 10, f"=I{r}-D{r}").number_format = _MONEY         # ETC = EAC - AC
        e.cell(r, 11, f"=$B$2-I{r}").number_format = _MONEY         # VAC = BAC - EAC
        for col in (2, 3, 4):
            e.cell(r, col).number_format = _MONEY
        r += 1
    e.sheet_view.showGridLines = False
    return wb
