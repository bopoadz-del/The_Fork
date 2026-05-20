"""Reusable Excel formatting helpers — Reasoning Engine Plan 6.

Built on openpyxl. No project logic — only spreadsheet shaping (headers,
Gantt-bar cells, histogram blocks) so write_schedule_excel stays thin.
"""

from typing import Dict, List

from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="8B5CF6")
_CRITICAL_FILL = PatternFill("solid", fgColor="EF4444")  # red
_NORMAL_FILL = PatternFill("solid", fgColor="60A5FA")    # blue


def header_row(ws: Worksheet, row: int, titles: List[str]) -> None:
    """Write a bold, filled header row starting at column A."""
    for col, title in enumerate(titles, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL


def paint_gantt_row(
    ws: Worksheet, row: int, first_col: int,
    start_day: int, end_day: int, total_days: int, is_critical: bool,
) -> None:
    """Fill the cells for one activity's bar across a day grid.

    The grid runs `total_days` columns from `first_col`. Cells for days in
    [start_day, end_day) are filled — red if critical, blue otherwise.
    """
    fill = _CRITICAL_FILL if is_critical else _NORMAL_FILL
    for day in range(total_days):
        if start_day <= day < end_day:
            ws.cell(row=row, column=first_col + day).fill = fill


def write_histogram_block(
    ws: Worksheet, start_row: int, periods: List[Dict],
) -> int:
    """Write a manpower histogram: a header then one row per period
    (label, total). Returns the row index just after the block."""
    header_row(ws, start_row, ["Period", "Manpower"])
    row = start_row + 1
    for p in periods:
        ws.cell(row=row, column=1, value=p.get("label"))
        ws.cell(row=row, column=2, value=p.get("total"))
        row += 1
    return row
