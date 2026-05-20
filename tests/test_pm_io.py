"""Tests for pm_computations I/O — Reasoning Engine Plan 6."""

import pytest

from app.lib.pm_computations import parse_xer


_MINIMAL_XER = "\t".join(["%T", "TASK"]) + "\n" + \
    "\t".join(["%F", "task_id", "task_code", "task_name", "target_drtn_hr_cnt"]) + "\n" + \
    "\t".join(["%R", "1001", "A", "Mobilise", "40"]) + "\n" + \
    "\t".join(["%R", "1002", "B", "Excavate", "80"]) + "\n" + \
    "\t".join(["%T", "TASKPRED"]) + "\n" + \
    "\t".join(["%F", "task_id", "pred_task_id", "pred_type", "lag_hr_cnt"]) + "\n" + \
    "\t".join(["%R", "1002", "1001", "PR_FS", "0"]) + "\n" + \
    "%E\n"


def test_parse_xer_reads_activities():
    acts = parse_xer(_MINIMAL_XER)
    assert {a.id for a in acts} == {"A", "B"}
    mob = next(a for a in acts if a.id == "A")
    assert mob.name == "Mobilise"
    assert mob.duration == 5            # 40 hr / 8 hr-per-day


def test_parse_xer_reads_predecessors():
    acts = parse_xer(_MINIMAL_XER)
    exc = next(a for a in acts if a.id == "B")
    assert len(exc.predecessors) == 1
    assert exc.predecessors[0].predecessor_id == "A"


def test_parse_xer_maps_relationship_types():
    xer = _MINIMAL_XER.replace("PR_FS", "PR_SS")
    exc = next(a for a in parse_xer(xer) if a.id == "B")
    from app.schemas.cpm import DependencyType
    assert exc.predecessors[0].type == DependencyType.SS


def test_parse_xer_empty_input_returns_empty_list():
    assert parse_xer("%E\n") == []


def test_parse_xer_survives_non_numeric_duration():
    # A malformed target_drtn_hr_cnt cell must default to 0, not abort parse.
    xer = _MINIMAL_XER.replace(
        "\t".join(["%R", "1001", "A", "Mobilise", "40"]),
        "\t".join(["%R", "1001", "A", "Mobilise", "N/A"]),
    )
    acts = parse_xer(xer)
    assert {a.id for a in acts} == {"A", "B"}
    mob = next(a for a in acts if a.id == "A")
    assert mob.duration == 0            # malformed cell -> 0


def test_parse_xer_survives_non_numeric_lag():
    # A malformed lag_hr_cnt cell must default to 0, not abort parse.
    xer = _MINIMAL_XER.replace(
        "\t".join(["%R", "1002", "1001", "PR_FS", "0"]),
        "\t".join(["%R", "1002", "1001", "PR_FS", "bad"]),
    )
    acts = parse_xer(xer)
    exc = next(a for a in acts if a.id == "B")
    assert exc.predecessors[0].lag == 0


from openpyxl import Workbook

from app.lib.excel_templates import (
    header_row, paint_gantt_row, write_histogram_block,
)


def test_header_row_writes_bold_titles():
    wb = Workbook(); ws = wb.active
    header_row(ws, 1, ["ID", "Name", "Duration"])
    assert ws["A1"].value == "ID"
    assert ws["C1"].value == "Duration"
    assert ws["A1"].font.bold is True


def test_paint_gantt_row_fills_only_the_bar_span():
    wb = Workbook(); ws = wb.active
    # bar from day 2..5 across a 0..9 day grid, starting at column 3
    paint_gantt_row(ws, row=2, first_col=3, start_day=2, end_day=5,
                    total_days=10, is_critical=False)
    # day 0,1 empty; day 2 (col 5) filled; day 5 (col 8) empty (end exclusive)
    assert ws.cell(row=2, column=3).fill.start_color.rgb in (None, "00000000")
    assert ws.cell(row=2, column=5).fill.start_color.rgb is not None
    assert ws.cell(row=2, column=5).fill.start_color.rgb != "00000000"


def test_write_histogram_block_writes_period_totals():
    wb = Workbook(); ws = wb.active
    next_row = write_histogram_block(
        ws, start_row=1,
        periods=[{"label": "W1", "total": 12}, {"label": "W2", "total": 8}],
    )
    assert ws["A2"].value == "W1" and ws["B2"].value == 12
    assert ws["A3"].value == "W2" and ws["B3"].value == 8
    assert next_row == 4            # row after the block


from openpyxl import load_workbook

from app.lib.pm_computations import compute_cpm, write_schedule_excel
from app.schemas.cpm import Activity, CPMInput, Dependency


def _chain():
    return [
        Activity(id="A", name="Mob", duration=3),
        Activity(id="B", name="Exc", duration=5,
                 predecessors=[Dependency(predecessor_id="A")]),
        Activity(id="C", name="Fnd", duration=2,
                 predecessors=[Dependency(predecessor_id="B")]),
    ]


def test_write_schedule_excel_creates_a_file(tmp_path):
    out = compute_cpm(CPMInput(activities=_chain()))
    path = tmp_path / "schedule.xlsx"
    written = write_schedule_excel(out, str(path))
    assert written == str(path)
    assert path.exists()


def test_write_schedule_excel_lists_every_activity(tmp_path):
    out = compute_cpm(CPMInput(activities=_chain()))
    path = tmp_path / "schedule.xlsx"
    write_schedule_excel(out, str(path))
    wb = load_workbook(path)
    ws = wb["Schedule"]
    ids = [ws.cell(row=r, column=1).value for r in range(2, 5)]
    assert ids == ["A", "B", "C"]


def test_write_schedule_excel_finish_date_subtracts_one_working_day(tmp_path):
    from datetime import date
    # A: duration 3, start Mon 2026-05-18 -> last working day is Wed 2026-05-20.
    out = compute_cpm(CPMInput(activities=[Activity(id="A", name="X", duration=3)],
                               project_start=date(2026, 5, 18)))
    path = tmp_path / "s.xlsx"
    write_schedule_excel(out, str(path))
    wb = load_workbook(path)
    ws = wb["Schedule"]
    # finish-date column shows the actual last working day, not EF+1.
    finish_cells = [ws.cell(row=2, column=c).value for c in range(1, 12)]
    assert "2026-05-20" in [str(v) for v in finish_cells]
