"""A schedule built from a BOQ must contain THAT BOQ's work.

Measured 2026-08-17 on the live service: /export/schedule-from-document was
handed a BOQ of structural steel truss, composite cladding, chilled-water
piping and epoxy flooring, and returned 46 activities of which ZERO mentioned
any of it — a per-project-type template, with the response headers reporting
'procurement items: 0, milestones: 0'. The BOQ contributed nothing.

These fences pin the replacement: quantities drive durations, the BOQ's own
scope becomes the activity list, sequencing is construction order, and a
missing productivity rate is refused by name rather than defaulted (a schedule
resting on an invented output rate looks exactly like a real one until it
slips).
"""
from __future__ import annotations

import json
import math

import pytest

from app.lib.boq_schedule import (
    CONSTRUCTION_SEQUENCE,
    MissingProductivity,
    activities_from_boq,
    duration_days,
    group_by_category,
    schedule_basis,
)

# A real-shaped BOQ: exactly what boq_processor emits.
BOQ = [
    {"item_key": "structural_steel_roof_truss", "description": "Structural steel roof truss fabrication",
     "quantity": 47, "unit": "ton", "unit_cost": 4850.0, "total_cost": 227950.0},
    {"item_key": "aluminium_composite_cladding", "description": "Aluminium composite cladding to facade",
     "quantity": 1310, "unit": "m2", "unit_cost": 395.0, "total_cost": 517450.0},
    {"item_key": "chilled_water_piping", "description": "Chilled water piping DN200 insulated",
     "quantity": 640, "unit": "m", "unit_cost": 720.0, "total_cost": 460800.0},
    {"item_key": "bulk_excavation", "description": "Bulk excavation to reduced level",
     "quantity": 2500, "unit": "m3", "unit_cost": 18.5, "total_cost": 46250.0},
]

# The shared classifier is pricing-tuned and puts "…piping DN200 insulated"
# in Waterproofing/Insulation (earliest match wins, on "insulated"). A planner
# corrects that with an override; the rates below cover the corrected result.
# Keyed by DESCRIPTION, not item_key: boq_processor derives item_key from
# the description itself ("chilled_water_piping_dn200_insulated"), so a
# hand-written key silently fails to match the extracted line.
OVERRIDES = {"Chilled water piping DN200 insulated": "Mechanical/HVAC (MEP)"}

# MAN-HOURS per unit (operator: "everything is in manhour"). Crew of 4 on an
# 8 h shift is the default, so e.g. facade 0.53 mh/m2 -> 1310 x 0.53 / 32 = 22 d.
RATES = {
    "Structural Steel": 8.0,          # mh / ton
    "Windows/Doors/Facade": 0.53,     # mh / m2
    "Mechanical/HVAC (MEP)": 0.8,     # mh / m
    "Earthworks/Excavation": 0.08,    # mh / m3
}


def _by_name(acts):
    return {a["name"]: a for a in acts}


# ── durations are derived, not invented ────────────────────────────────────

def test_manhours_are_quantity_times_norm():
    from app.lib.boq_schedule import manhours
    assert manhours(1250, 0.8) == 1000.0
    assert manhours(420, 1.37) == pytest.approx(575.4)


def test_duration_is_manhours_over_crew_and_shift():
    # 1000 mh, crew of 4, 8 h shift -> 31.25 -> 32 working days
    assert duration_days(1000, 4, 8) == 32
    assert duration_days(576, 6, 8) == 12


def test_a_bigger_gang_shortens_the_same_manhours():
    assert duration_days(1000, 8, 8) == 16
    assert duration_days(1000, 4, 8) == 32


def test_shift_length_is_a_variable_not_a_constant():
    """10 h shifts and Ramadan 6 h both happen; the norm does not change."""
    assert duration_days(1000, 4, 10) == 25
    assert duration_days(1000, 4, 6) == 42


def test_a_measurable_item_never_takes_zero_days():
    assert duration_days(1.0, 4, 8) == 1


def test_zero_crew_or_shift_is_an_error_not_a_division():
    with pytest.raises(ValueError):
        duration_days(100, 0, 8)
    with pytest.raises(ValueError):
        duration_days(100, 4, 0)
    from app.lib.boq_schedule import manhours as _mh
    with pytest.raises(ValueError):
        _mh(100, 0)


def test_published_daily_output_converts_to_a_manhour_norm():
    """Handbooks publish units/man-day; planning needs mh/unit."""
    from app.lib.boq_schedule import from_daily_output
    assert from_daily_output(10) == 0.8        # 10 m2/man-day -> 0.8 mh/m2
    assert from_daily_output(18) == pytest.approx(0.444, abs=0.001)


# ── the BOQ's own scope becomes the activity list ──────────────────────────

def test_activities_carry_the_boq_scope_not_a_template():
    acts = activities_from_boq(BOQ, manhours_per_unit=RATES, category_overrides=OVERRIDES)
    names = " | ".join(a["name"].lower() for a in acts)
    for term in ("steel roof truss", "cladding", "chilled water piping", "excavation"):
        assert term in names, f"{term!r} missing from {names}"
    # and none of the canned template's activities appear
    for template_item in ("topographic survey", "building permit"):
        assert template_item not in names


def test_each_activity_traces_back_to_its_boq_line():
    acts = _by_name(activities_from_boq(BOQ, manhours_per_unit=RATES, category_overrides=OVERRIDES))
    cladding = acts["Aluminium composite cladding to facade"]
    assert cladding["boq"]["quantity"] == 1310
    assert cladding["boq"]["unit"] == "m2"
    assert cladding["boq"]["manhours_per_unit"] == 0.53
    assert cladding["total_manhours"] == pytest.approx(694.3, abs=0.1)
    # 694.3 mh / (4 heads x 8 h) = 21.7 -> 22 days
    assert cladding["duration_days"] == 22
    assert cladding["boq"]["total_cost"] == 517450.0


def test_activity_shape_matches_generate_wbs():
    """The CPM engine, cost bridge and Excel writer consume this unchanged."""
    for a in activities_from_boq(BOQ, manhours_per_unit=RATES, category_overrides=OVERRIDES):
        assert {"id", "code", "name", "duration_days", "predecessors",
                "resources", "wbs_phase"} <= set(a)
        assert isinstance(a["duration_days"], int) and a["duration_days"] >= 1
        assert isinstance(a["predecessors"], list)
        assert a["resources"] and all(isinstance(r, str) for r in a["resources"])


# ── sequencing is construction order ───────────────────────────────────────

def test_packages_run_in_construction_order():
    acts = activities_from_boq(BOQ, manhours_per_unit=RATES, category_overrides=OVERRIDES)
    phases = [a["wbs_phase"] for a in acts]
    order = {p: i for i, p in enumerate(dict.fromkeys(phases))}
    # earthworks precedes steel precedes facade precedes MEP
    assert order["earthworks_excavation"] < order["structural_steel"]
    assert order["structural_steel"] < order["windows_doors_facade"]
    assert order["windows_doors_facade"] < order["mechanical_hvac_mep"]


def test_first_activity_has_no_predecessor_and_the_rest_chain():
    acts = activities_from_boq(BOQ, manhours_per_unit=RATES, category_overrides=OVERRIDES)
    assert acts[0]["predecessors"] == []
    ids = {a["id"] for a in acts}
    for a in acts[1:]:
        assert a["predecessors"], f"{a['name']} floats with no predecessor"
        for p in a["predecessors"]:
            assert p in ids, f"dangling predecessor {p}"


def test_sequence_covers_every_category_the_classifier_can_emit():
    from app.lib.boq_pricing import CATS
    known = {c[0] for c in CATS} | {"Other/Uncategorized"}
    assert known <= set(CONSTRUCTION_SEQUENCE), known - set(CONSTRUCTION_SEQUENCE)


def test_an_unrecognised_line_is_still_scheduled_and_reported():
    """Unknown work must never vanish from the programme — and the planner
    must be told the taxonomy did not understand it."""
    from app.lib.boq_schedule import uncategorized
    items = BOQ + [{"item_key": "novel", "description": "Zzz unknown novel work",
                    "quantity": 10, "unit": "no"}]
    assert "Zzz unknown novel work" in uncategorized(items, OVERRIDES)
    rates = dict(RATES); rates["Other/Uncategorized"] = 1.0
    acts = activities_from_boq(items, manhours_per_unit=rates, category_overrides=OVERRIDES)
    assert any("novel work" in a["name"].lower() for a in acts)
    # and it sorts last, after the recognised packages
    assert acts[-1]["wbs_phase"] == "other_uncategorized"


def test_an_override_moves_a_line_to_the_right_package():
    """Without the override the chilled-water line schedules as waterproofing."""
    default = group_by_category(BOQ)
    assert any("Chilled water" in (i["description"]) for i in
               default["Waterproofing/Insulation"])
    corrected = group_by_category(BOQ, OVERRIDES)
    assert any("Chilled water" in (i["description"]) for i in
               corrected["Mechanical/HVAC (MEP)"])
    assert "Waterproofing/Insulation" not in corrected


# ── missing productivity is refused, never defaulted ───────────────────────

def test_missing_productivity_refuses_and_names_the_gap():
    with pytest.raises(MissingProductivity) as exc:
        activities_from_boq(BOQ, manhours_per_unit={"Structural Steel": 8.0},
                            category_overrides=OVERRIDES)
    msg = str(exc.value)
    assert "Windows/Doors/Facade" in msg and "m2" in msg
    assert "man-hour" in msg.lower()
    assert "Mechanical/HVAC (MEP)" in msg
    # the machine-readable form drives the operator question
    assert exc.value.missing["Earthworks/Excavation"] == "m3"


def test_no_category_carries_a_built_in_rate():
    """Process, not facts: nothing in the module may supply an output rate."""
    import inspect

    from app.lib import boq_schedule
    src = inspect.getsource(boq_schedule)
    body = src.split("CONSTRUCTION_SEQUENCE")[0]
    assert "manhours_per_unit = " not in body
    with pytest.raises(MissingProductivity):
        activities_from_boq(BOQ, manhours_per_unit={}, category_overrides=OVERRIDES)


def test_basis_states_what_every_duration_rests_on():
    basis = " ".join(schedule_basis(RATES, crew_size={"Structural Steel": 8}))
    assert "Man-hours = quantity x norm" in basis
    assert "Structural Steel: 8 man-hours/unit, crew of 8" in basis
    assert "not assumed" in basis


# ── the endpoint: BOQ in, real programme out ───────────────────────────────

@pytest.mark.asyncio
async def test_endpoint_builds_a_workbook_from_the_boq(tmp_path, monkeypatch):
    """End-to-end through the route: BOQ line items -> CPM -> workbook."""
    import httpx
    from openpyxl import Workbook, load_workbook

    from app import main as app_main
    from app.core import projects as projects_store
    from app.routers import exports as exports_mod
    from app.dependencies import require_user

    src = tmp_path / "priced_boq.xlsx"
    wb = Workbook(); ws = wb.active
    ws.append(["Item", "Description", "Unit", "Qty", "Rate", "Amount"])
    for i, r in enumerate(BOQ):
        ws.append([f"{i+1}.0", r["description"], r["unit"], r["quantity"],
                   r["unit_cost"], r["total_cost"]])
    wb.save(src)

    monkeypatch.setattr(projects_store, "get_document",
                        lambda doc_id: {"id": doc_id, "file_path": str(src)},
                        raising=False)
    monkeypatch.setattr(exports_mod, "_check_owner",
                        lambda pid, uid: {"name": "Fence Project"}, raising=False)
    app_main.app.dependency_overrides[require_user] = lambda: {"user_id": "u1"}
    try:
        transport = httpx.ASGITransport(app=app_main.app)
        async with httpx.AsyncClient(transport=transport,
                                     base_url="http://testserver") as client:
            res = await client.post(
                "/v1/projects/p1/export/schedule-from-boq",
                json={"document_id": "d1", "manhours_per_unit": RATES,
                      "category_overrides": OVERRIDES, "start_date": "2026-09-01"})
    finally:
        app_main.app.dependency_overrides.pop(require_user, None)

    assert res.status_code == 200, res.text[:300]
    assert int(res.headers["X-Activities"]) == len(BOQ)
    assert int(res.headers["X-Duration-Days"]) > 0

    out = tmp_path / "out.xlsx"
    out.write_bytes(res.content)
    book = load_workbook(out)
    assert "L2 Schedule" in book.sheetnames
    names = " ".join(str(r[2]).lower() for r in
                     book["L2 Schedule"].iter_rows(min_row=4, values_only=True) if r and r[2])
    for term in ("truss", "cladding", "chilled water", "excavation"):
        assert term in names, f"{term!r} missing from the generated schedule"


@pytest.mark.asyncio
async def test_endpoint_refuses_and_names_the_missing_rates(tmp_path, monkeypatch):
    """The operator must be asked for the exact rates, not given a guess."""
    import httpx
    from openpyxl import Workbook

    from app import main as app_main
    from app.core import projects as projects_store
    from app.routers import exports as exports_mod
    from app.dependencies import require_user

    src = tmp_path / "b.xlsx"
    wb = Workbook(); ws = wb.active
    ws.append(["Item", "Description", "Unit", "Qty", "Rate", "Amount"])
    ws.append(["1.0", "Bulk excavation to reduced level", "m3", 2500, 18.5, 46250])
    wb.save(src)

    monkeypatch.setattr(projects_store, "get_document",
                        lambda doc_id: {"id": doc_id, "file_path": str(src)},
                        raising=False)
    monkeypatch.setattr(exports_mod, "_check_owner",
                        lambda pid, uid: {"name": "P"}, raising=False)
    app_main.app.dependency_overrides[require_user] = lambda: {"user_id": "u1"}
    try:
        transport = httpx.ASGITransport(app=app_main.app)
        async with httpx.AsyncClient(transport=transport,
                                     base_url="http://testserver") as client:
            res = await client.post("/v1/projects/p1/export/schedule-from-boq",
                                    json={"document_id": "d1", "manhours_per_unit": {}})
    finally:
        app_main.app.dependency_overrides.pop(require_user, None)

    assert res.status_code == 422
    body = res.json()
    detail = body.get("detail", body)
    assert "Earthworks/Excavation" in json.dumps(detail)
    assert "m3" in json.dumps(detail)


# ── substructure before superstructure (operator-reported, 2026-08-17) ─────
#
# "rc for foundation?" — the pricing categorizer answers WHICH TRADE, never
# WHICH PART OF THE BUILDING, so "Reinforced concrete C40 to raft foundation"
# and "...to columns" are both plain Concrete. Scheduled as one package in BOQ
# order, a raft listed second lands AFTER the columns it carries: a wrong
# programme, not a cosmetic one.

FOUNDATION_BOQ = [
    # deliberately listed columns-first, the order that used to break it
    {"item_key": "rc_columns", "description": "Reinforced concrete C40 to columns",
     "quantity": 180, "unit": "m3"},
    {"item_key": "rc_raft", "description": "Reinforced concrete C40 to raft foundation",
     "quantity": 420, "unit": "m3"},
    {"item_key": "blinding", "description": "Plain concrete blinding to foundations",
     "quantity": 160, "unit": "m2"},
    {"item_key": "rebar_found", "description": "Rebar cut bend and fix to foundations",
     "quantity": 38000, "unit": "kg"},
    {"item_key": "blockwork", "description": "Supply and install 200mm blockwork wall",
     "quantity": 1250, "unit": "m2"},
]
FOUNDATION_RATES = {"Concrete": 1.37, "Reinforcement": 0.023, "Masonry/Blockwork": 0.8}


def _order(acts):
    return {a["name"]: i for i, a in enumerate(acts)}


def test_foundation_concrete_precedes_column_concrete():
    acts = activities_from_boq(FOUNDATION_BOQ, manhours_per_unit=FOUNDATION_RATES)
    pos = _order(acts)
    assert pos["Reinforced concrete C40 to raft foundation"] < pos["Reinforced concrete C40 to columns"]
    assert pos["Plain concrete blinding to foundations"] < pos["Reinforced concrete C40 to columns"]


def test_substructure_work_is_marked_and_phased_separately():
    acts = activities_from_boq(FOUNDATION_BOQ, manhours_per_unit=FOUNDATION_RATES)
    by_name = {a["name"]: a for a in acts}
    raft = by_name["Reinforced concrete C40 to raft foundation"]
    cols = by_name["Reinforced concrete C40 to columns"]
    assert raft["boq"]["stage"] == "substructure"
    assert cols["boq"]["stage"] == "superstructure"
    # and they are different WBS phases, so a planner sees two packages
    assert raft["wbs_phase"] != cols["wbs_phase"]
    assert raft["wbs_phase"].endswith("_substructure")


def test_the_whole_substructure_finishes_before_the_superstructure_starts():
    acts = activities_from_boq(FOUNDATION_BOQ, manhours_per_unit=FOUNDATION_RATES)
    subs = [i for i, a in enumerate(acts) if a["boq"]["stage"] == "substructure"]
    supers = [i for i, a in enumerate(acts) if a["boq"]["stage"] == "superstructure"]
    assert subs and supers
    assert max(subs) < min(supers)


def test_foundation_rebar_is_staged_too_not_just_concrete():
    acts = activities_from_boq(FOUNDATION_BOQ, manhours_per_unit=FOUNDATION_RATES)
    rebar = next(a for a in acts if "Rebar" in a["name"])
    assert rebar["boq"]["stage"] == "substructure"


@pytest.mark.parametrize("desc", [
    "Reinforced concrete C40 to raft foundation",
    "RC foundation footings C35",
    "Plain concrete blinding to foundations",
    "Reinforced concrete to pile caps",
    "Concrete to ground beam",
    "Waterproofing/tanking to substructure walls",
])
def test_substructure_phrasings_are_recognised(desc):
    from app.lib.boq_pricing import categorize
    from app.lib.boq_schedule import SUBSTRUCTURE, element_stage
    assert element_stage(desc, categorize(desc)) == SUBSTRUCTURE, categorize(desc)


def test_superstructure_work_is_not_swept_into_substructure():
    from app.lib.boq_pricing import categorize
    from app.lib.boq_schedule import SUPERSTRUCTURE, element_stage
    for desc in ("Reinforced concrete C40 to columns",
                 "Reinforced concrete to suspended slab",
                 "Cement plaster 15mm to internal walls",
                 "Waterproofing membrane to roof slab"):
        assert element_stage(desc, categorize(desc)) == SUPERSTRUCTURE, desc


# ── the built programme teaches the norms (operator, 2026-08-17) ───────────
#
# "the built programs and its manpower histogram should tell u the
# productivity per manhour" — a completed programme IS a productivity record.
# Man-hours consumed / quantity delivered = the norm for that work, measured
# with these crews on this project, which outranks any published reference.

BUILT_PROGRAMME = [
    {"id": "A", "name": "Supply and install 200mm blockwork wall",
     "duration_days": 32, "crew_size": 4},
    {"id": "B", "name": "Reinforced concrete C40 to raft foundation",
     "duration_days": 12, "crew_size": 6},
    {"id": "C", "name": "Blockwork to lift shaft walls",
     "duration_days": 5, "crew_size": 4},
]
BUILT_QUANTITIES = {"A": 1250, "B": 420, "C": 190}


def test_norms_are_derived_from_a_built_programme():
    from app.lib.boq_schedule import norms_from_programme
    norms = norms_from_programme(BUILT_PROGRAMME, BUILT_QUANTITIES)
    block = norms["Masonry/Blockwork"]
    # (32d x 4 x 8h) + (5d x 4 x 8h) = 1184 mh over 1440 m2
    assert block["total_manhours"] == pytest.approx(1184.0)
    assert block["total_quantity"] == pytest.approx(1440.0)
    assert block["manhours_per_unit"] == pytest.approx(1184 / 1440, abs=0.001)
    assert block["samples"] == 2


def test_a_derived_norm_agrees_with_the_published_reference():
    """Sanity: 32 days for 1250 m2 with 4 men lands on the handbook's 0.8 mh/m2."""
    from app.lib.boq_schedule import from_daily_output, norms_from_programme
    norms = norms_from_programme([BUILT_PROGRAMME[0]], {"A": 1250})
    assert norms["Masonry/Blockwork"]["manhours_per_unit"] == pytest.approx(
        from_daily_output(10), abs=0.03)


def test_the_norm_is_quantity_weighted_not_an_average_of_ratios():
    """A 4,000 m2 slab and a 20 m2 landing must not carry equal weight."""
    from app.lib.boq_schedule import norms_from_programme
    prog = [
        {"id": "big", "name": "Cement plaster to walls", "duration_days": 40, "crew_size": 10},
        {"id": "small", "name": "Cement plaster to lift lobby", "duration_days": 4, "crew_size": 2},
    ]
    norms = norms_from_programme(prog, {"big": 4000, "small": 20})
    # weighted: (3200 + 64) / 4020 = 0.812  |  mean-of-ratios would be 2.0
    assert norms["Finishes"]["manhours_per_unit"] == pytest.approx(0.812, abs=0.005)


def test_explicit_manhours_beat_crew_times_duration():
    """A P6 resource assignment carries real man-hours; prefer it."""
    from app.lib.boq_schedule import activity_manhours
    assert activity_manhours({"total_manhours": 900, "crew_size": 4,
                              "duration_days": 32}) == 900
    assert activity_manhours({"target_qty": 750, "duration_days": 10}) == 750
    assert activity_manhours({"crew_size": 4, "duration_days": 10}) == 320


def test_an_activity_with_no_manhours_teaches_nothing_rather_than_zero():
    from app.lib.boq_schedule import activity_manhours, norms_from_programme
    assert activity_manhours({"name": "milestone only"}) is None
    norms = norms_from_programme(
        [{"id": "m", "name": "Blockwork milestone"}], {"m": 100})
    assert norms == {}


def test_unmeasured_activities_are_skipped_not_guessed():
    """No quantity means no norm — an unmeasured activity cannot teach one."""
    from app.lib.boq_schedule import norms_from_programme
    norms = norms_from_programme(BUILT_PROGRAMME, {"A": 1250})  # B and C unmeasured
    assert set(norms) == {"Masonry/Blockwork"}
    assert norms["Masonry/Blockwork"]["samples"] == 1


def test_derived_norms_feed_straight_back_into_a_new_programme():
    """The loop closes: measure on the last job, plan the next one with it."""
    from app.lib.boq_schedule import norms_from_programme
    derived = norms_from_programme(BUILT_PROGRAMME, BUILT_QUANTITIES)
    norms = {cat: v["manhours_per_unit"] for cat, v in derived.items()}
    acts = activities_from_boq(
        [{"description": "Supply and install 200mm blockwork wall",
          "quantity": 2500, "unit": "m2"}],
        manhours_per_unit=norms)
    # 2500 x 0.822 mh = 2056 mh / (4 x 8) = 64.2 -> 65 days
    assert acts[0]["duration_days"] == 65
    assert acts[0]["boq"]["manhours_per_unit"] == pytest.approx(0.822, abs=0.002)


# ── the histogram must agree with the planned man-hours ────────────────────

def test_bridge_reads_crew_size_instead_of_multiplying_heads_by_trades():
    """resources carries one entry per HEAD for man-hour activities; the old
    len(resources) x crew_per_trade turned a crew of 4 into 16 men and
    inflated man-days, the S-curve and the histogram together."""
    from app.lib.schedule_bridge import bridge_activity
    act = {"id": "1.1", "name": "Blockwork", "duration_days": 32,
           "resources": ["masonry"] * 4, "crew_size": 4, "total_manhours": 1000.0}
    out = bridge_activity(act, crew_per_trade=4)
    assert out["manpower"] == 4
    assert out["manhours"] == 1000.0


def test_bridge_still_guesses_for_template_activities():
    """generate_wbs puts one entry per TRADE and states no crew — unchanged."""
    from app.lib.schedule_bridge import bridge_activity
    out = bridge_activity({"id": "1.1", "name": "Survey", "duration_days": 7,
                           "resources": ["geotech"]}, crew_per_trade=4)
    assert out["manpower"] == 4
    assert "manhours" not in out


def test_manhours_reconcile_with_crew_times_duration_times_shift():
    """The two derivations of the same labour must land together."""
    from app.lib.schedule_bridge import bridge_wbs_to_cost_loaded
    acts = activities_from_boq(
        [{"description": "Supply and install 200mm blockwork wall",
          "quantity": 1250, "unit": "m2"}],
        manhours_per_unit={"Masonry/Blockwork": 0.8})
    bridged = bridge_wbs_to_cost_loaded(acts, crew_per_trade=4)
    b = bridged[0]
    man_days = b["duration"] * b["manpower"]
    assert b["manhours"] == pytest.approx(1000.0)
    # 32 d x 4 heads x 8 h = 1024 mh vs 1000 planned: the rounding-up of the
    # final part-day, never more than one crew-shift.
    assert 0 <= man_days * 8 - b["manhours"] <= b["manpower"] * 8


def test_workbook_shows_manhours_when_planned_in_them(tmp_path):
    from openpyxl import load_workbook

    from app.lib.pm_excel import generate_cost_loaded_schedule
    from app.lib.schedule_bridge import bridge_wbs_to_cost_loaded
    from app.containers.construction import ConstructionContainer

    acts = activities_from_boq(
        [{"description": "Supply and install 200mm blockwork wall", "quantity": 1250, "unit": "m2"},
         {"description": "Reinforced concrete C40 to raft foundation", "quantity": 420, "unit": "m3"}],
        manhours_per_unit={"Masonry/Blockwork": 0.8, "Concrete": 1.37})
    enriched, _summary, err = ConstructionContainer()._attach_cpm_to_activities(acts, None)
    assert not err
    wb = generate_cost_loaded_schedule(
        {"project": "P", "currency": "SAR", "hours_per_day": 8},
        bridge_wbs_to_cost_loaded(enriched, crew_per_trade=4))
    out = tmp_path / "s.xlsx"; wb.save(out)
    ws = load_workbook(out)["Manpower Histogram"]
    header = [c.value for c in ws[3]]
    assert "Planned man-hours" in header
    col = header.index("Planned man-hours") + 1
    # look the row up by name: substructure sorts FIRST, so row 4 is the raft
    rows = {ws.cell(r, 2).value: r for r in range(4, ws.max_row + 1)}
    assert ws.cell(rows["Supply and install 200mm blockwork wall"], col).value         == pytest.approx(1000.0)
    assert ws.cell(rows["Reinforced concrete C40 to raft foundation"], col).value         == pytest.approx(575.4, abs=0.1)


def test_workbook_omits_the_manhour_columns_for_template_schedules(tmp_path):
    """A brief-driven schedule has no man-hours; the sheet must look unchanged."""
    from openpyxl import load_workbook

    from app.lib.pm_excel import generate_cost_loaded_schedule
    from app.lib.schedule_bridge import bridge_wbs_to_cost_loaded

    plain = [{"id": "1.1", "name": "Survey", "duration_days": 7,
              "resources": ["geotech"], "wbs_phase": "site", "predecessors": []}]
    wb = generate_cost_loaded_schedule({"project": "P", "currency": "SAR"},
                                       bridge_wbs_to_cost_loaded(plain))
    out = tmp_path / "t.xlsx"; wb.save(out)
    header = [c.value for c in load_workbook(out)["Manpower Histogram"][3]]
    assert header[:5] == ["ID", "Activity", "Dur", "Manpower", "Man-days"]
    assert "Planned man-hours" not in header
