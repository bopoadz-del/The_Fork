"""The BOQ processor advertises .xls -- it must actually parse one.

Live-fire campaign phase 3, run T1c: a real 652KB contract-volume BOQ in the
old BIFF .xls format (12,971 rows, readable by pandas/xlrd in one call) came
back "could not extract line items". Mechanism: `_parse_excel` hardcodes
`engine="openpyxl"`, and openpyxl reads only zip-based .xlsx -- every .xls
dies with "File is not a zip file" while the block's own ui_schema lists
".xls" in its accept list. An advertised capability that could never work.

The fixture is generated with xlwt if available; otherwise the test that
needs a real BIFF file skips LOUDLY rather than silently passing.
"""
from __future__ import annotations

import pathlib

import pytest

from app.blocks.boq_processor import BOQProcessorBlock

ROWS = [
    ("Description", "Quantity", "Unit", "Rate", "Amount"),
    ("Excavation to reduced level", 1200, "m3", 25, 30000),
    ("Concrete blinding 75mm", 300, "m2", 45, 13500),
    ("Reinforced concrete in raft", 950, "m3", 380, 361000),
]


@pytest.fixture
def xls_boq(tmp_path):
    xlwt = pytest.importorskip(
        "xlwt", reason="xlwt needed to author a real BIFF .xls fixture")
    wb = xlwt.Workbook()
    ws = wb.add_sheet("BOQ")
    for r, row in enumerate(ROWS):
        for c, val in enumerate(row):
            ws.write(r, c, val)
    path = str(tmp_path / "legacy_boq.xls")
    wb.save(path)
    return path


@pytest.mark.asyncio
async def test_a_biff_xls_boq_parses_to_line_items(xls_boq):
    """RED before the fix: 'Parse error: File is not a zip file'."""
    res = await BOQProcessorBlock().process({"file_path": xls_boq}, {})
    assert res.get("status") == "success", res.get("error")
    items = res.get("line_items") or []
    assert len(items) == 3, items
    descs = " ".join(str(i.get("description", "")) for i in items)
    assert "Excavation" in descs and "raft" in descs


@pytest.mark.asyncio
async def test_xlsx_still_parses_after_the_engine_change(tmp_path):
    """The other direction: fixing .xls must not regress .xlsx."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    for row in ROWS:
        ws.append(list(row))
    path = str(tmp_path / "modern_boq.xlsx")
    wb.save(path)

    res = await BOQProcessorBlock().process({"file_path": path}, {})
    assert res.get("status") == "success", res.get("error")
    assert len(res.get("line_items") or []) == 3


@pytest.mark.asyncio
async def test_engine_is_selected_by_extension(monkeypatch, tmp_path):
    """The machinery decision, runnable in CI without a BIFF author.

    The real-file test above needs xlwt and skips where absent; this one pins
    the mechanism itself -- .xls must go to xlrd, .xlsx to openpyxl -- so the
    regression cannot come back invisibly on a runner without xlwt.
    """
    import pandas as pd
    seen = {}

    def fake_read_excel(path, sheet_name=0, engine=None, header="unset"):
        seen[pathlib.Path(str(path)).suffix] = engine
        # header=None is part of the contract now (headerless read + detect);
        # mirror pandas' shape for that mode.
        return pd.DataFrame(
            [["Description", "Quantity", "Unit", "Rate"],
             ["x", 1, "m", 1]])

    monkeypatch.setattr(pd, "read_excel", fake_read_excel)
    block = BOQProcessorBlock()
    await block._parse_excel(str(tmp_path / "legacy.xls"), {})
    await block._parse_excel(str(tmp_path / "modern.xlsx"), {})
    assert seen[".xls"] == "xlrd", seen
    assert seen[".xlsx"] == "openpyxl", seen


@pytest.mark.asyncio
async def test_a_boq_with_a_title_block_above_the_header_still_parses(tmp_path):
    """Real contract volumes open with project/employer/bill-number rows; the
    column header sits many rows down. Found live on a 12,971-row Vol III bill:
    the parser assumed headers at row 0, resolved no columns, and extracted
    ZERO line items from a file pandas read perfectly.

    Reproduced here as .xlsx so the fence runs everywhere (the .xls variant of
    the same file needs xlwt to author and skips where absent).
    """
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["PROJECT: JUMEIRAH ISLANDS"])
    ws.append(["For M/s EMPLOYER"])
    ws.append([])
    ws.append(["BILL NO.4 - APARTMENTS"])
    ws.append(["Item", "Description", "Qty", "Unit", "Rate", "Amount"])
    ws.append(["A", "Shear wall", 3533, "m3", 380, 1342540])
    ws.append(["B", "Lift walls", 1018, "m3", 380, 386840])
    ws.append(["C", "Staircase & landings", 344, "m3", 395, 135880])
    path = str(tmp_path / "titled_boq.xlsx")
    wb.save(path)

    res = await BOQProcessorBlock().process({"file_path": path}, {})
    assert res.get("status") == "success", res.get("error")
    items = res.get("line_items") or []
    assert len(items) == 3, f"title block swallowed the bill: {len(items)} items"
    descs = " ".join(str(i.get("description")) for i in items)
    assert "Shear wall" in descs and "Staircase" in descs
