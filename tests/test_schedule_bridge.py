"""The bridge turns raw generate_wbs activities into a cost-loaded schedule
workbook with an honest man-days S-curve + milestones, and never fabricates a
cost column when no rate is supplied."""
from __future__ import annotations

from app.lib.schedule_bridge import bridge_wbs_to_cost_loaded, bridge_activity
from app.lib.pm_excel import generate_cost_loaded_schedule

# Shape as emitted by ConstructionContainer.generate_wbs / _build_wbs_activities.
WBS = [
    {"id": "1.1.1", "name": "Mobilization", "duration_days": 5, "predecessors": [],
     "resources": ["labor"], "wbs_phase": "Enabling"},
    {"id": "1.1.2", "name": "Earthworks", "duration_days": 10, "predecessors": ["1.1.1"],
     "resources": ["excavator", "labor"], "wbs_phase": "Substructure"},
    {"id": "1.1.3", "name": "Foundations", "duration_days": 8, "predecessors": ["1.1.2"],
     "resources": ["labor", "concrete", "steelfixer"], "wbs_phase": "Substructure"},
]


def test_bridge_maps_fields_and_derives_manpower():
    out = bridge_wbs_to_cost_loaded(WBS, crew_per_trade=4)
    a = {x["id"]: x for x in out}
    assert a["1.1.1"]["duration"] == 5 and a["1.1.1"]["wbs"] == "Enabling"
    # manpower = trades x crew_per_trade
    assert a["1.1.1"]["manpower"] == 4       # 1 trade
    assert a["1.1.2"]["manpower"] == 8       # 2 trades
    assert a["1.1.3"]["manpower"] == 12      # 3 trades
    assert a["1.1.2"]["predecessors"] == ["1.1.1"]


def test_bridge_never_fabricates_cost_without_rate():
    out = bridge_wbs_to_cost_loaded(WBS)
    assert all("cost" not in x for x in out), "no cost may appear without a day-rate"


def test_bridge_indicative_cost_with_rate():
    out = bridge_wbs_to_cost_loaded(WBS, crew_per_trade=4, day_rate=100)
    a = {x["id"]: x for x in out}
    # indicative labor = manpower x duration x rate
    assert a["1.1.1"]["cost"] == 4 * 5 * 100
    assert a["1.1.3"]["cost"] == 12 * 8 * 100


def test_bridge_preserves_existing_cost_and_manpower():
    a = bridge_activity({"id": "9", "name": "X", "duration": 3,
                         "cost": 5000, "manpower": 7, "wbs": "1.4"})
    assert a["cost"] == 5000 and a["manpower"] == 7 and a["duration"] == 3


def test_workbook_from_bridged_wbs_has_scurve_and_milestones_no_fake_cost():
    acts = bridge_wbs_to_cost_loaded(WBS)          # no rate -> no cost
    wb = generate_cost_loaded_schedule({"project": "DC", "start_date": "2026-01-05"}, acts)
    assert "Milestones" in wb.sheetnames
    # the S-curve is a LINE chart on the progress-baseline sheet
    from openpyxl.chart import LineChart
    assert any(isinstance(c, LineChart) for c in wb["Cost Loading"]._charts), "man-days S-curve line chart"
    # cumulative man-days present (a running =prev+curr on the baseline)
    cl_txt = "\n".join(str(c.value) for row in wb["Cost Loading"].iter_rows()
                       for c in row if c.value is not None)
    assert "Man-days" in cl_txt and "%" not in cl_txt[:20]
    # no fabricated cost: L2 cost column is blank, Summary shows no Total Cost row
    summ_txt = "\n".join(str(c.value) for row in wb["Summary"].iter_rows()
                         for c in row if c.value is not None)
    assert "Total Cost" not in summ_txt and "Indicative" not in summ_txt
    assert "Total Man-days" in summ_txt
    # milestones: 2 phase groups + project completion = 3 rows, dated
    ms_txt = "\n".join(str(c.value) for row in wb["Milestones"].iter_rows()
                       for c in row if c.value is not None)
    assert "PROJECT COMPLETION" in ms_txt and "2026-" in ms_txt


def test_workbook_with_rate_labels_cost_indicative():
    acts = bridge_wbs_to_cost_loaded(WBS, day_rate=100)
    wb = generate_cost_loaded_schedule(
        {"project": "DC", "cost_basis": "Indicative Labor"}, acts)
    summ_txt = "\n".join(str(c.value) for row in wb["Summary"].iter_rows()
                         for c in row if c.value is not None)
    assert "Indicative Labor" in summ_txt
