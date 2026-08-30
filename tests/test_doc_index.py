"""Tests for app/core/doc_index.py — document text extraction, chunking,
and per-project index persistence.

TDD order:
  Wave 1 — Task 1: text extraction
  Wave 2 — Task 2: chunking
  Wave 3 — Task 3: index build + JSON persistence
"""

import importlib
import os
import sys
import threading

import pytest
from cryptography.fernet import Fernet

from app.core import file_crypto
from app.core import projects as projects_mod


# ────────────────────────────────────────────────────────────────────────────
# Helpers
# ────────────────────────────────────────────────────────────────────────────

FIXTURES_DIR = os.path.join(os.path.dirname(__file__), "fixtures")
SAMPLE_TXT = os.path.join(FIXTURES_DIR, "sample.txt")


# ────────────────────────────────────────────────────────────────────────────
# WAVE 1 — Task 1: extract_document_text
# ────────────────────────────────────────────────────────────────────────────

def test_extract_txt(tmp_path, monkeypatch):
    """Plain-text fixture round-trips; known phrase appears in result."""
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    from app.core.doc_index import extract_document_text

    # Copy sample.txt to a tmp location so we have a real writable path
    content = open(SAMPLE_TXT, "rb").read()
    doc_path = str(tmp_path / "sample.txt")
    file_crypto.write_document(doc_path, content)

    text = extract_document_text(doc_path, "sample.txt")
    assert "Kingsbridge Tower" in text
    assert "Sandra Okafor" in text


def test_extract_pdf_mocked(tmp_path, monkeypatch):
    """PDF extraction: monkeypatch fitz.open; assert concatenated page text."""
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)

    # Write a dummy file so open_plaintext has something to open
    doc_path = str(tmp_path / "report.pdf")
    file_crypto.write_document(doc_path, b"%PDF-1.4 dummy")

    # Build fake fitz objects
    class FakePage:
        def get_text(self):
            return "Page text alpha. "

    class FakeDoc:
        def __iter__(self):
            return iter([FakePage(), FakePage()])

        def close(self):
            pass

    import fitz as real_fitz
    monkeypatch.setattr(real_fitz, "open", lambda path: FakeDoc())

    from app.core import doc_index
    importlib.reload(doc_index)
    text = doc_index.extract_document_text(doc_path, "report.pdf")
    assert "Page text alpha." in text
    assert text.count("Page text alpha.") == 2


def test_extract_pdf_ocrs_image_pages_even_when_cover_has_text(tmp_path, monkeypatch):
    """Cover-page bug fix: a PDF whose FIRST page has a text layer but whose
    body pages are image-only must still OCR the body pages (per-page), instead
    of skipping OCR because the whole-doc text passed the threshold."""
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    doc_path = str(tmp_path / "scanned_boq.pdf")
    file_crypto.write_document(doc_path, b"%PDF-1.4 dummy")

    class FakePage:
        def __init__(self, txt):
            self._t = txt

        def get_text(self):
            return self._t

    class FakeDoc:
        def __iter__(self):
            return iter([
                FakePage("Cover: Site Demolition Works Package 1 Volume 2 specification."),
                FakePage(""),   # image-only body page
                FakePage(""),   # image-only body page
            ])

        def close(self):
            pass

    import fitz as real_fitz
    monkeypatch.setattr(real_fitz, "open", lambda path: FakeDoc())

    from app.core import doc_index
    importlib.reload(doc_index)
    # Stub the single-page OCR so the test doesn't need tesseract.
    monkeypatch.setattr(doc_index, "_ocr_pdf_page",
                        lambda page: "OCR body text: excavation 5000 m3")

    text, meta = doc_index._extract_with_meta(doc_path, "scanned_boq.pdf")
    assert "Cover:" in text, "the cover text-layer page must be kept"
    assert "OCR body text" in text, "image body pages must be OCR'd per-page (cover-page bug)"
    assert meta.get("ocr_low_quality") is True


def test_pdf_tables_disabled_for_large_files(monkeypatch):
    """pdfplumber loads the whole PDF into memory — the OOM hazard on a large
    scan. Skip it above PDF_TABLES_MAX_MB (per-page OCR still captures text;
    image-only scans have no extractable tables anyway). Small files keep it."""
    from app.core import doc_index
    importlib.reload(doc_index)
    monkeypatch.setenv("PDF_TABLES_MAX_MB", "25")
    monkeypatch.setattr("os.path.getsize", lambda p: 60 * 1024 * 1024)  # 60 MB
    assert doc_index._pdf_tables_enabled("big_scan.pdf") is False
    monkeypatch.setattr("os.path.getsize", lambda p: 2 * 1024 * 1024)   # 2 MB
    assert doc_index._pdf_tables_enabled("small.pdf") is True


def test_extract_pdf_ocr_page_cap_bounds_memory(tmp_path, monkeypatch):
    """OCR is capped so a long scan can't OOM the 2GB box: pages beyond the
    cap set ocr_truncated and are NOT OCR'd."""
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    monkeypatch.setenv("PDF_OCR_PAGE_CAP", "1")
    doc_path = str(tmp_path / "long_scan.pdf")
    file_crypto.write_document(doc_path, b"%PDF-1.4 dummy")

    class FakePage:
        def get_text(self):
            return ""  # all image pages

    class FakeDoc:
        def __iter__(self):
            return iter([FakePage(), FakePage(), FakePage()])

        def close(self):
            pass

    import fitz as real_fitz
    monkeypatch.setattr(real_fitz, "open", lambda path: FakeDoc())

    from app.core import doc_index
    importlib.reload(doc_index)
    calls = {"n": 0}

    def _fake_ocr(page):
        calls["n"] += 1
        return "page ocr text"

    monkeypatch.setattr(doc_index, "_ocr_pdf_page", _fake_ocr)
    text, meta = doc_index._extract_with_meta(doc_path, "long_scan.pdf")
    assert calls["n"] == 1, "OCR must stop at the page cap (memory bound)"
    assert meta.get("ocr_truncated") is True


def test_pdf_ocr_limits_read_at_call_time(monkeypatch):
    """Dashboard env changes must apply without re-importing doc_index.

    The old import-bound ``_PDF_OCR_MAX_SIZE_MB`` silently ignored a live
    ``PDF_OCR_MAX_SIZE_MB`` bump until the worker restarted and re-imported.
    """
    from app.core import doc_index

    monkeypatch.setenv("PDF_OCR_MAX_SIZE_MB", "32")
    monkeypatch.setenv("PDF_OCR_PAGE_CAP", "80")
    monkeypatch.setenv("PDF_OCR_BOQ_PAGE_CAP", "160")
    assert doc_index.pdf_ocr_max_size_mb() == 32
    assert doc_index.pdf_ocr_page_cap("drawing.pdf") == 80
    assert doc_index.pdf_ocr_page_cap(
        "IP-INF-053-0000-JCB-BOQ-CA-000007-B_Bill of Quantities (Priced).pdf"
    ) == 160
    monkeypatch.setenv("PDF_OCR_MAX_SIZE_MB", "25")
    assert doc_index.pdf_ocr_max_size_mb() == 25
    # A leftover dashboard 25 must not re-skip priced BOQs.
    assert doc_index.pdf_ocr_max_size_mb(
        "IP-INF-053-0000-JCB-BOQ-CA-000007-B_Bill of Quantities (Priced).pdf"
    ) == 100


def test_extract_pdf_ocrs_priced_boq_sized_scan(tmp_path, monkeypatch):
    """A ~28 MB image-only PDF must still OCR.

    Live priced BOQ 20ac033d is ~28 MB. The old PDF_OCR_MAX_SIZE_MB=25 set
    page_cap=0, so item codes never entered the index.
    """
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    monkeypatch.setenv("PDF_OCR_MAX_SIZE_MB", "32")
    doc_path = str(tmp_path / "priced_boq.pdf")
    file_crypto.write_document(doc_path, b"%PDF-1.4 dummy")

    class FakePage:
        def get_text(self):
            return ""

    class FakeDoc:
        def __iter__(self):
            return iter([FakePage()])

        def close(self):
            pass

    import fitz as real_fitz
    monkeypatch.setattr(real_fitz, "open", lambda path: FakeDoc())

    from app.core import doc_index
    importlib.reload(doc_index)

    real_getsize = os.path.getsize

    def _getsize(path):
        if os.path.abspath(path) == os.path.abspath(doc_path):
            return 28 * 1024 * 1024
        return real_getsize(path)

    monkeypatch.setattr(os.path, "getsize", _getsize)
    monkeypatch.setattr(
        doc_index, "_ocr_pdf_page",
        lambda page: "D599.5 Breaking out existing carriageway 1200 m2 18.00 21600",
    )
    monkeypatch.setattr(doc_index, "_pdf_tables_enabled", lambda *a, **k: False)

    text, meta = doc_index._extract_pdf(
        doc_path,
        "IP-INF-053-0000-JCB-BOQ-CA-000007-B_Bill of Quantities (Priced).pdf",
    )
    assert "D599.5" in text, "priced-BOQ item codes must be OCR'd at ~28 MB"
    assert meta.get("ocr_skipped_too_large") is not True
    assert meta.get("ocr_pages", 0) >= 1


def test_extract_pdf_skips_ocr_above_raised_cap(tmp_path, monkeypatch):
    """Files above the new 32 MB OCR ceiling still skip OCR (OOM bound)."""
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    monkeypatch.setenv("PDF_OCR_MAX_SIZE_MB", "32")
    doc_path = str(tmp_path / "huge_scan.pdf")
    file_crypto.write_document(doc_path, b"%PDF-1.4 dummy")

    class FakePage:
        def get_text(self):
            return ""

    class FakeDoc:
        def __iter__(self):
            return iter([FakePage(), FakePage()])

        def close(self):
            pass

    import fitz as real_fitz
    monkeypatch.setattr(real_fitz, "open", lambda path: FakeDoc())

    from app.core import doc_index
    importlib.reload(doc_index)
    real_getsize = os.path.getsize

    def _getsize(path):
        if os.path.abspath(path) == os.path.abspath(doc_path):
            return 40 * 1024 * 1024
        return real_getsize(path)

    monkeypatch.setattr(os.path, "getsize", _getsize)
    calls = {"n": 0}

    def _fake_ocr(page):
        calls["n"] += 1
        return "should not run"

    monkeypatch.setattr(doc_index, "_ocr_pdf_page", _fake_ocr)
    monkeypatch.setattr(doc_index, "_pdf_tables_enabled", lambda *a, **k: False)

    text, meta = doc_index._extract_pdf(doc_path, "huge_scan.pdf")
    assert calls["n"] == 0
    assert meta.get("ocr_skipped_too_large") is True
    assert "should not run" not in text


def test_boq_filename_uses_higher_ocr_page_cap(tmp_path, monkeypatch):
    """A BOQ-named scan uses PDF_OCR_BOQ_PAGE_CAP, not the generic cap.

    Demolition was PARTIAL at 40 pages; later codes (D549.2, D599.5) sat
    past the generic budget.
    """
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    monkeypatch.setenv("PDF_OCR_PAGE_CAP", "1")
    monkeypatch.setenv("PDF_OCR_BOQ_PAGE_CAP", "3")
    doc_path = str(tmp_path / "demolition_boq.pdf")
    file_crypto.write_document(doc_path, b"%PDF-1.4 dummy")

    class FakePage:
        def get_text(self):
            return ""

    class FakeDoc:
        def __iter__(self):
            return iter([FakePage() for _ in range(5)])

        def close(self):
            pass

    import fitz as real_fitz
    monkeypatch.setattr(real_fitz, "open", lambda path: FakeDoc())

    from app.core import doc_index
    importlib.reload(doc_index)
    calls = {"n": 0}

    def _fake_ocr(page):
        calls["n"] += 1
        return f"ocr page {calls['n']} D549.2 D599.5"

    monkeypatch.setattr(doc_index, "_ocr_pdf_page", _fake_ocr)
    monkeypatch.setattr(doc_index, "_pdf_tables_enabled", lambda *a, **k: False)

    text, meta = doc_index._extract_pdf(doc_path, "Demolition Bill of Quantities.pdf")
    assert calls["n"] == 3, "BOQ page cap must apply, not the generic cap of 1"
    assert "D549.2" in text and "D599.5" in text
    assert meta.get("ocr_truncated") is True


def test_extract_pdf_batched_keeps_prior_pages_on_timeout(tmp_path, monkeypatch):
    """A later-range timeout must keep earlier OCR, not ZERO_CHUNK the doc."""
    from app.core import doc_index
    importlib.reload(doc_index)

    monkeypatch.setenv("PDF_OCR_BATCH_PAGES", "20")
    monkeypatch.setenv("PDF_OCR_BOQ_PAGE_CAP", "160")
    ranges: list[tuple[int, int]] = []

    def _fake_isolated(fn, args, fallback, label=""):
        start, end = args[2], args[3]
        ranges.append((start, end))
        if start >= 20:
            return fallback, {
                "extract_failed": "timeout",
                "extract_failed_detail": "extraction exceeded 600s",
            }
        return (
            "pages 0-20 D549.2 Removal of existing chain link fence",
            {"ocr_pages": 5, "ocr_low_quality": True},
        ), {}

    monkeypatch.setattr(doc_index, "_pdf_page_count", lambda p: 45)
    monkeypatch.setattr("app.core.extract_isolated.run_isolated", _fake_isolated)

    text, meta = doc_index._extract_pdf_batched(
        str(tmp_path / "priced.pdf"),
        "IP-INF-053-0000-JCB-BOQ-CA-000007-B_Bill of Quantities (Priced).pdf",
    )
    assert ranges[0] == (0, 20)
    assert ranges[1] == (20, 40)
    assert "D549.2" in text
    assert meta.get("extract_partial") is True
    assert meta.get("extract_failed") == "timeout"
    assert meta.get("ocr_pages") == 5


def test_chunker_for_boq_filename_is_finer():
    from app.core.doc_index import _chunker_for_document

    priced = "IP-INF-053-0000-JCB-BOQ-CA-000007-B_Bill of Quantities (Priced).pdf"
    assert _chunker_for_document(priced) == "finer"
    assert _chunker_for_document("specification_part8.pdf") == "default"
    assert _chunker_for_document(priced, "markdown") == "markdown"


def test_ocr_size_gate_uses_plaintext_not_ciphertext(tmp_path, monkeypatch):
    """Fernet on-disk size must not skip OCR of a 28 MB plaintext scan.

    Live 20ac033d is 28,180,707 plaintext bytes. Encrypted at rest that is
    ~36 MB on disk — above the 32 MB gate — so #448 still skipped OCR and
    a 15s reindex recast the cover layer as 6 chunks.
    """
    monkeypatch.setenv("DATA_ENCRYPTION_KEY", Fernet.generate_key().decode())
    monkeypatch.setenv("PDF_OCR_MAX_SIZE_MB", "32")
    doc_path = str(tmp_path / "priced_boq.pdf")
    file_crypto.write_document(doc_path, b"%PDF-1.4 dummy")

    class FakePage:
        def get_text(self):
            return ""

    class FakeDoc:
        def __iter__(self):
            return iter([FakePage()])

        def close(self):
            pass

    import fitz as real_fitz
    monkeypatch.setattr(real_fitz, "open", lambda path: FakeDoc())

    from app.core import doc_index
    importlib.reload(doc_index)

    monkeypatch.setattr(doc_index.file_crypto, "plaintext_size", lambda p: 28 * 1024 * 1024)
    real_getsize = os.path.getsize

    def _getsize(path):
        if os.path.abspath(path) == os.path.abspath(doc_path):
            return 38 * 1024 * 1024
        return real_getsize(path)

    monkeypatch.setattr(os.path, "getsize", _getsize)
    monkeypatch.setattr(
        doc_index, "_ocr_pdf_page",
        lambda page: "D599.5 Breaking out existing carriageway 340904 m2 31.00",
    )
    monkeypatch.setattr(doc_index, "_pdf_tables_enabled", lambda *a, **k: False)

    text, meta = doc_index._extract_pdf(
        doc_path,
        "IP-INF-053-0000-JCB-BOQ-CA-000007-B_Bill of Quantities (Priced).pdf",
    )
    assert "D599.5" in text, "ciphertext size must not skip OCR of a 28 MB scan"
    assert meta.get("ocr_skipped_too_large") is not True
    assert meta.get("ocr_pages", 0) >= 1
    assert meta.get("ocr_required") is not True


def test_index_document_scanned_boq_without_ocr_is_not_ok(
    fresh_db, tmp_path, monkeypatch
):
    """Regression: a scanned BOQ with empty body pages must not report
    status=ok just because the finer chunker split the cover + guard.

    This is the live 20ac033d failure mode: 6 chunks, 15s, D599.5 absent.
    """
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    monkeypatch.setenv("RAG_EMBEDDING_MODEL", "fake")
    from app.core import doc_index
    importlib.reload(doc_index)
    from app.core.rag import vector_store as _vs
    _vs.reset_store_cache()

    proj = projects_mod.create_project("Scanned BOQ No OCR")
    pid = proj["id"]
    doc_path = _write_txt_doc(tmp_path, "Priced BOQ.pdf", b"%PDF-1.4 cover")
    doc = projects_mod.add_document(
        pid, "Bill of Quantities (Priced).pdf", file_path=doc_path, size=100
    )

    cover = (
        "VERIFIED-TOTAL GUARD for 'Bill of Quantities (Priced).pdf': "
        "this is a Bill of Quantities, but NO verified total could be "
        "computed from it — it appears to be a scanned/image PDF."
    )
    monkeypatch.setattr(
        doc_index,
        "_extract_with_meta",
        lambda *a, **k: (
            cover,
            {
                "empty_text_pages": 368,
                "ocr_pages": 0,
                "ocr_attempts": 0,
                "ocr_required": True,
                "ocr_skipped_too_large": True,
            },
        ),
    )
    monkeypatch.setattr(doc_index, "_boq_chunks_for_document", lambda *a, **k: [])
    monkeypatch.setattr(doc_index, "_drawing_chunks_for_document", lambda *a, **k: [])
    monkeypatch.setattr(doc_index, "_ifc_chunks_for_document", lambda *a, **k: [])

    result = doc_index.index_document(pid, doc["id"], chunker="finer")
    assert result["status"] != "ok", result
    assert result["error"] == "OCR_REQUIRED", result
    assert result.get("ocr_pages", 0) == 0
    assert result.get("indexed") == 0


def test_index_document_scanned_boq_with_ocr_indexes_item_codes(
    fresh_db, tmp_path, monkeypatch
):
    """When OCR actually ran, D599.5 / D549.2 must land in searchable chunks."""
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    monkeypatch.setenv("RAG_EMBEDDING_MODEL", "fake")
    from app.core import doc_index
    importlib.reload(doc_index)
    from app.core.rag import vector_store as _vs
    _vs.reset_store_cache()

    proj = projects_mod.create_project("Scanned BOQ OCR")
    pid = proj["id"]
    doc_path = _write_txt_doc(tmp_path, "Priced BOQ.pdf", b"%PDF-1.4 cover")
    doc = projects_mod.add_document(
        pid, "Bill of Quantities (Priced).pdf", file_path=doc_path, size=100
    )

    body = (
        "D549.2 Removal of existing chain link fence 3504 m 80.00 280320\n"
        "D599.5 Breaking out existing carriageway 340904 m2 31.00 10568024\n"
    )
    monkeypatch.setattr(
        doc_index,
        "_extract_with_meta",
        lambda *a, **k: (
            body,
            {"ocr_pages": 12, "empty_text_pages": 12, "ocr_attempts": 12},
        ),
    )
    monkeypatch.setattr(doc_index, "_boq_chunks_for_document", lambda *a, **k: [])
    monkeypatch.setattr(doc_index, "_drawing_chunks_for_document", lambda *a, **k: [])
    monkeypatch.setattr(doc_index, "_ifc_chunks_for_document", lambda *a, **k: [])

    result = doc_index.index_document(pid, doc["id"], chunker="finer")
    assert result["status"] == "ok", result
    assert result.get("ocr_pages") == 12
    saved = doc_index._load_index(pid)
    blob = "\n".join(saved["documents"][0]["chunks"])
    assert "D599.5" in blob and "D549.2" in blob
    assert "340904" in blob or "340,904" in blob
    assert "31.00" in blob


def test_force_ocr_runs_when_size_gate_would_skip(tmp_path, monkeypatch):
    """Admin reindex force_ocr=true OCRs a file above PDF_OCR_MAX_SIZE_MB."""
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    monkeypatch.setenv("PDF_OCR_MAX_SIZE_MB", "25")
    monkeypatch.setenv("PDF_OCR_BOQ_MAX_SIZE_MB", "25")
    doc_path = str(tmp_path / "priced_boq.pdf")
    file_crypto.write_document(doc_path, b"%PDF-1.4 dummy")

    class FakePage:
        def get_text(self):
            return ""

    class FakeDoc:
        def __iter__(self):
            return iter([FakePage()])

        def close(self):
            pass

    import fitz as real_fitz
    monkeypatch.setattr(real_fitz, "open", lambda path: FakeDoc())

    from app.core import doc_index
    importlib.reload(doc_index)

    real_getsize = os.path.getsize

    def _getsize(path):
        if os.path.abspath(path) == os.path.abspath(doc_path):
            return 28 * 1024 * 1024
        return real_getsize(path)

    monkeypatch.setattr(os.path, "getsize", _getsize)
    monkeypatch.setattr(doc_index.file_crypto, "plaintext_size", lambda p: 28 * 1024 * 1024)
    monkeypatch.setattr(
        doc_index, "_ocr_pdf_page",
        lambda page: "D549.2 Removal of existing chain link fence 3504 m 80.00",
    )
    monkeypatch.setattr(doc_index, "_pdf_tables_enabled", lambda *a, **k: False)

    text, meta = doc_index._extract_pdf(
        doc_path,
        "IP-INF-053-0000-JCB-BOQ-CA-000007-B_Bill of Quantities (Priced).pdf",
        force_ocr=True,
    )
    assert "D549.2" in text
    assert meta.get("ocr_pages", 0) >= 1
    assert meta.get("ocr_skipped_too_large") is not True


def test_finer_chunk_keeps_priced_boq_item_codes():
    """D599.5 / D549.2 must survive extract → chunk so WAVE 2 B4/B5 can hit."""
    from app.core.doc_index import chunk_extracted_document, _chunker_for_document

    rows = [
        "D110 General site clearance 12 ha 2500.00 30000.00",
        "D290.1 Removal of trees in existing sidewalks 48 nr 220.00 10560.00",
        "D549.2 Removal of existing chain link fence 80 m 15.00 1200.00",
        "D599.5 Breaking out existing carriageway including road markings "
        "1200 m2 18.00 21600.00",
    ]
    # Repeat filler so the default 500-word splitter would bury the codes
    # if the BOQ filename failed to select the finer chunker.
    filler = " ".join(f"padding{i}" for i in range(400))
    text = filler + "\n" + "\n".join(rows) + "\n" + filler
    filename = "IP-INF-053-0000-JCB-BOQ-CA-000007-B_Bill of Quantities (Priced).pdf"
    chunks = chunk_extracted_document(
        text, chunker=_chunker_for_document(filename), filename=filename,
    )
    blob = "\n".join(chunks)
    assert "D599.5" in blob
    assert "D549.2" in blob
    # Finer chunks keep a code on a short window, not a 500-word blob.
    hit = next(c for c in chunks if "D599.5" in c)
    assert len(hit.split()) < 200


def test_extract_docx_mocked(tmp_path, monkeypatch):
    """DOCX extraction: monkeypatch docx.Document; assert joined paragraph text."""
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)

    doc_path = str(tmp_path / "contract.docx")
    file_crypto.write_document(doc_path, b"PK\x03\x04 fake docx bytes")

    class FakePara:
        def __init__(self, txt):
            self.text = txt

    class FakeCell:
        def __init__(self, txt):
            self.text = txt

    class FakeRow:
        def __init__(self, cells):
            self.cells = [FakeCell(c) for c in cells]

    class FakeTable:
        def __init__(self, rows):
            self.rows = [FakeRow(r) for r in rows]

    class FakeDocxDoc:
        paragraphs = [FakePara("Clause one text."), FakePara("Clause two text.")]
        # python-docx `.paragraphs` excludes table cells; the extractor must
        # pull both so BOQ/spec tables aren't dropped.
        tables = [FakeTable([["Item", "Qty"], ["Excavation", "100 m3"]])]

    import docx as real_docx
    monkeypatch.setattr(real_docx, "Document", lambda path: FakeDocxDoc())

    from app.core import doc_index
    importlib.reload(doc_index)
    text = doc_index.extract_document_text(doc_path, "contract.docx")
    assert "Clause one text." in text
    assert "Clause two text." in text
    # Table cells are extracted too (pipe-joined per row).
    assert "Excavation" in text and "100 m3" in text


def test_extract_xlsx_keeps_priced_boq_row_intact(tmp_path, monkeypatch):
    """A priced-BOQ .xlsx must extract row-wise, so each unit rate stays
    adjacent to its item description in one line. The old cell-wise flatten
    ("<all cells> ".join) scattered a rate away from its item and made company
    unit rates unretrievable."""
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    openpyxl = pytest.importorskip("openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Description", "Qty", "Unit", "Rate (SAR)", "Amount"])
    ws.append(["Ready-mix concrete C30", 500, "m3", 220, 110000])
    ws.append(["Rebar 12mm", 40, "tonne", 3000, 120000])
    raw_path = str(tmp_path / "_raw.xlsx")
    wb.save(raw_path)
    with open(raw_path, "rb") as fh:
        xlsx_bytes = fh.read()

    doc_path = str(tmp_path / "priced_boq.xlsx")
    file_crypto.write_document(doc_path, xlsx_bytes)

    from app.core import doc_index
    importlib.reload(doc_index)
    text = doc_index.extract_document_text(doc_path, "priced_boq.xlsx")

    # The concrete rate (220) must sit on the same line as its description.
    line = next(ln for ln in text.splitlines() if "Ready-mix concrete C30" in ln)
    assert "220" in line and "m3" in line
    # Rebar row is a distinct line — rows are not merged into one blob.
    assert "Rebar 12mm" not in line
    assert any("Rebar 12mm" in ln and "3000" in ln for ln in text.splitlines())


def test_extract_kmz_indexes_labels_not_coordinates(tmp_path, monkeypatch):
    """KMZ extraction indexes the human-meaningful <name>/<description> labels
    and drops the raw <coordinates> dump (which otherwise explodes into
    coordinate-noise chunks). Duplicate labels are de-duped."""
    import io
    import zipfile

    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)

    kml = (
        '<?xml version="1.0"?>\n'
        "<kml><Document>\n"
        "  <name>Site Boundary</name>\n"
        "  <Placemark>\n"
        "    <name>Manhole MH-12</name>\n"
        "    <description>Invert level 45.20</description>\n"
        "    <Point><coordinates>46.6,24.7,0 46.7,24.8,0</coordinates></Point>\n"
        "  </Placemark>\n"
        "  <Placemark>\n"
        "    <name>Manhole MH-12</name>\n"            # duplicate -> de-duped
        "    <description>-46.6,24.7</description>\n"  # pure coords -> dropped
        "  </Placemark>\n"
        "</Document></kml>\n"
    )
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("doc.kml", kml)
    doc_path = str(tmp_path / "survey.kmz")
    file_crypto.write_document(doc_path, buf.getvalue())

    from app.core import doc_index
    importlib.reload(doc_index)
    text = doc_index.extract_document_text(doc_path, "survey.kmz")

    assert "Site Boundary" in text
    assert "Manhole MH-12" in text
    assert "Invert level 45.20" in text
    assert text.count("Manhole MH-12") == 1        # de-duped
    assert "46.6,24.7,0" not in text               # raw coordinate dump excluded
    assert "-46.6,24.7" not in text                # pure-coordinate label dropped


def test_extract_xlsx_mocked(tmp_path, monkeypatch):
    """XLSX extraction: monkeypatch openpyxl.load_workbook; assert cell text."""
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)

    doc_path = str(tmp_path / "budget.xlsx")
    file_crypto.write_document(doc_path, b"PK\x03\x04 fake xlsx bytes")

    class FakeCell:
        def __init__(self, val):
            self.value = val

    class FakeSheet:
        def __iter__(self):
            return iter([
                [FakeCell("Activity"), FakeCell("Cost"), FakeCell(None)],
                [FakeCell("Excavation"), FakeCell(15000), FakeCell("")],
            ])

    class FakeWorkbook:
        sheetnames = ["Sheet1"]

        def __getitem__(self, name):
            return FakeSheet()

    import openpyxl as real_openpyxl
    monkeypatch.setattr(
        real_openpyxl, "load_workbook",
        lambda path, data_only=True: FakeWorkbook()
    )

    from app.core import doc_index
    importlib.reload(doc_index)
    text = doc_index.extract_document_text(doc_path, "budget.xlsx")
    assert "Activity" in text
    assert "Cost" in text
    assert "Excavation" in text


def test_extract_unsupported_returns_empty(tmp_path, monkeypatch):
    """Unsupported extension returns empty string, never raises.

    Stream F: images (.jpg/.png/...) are now SUPPORTED via OCR, so this uses a
    genuinely unsupported extension (.dwg) instead.
    """
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    from app.core.doc_index import extract_document_text

    doc_path = str(tmp_path / "drawing.dwg")
    file_crypto.write_document(doc_path, b"AutoCAD DWG binary data")

    result = extract_document_text(doc_path, "drawing.dwg")
    assert result == ""


def test_extract_encrypted_txt(tmp_path, monkeypatch):
    """Encrypted .txt file is transparently decrypted; plaintext is returned."""
    key = Fernet.generate_key().decode()
    monkeypatch.setenv("DATA_ENCRYPTION_KEY", key)

    doc_path = str(tmp_path / "secret.txt")
    plaintext = b"Confidential: the bridge budget is forty million dirhams."
    file_crypto.write_document(doc_path, plaintext)

    # Confirm it is actually encrypted on disk
    on_disk = open(doc_path, "rb").read()
    assert on_disk != plaintext

    from app.core import doc_index
    importlib.reload(doc_index)
    text = doc_index.extract_document_text(doc_path, "secret.txt")
    assert "forty million dirhams" in text


def test_extract_missing_file_returns_empty(monkeypatch):
    """Missing file path returns '' without raising."""
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    from app.core.doc_index import extract_document_text

    result = extract_document_text("/nonexistent/path/doc.txt", "doc.txt")
    assert result == ""


# ────────────────────────────────────────────────────────────────────────────
# WAVE 2 — Task 2: chunk_text
# ────────────────────────────────────────────────────────────────────────────

def _make_words(n: int) -> str:
    return " ".join(f"word{i}" for i in range(n))


def test_chunk_1200_words_gives_three_chunks():
    from app.core.doc_index import chunk_text
    text = _make_words(1200)
    chunks = chunk_text(text, words_per_chunk=500)
    assert len(chunks) == 3
    for ch in chunks:
        assert len(ch.split()) <= 500


def test_chunk_20_words_gives_one_chunk():
    from app.core.doc_index import chunk_text
    text = _make_words(20)
    chunks = chunk_text(text)
    assert len(chunks) == 1
    assert len(chunks[0].split()) == 20


def test_chunk_empty_string_gives_empty_list():
    from app.core.doc_index import chunk_text
    assert chunk_text("") == []


def test_chunk_whitespace_only_gives_empty_list():
    from app.core.doc_index import chunk_text
    assert chunk_text("   \n\t  ") == []


# ────────────────────────────────────────────────────────────────────────────
# WAVE 3 — Task 3: index build + JSON persistence
# ────────────────────────────────────────────────────────────────────────────

@pytest.fixture
def fresh_db(tmp_path, monkeypatch):
    """Isolated DATA_DIR + fresh projects DB for each test."""
    monkeypatch.setenv("DATA_DIR", str(tmp_path))
    monkeypatch.setattr(projects_mod, "_initialized", False)
    projects_mod.init_db()
    return tmp_path


def _write_txt_doc(tmp_path, filename, content_bytes):
    """Write a plaintext doc file under tmp_path, return its path."""
    p = str(tmp_path / filename)
    file_crypto.write_document(p, content_bytes)
    return p


def test_index_project_writes_file_and_returns_summary(fresh_db, tmp_path, monkeypatch):
    """index_project indexes supported docs and writes the JSON index file."""
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    from app.core import doc_index
    importlib.reload(doc_index)
    from app.core.rag import vector_store as _vs
    _vs.reset_store_cache()

    proj = projects_mod.create_project("Alpha Project")
    pid = proj["id"]

    content = b"The structural survey confirmed no subsidence was detected."
    doc_path = _write_txt_doc(tmp_path, "report.txt", content)
    projects_mod.add_document(pid, "report.txt", file_path=doc_path, size=len(content))

    result = doc_index.index_project(pid)

    assert result["project_id"] == pid
    assert result["indexed"] == 1
    assert result["skipped_unsupported"] == 0
    assert result["total_chunks"] >= 1

    # Persisted index
    saved = doc_index._load_index(pid)
    assert saved is not None
    assert saved["project_id"] == pid
    assert len(saved["documents"]) == 1
    assert "subsidence" in saved["documents"][0]["chunks"][0]


def test_index_project_skips_unsupported_type(fresh_db, tmp_path, monkeypatch):
    """A genuinely unsupported document lands in 'skipped', not 'documents'.

    Stream F: images are now OCR-able and SUPPORTED, so this uses .dwg — a
    type that remains unsupported.
    """
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    from app.core import doc_index
    importlib.reload(doc_index)
    from app.core.rag import vector_store as _vs
    _vs.reset_store_cache()

    proj = projects_mod.create_project("Beta Project")
    pid = proj["id"]

    img_path = str(tmp_path / "model.dwg")
    file_crypto.write_document(img_path, b"AutoCAD DWG")
    projects_mod.add_document(pid, "model.dwg", file_path=img_path, size=6)

    result = doc_index.index_project(pid)

    assert result["indexed"] == 0
    assert result["skipped_unsupported"] == 1

    import json
    saved = doc_index._load_index(pid)
    assert saved["documents"] == []
    assert len(saved["skipped"]) == 1
    assert saved["skipped"][0]["reason"] == "unsupported_type"


def test_index_document_indexes_ifc_census_not_unsupported(
    fresh_db, tmp_path, monkeypatch
):
    """Leftover live UI: sample_office.ifc sat at chunk_count=0 / Not indexed
    because .ifc was not a supported extractor type. Clash stays off."""
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    monkeypatch.setenv("RAG_EMBEDDING_MODEL", "fake")
    from app.core import doc_index
    importlib.reload(doc_index)
    from app.core.rag import vector_store as _vs
    _vs.reset_store_cache()

    proj = projects_mod.create_project("IFC Project")
    pid = proj["id"]
    sample = os.path.join(FIXTURES_DIR, "sample_office.ifc")
    content = open(sample, "rb").read()
    doc_path = _write_txt_doc(tmp_path, "sample_office.ifc", content)
    doc = projects_mod.add_document(
        pid, "sample_office.ifc", file_path=doc_path, size=len(content),
    )

    result = doc_index.index_document(pid, doc["id"])
    assert result.get("error") != "ZERO_CHUNK", result
    assert result.get("indexed") == 1, result
    saved = doc_index._load_index(pid)
    assert saved is not None
    chunks = saved["documents"][0]["chunks"]
    blob = "\n".join(chunks).lower()
    assert "ifc" in blob
    assert "wall" in blob
    skipped = saved.get("skipped") or []
    assert all(s.get("filename") != "sample_office.ifc" for s in skipped)

    from app.core.rag.embeddings import get_embedder
    from app.core.rag.vector_store import get_store
    counts = get_store(dim=get_embedder().dim).count_by_doc(pid)
    assert counts.get(doc["id"], 0) >= 1


def test_index_document_wires_boq_total_into_rag(fresh_db, tmp_path, monkeypatch):
    """A priced BOQ's total + line items become retrievable chunks so the
    platform can answer 'what is the total package value?' from the corpus.
    Wires app/blocks/boq_processor into index_document. The summed total
    (105000) is NOT present in the raw CSV text — only the BOQ wiring emits it.
    """
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    monkeypatch.setenv("RAG_EMBEDDING_MODEL", "fake")
    from app.core import doc_index
    importlib.reload(doc_index)
    from app.core.rag import vector_store as _vs
    _vs.reset_store_cache()

    proj = projects_mod.create_project("BOQ Project")
    pid = proj["id"]

    csv = (
        b"Description,Quantity,Rate,Amount\n"
        b"Excavation works,100,50,5000\n"
        b"Concrete grade 40,200,300,60000\n"
        b"Reinforcement steel,50,800,40000\n"
    )
    doc_path = _write_txt_doc(tmp_path, "Priced BOQ.csv", csv)
    doc = projects_mod.add_document(pid, "Priced BOQ.csv", file_path=doc_path, size=len(csv))

    doc_index.index_document(pid, doc["id"])

    saved = doc_index._load_index(pid)
    chunks = saved["documents"][0]["chunks"]
    blob = "\n".join(chunks).lower()
    assert "boq total" in blob, f"no BOQ summary chunk; chunks={chunks}"
    # 5000 + 60000 + 40000 = 105000 — appears only via the BOQ wiring.
    assert "105000" in blob or "105,000" in blob, f"total not wired; chunks={chunks}"
    # The total chunk must phrase itself the way users actually ask, so a
    # broad "what is the total package value?" matches it instead of a generic
    # contract template (the question that started this thread).
    assert "package value" in blob, f"total chunk should match 'package value'; chunks={chunks}"
    assert "total contract value" in blob, f"total chunk should match 'contract value'; chunks={chunks}"


def test_boq_line_item_chunk_carries_five_fields_and_source():
    """Each per-item chunk must carry all five fields operators ask about
    (description, total quantity, unit, unit price, total price) plus the source
    BOQ name, so 'unit price of X' / 'total quantity of X' retrieve cleanly and
    disambiguate between several BOQs in one project."""
    from app.core.doc_index import _boq_summary_chunks

    result = {
        "status": "success",
        "currency": "SAR",
        "total_cost": 525000,
        "source_name": "Al-Ostool Demolition BOQ",
        "line_items": [
            {
                "description": "300mm gravity sewer pipe",
                "quantity": 1250,
                "unit": "m",
                "unit_cost": 420,
                "total_cost": 525000,
            }
        ],
    }
    item_chunks = [
        c for c in _boq_summary_chunks(result)
        if c.lower().startswith("boq line item")
    ]
    assert item_chunks, "no per-item chunk emitted"
    c = item_chunks[0].lower()
    assert "al-ostool demolition boq" in c, f"source BOQ not named: {c}"
    assert "300mm gravity sewer pipe" in c
    assert "1250" in c and "m," in c, f"quantity+unit missing: {c}"
    assert "unit price" in c and "420" in c, f"unit price missing: {c}"
    assert "total price" in c and "525000" in c, f"total price missing: {c}"


def test_boq_no_software_aggregate_but_has_terminology_chunk():
    """No blanket software 'aggregate total' chunk — summing OCR'd line items is
    wrong (units are mixed and often mis-read: excavation/backfill/concrete are
    m3, pipe-laying is LM). But a sewer/waste-water BOQ MUST emit a terminology
    chunk so 'sewer' and 'waste water' queries match the same network."""
    from app.core.doc_index import _boq_summary_chunks

    result = {
        "status": "success", "currency": "SAR", "source_name": "the client project Sewer BOQ",
        "line_items": [
            {"description": "Waste water (foul/sewer) gravity pipe 200mm",
             "quantity": 7790, "unit": "m", "unit_cost": 10, "total_cost": 77900},
        ],
    }
    chunks = _boq_summary_chunks(result)
    # No software aggregate / summed-total chunk.
    assert not [c for c in chunks if c.lower().startswith("boq total quantity")]
    # Terminology alias IS present and bridges sewer <-> waste water.
    term = [c for c in chunks if c.startswith("TERMINOLOGY")]
    assert term, "no terminology chunk for a sewer/waste-water BOQ"
    t = term[0].lower()
    assert "sewer" in t and "waste water" in t


def test_index_document_boq_total_hedged_when_pages_skipped(fresh_db, tmp_path, monkeypatch):
    """Accuracy guard: if the BOQ parse skipped pages, the chunk must say the
    total is PARTIAL, never a confident number (no-assumptions rule)."""
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    monkeypatch.setenv("RAG_EMBEDDING_MODEL", "fake")
    from app.core import doc_index
    importlib.reload(doc_index)
    from app.core.rag import vector_store as _vs
    _vs.reset_store_cache()

    # Force the BOQ processor to report a skipped page.
    import app.blocks.boq_processor as boq
    async def _fake_process(self, input_data, params=None):
        return {
            "status": "success", "item_count": 2, "total_cost": 105000.0,
            "currency": "USD", "line_items": [
                {"description": "Excavation", "total_cost": 5000.0, "section": "Civil"},
                {"description": "Concrete", "total_cost": 100000.0, "section": "Civil"},
            ],
            "cost_breakdown": {"Civil": {"total": 105000.0, "percentage": 100.0}},
            "pages_skipped": 2,
        }
    monkeypatch.setattr(boq.BOQProcessorBlock, "process", _fake_process)

    proj = projects_mod.create_project("Partial BOQ")
    pid = proj["id"]
    csv = b"Description,Amount\nExcavation,5000\n"
    doc_path = _write_txt_doc(tmp_path, "scanned BOQ.pdf", csv)  # .pdf -> boq path
    doc = projects_mod.add_document(pid, "scanned BOQ.pdf", file_path=doc_path, size=len(csv))

    doc_index.index_document(pid, doc["id"])
    saved = doc_index._load_index(pid)
    blob = "\n".join(saved["documents"][0]["chunks"]).lower()
    assert "partial" in blob, f"pages_skipped>0 must hedge; chunks missing 'partial': {blob[:300]}"


def test_boq_named_pdf_without_computable_total_emits_guard(fresh_db, tmp_path, monkeypatch):
    """A BOQ-named PDF that boq_processor CANNOT total (scanned/unparseable)
    must emit a guard chunk telling the model NOT to state a total — so it
    can't synthesise a false 'total package value' from partial OCR text
    (the demolition-BOQ failure mode: 8 chunks of a 675-page scan)."""
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    monkeypatch.setenv("RAG_EMBEDDING_MODEL", "fake")
    from app.core import doc_index
    importlib.reload(doc_index)
    from app.core.rag import vector_store as _vs
    _vs.reset_store_cache()
    import app.blocks.boq_processor as boq

    async def _no_total(self, input_data, params=None):
        return {"status": "error", "error": "no tables found (scanned image PDF)"}
    monkeypatch.setattr(boq.BOQProcessorBlock, "process", _no_total)

    proj = projects_mod.create_project("Scanned BOQ")
    pid = proj["id"]
    doc_path = _write_txt_doc(tmp_path, "Demolition BOQ.pdf", b"scanned image content")
    doc = projects_mod.add_document(pid, "Demolition BOQ.pdf", file_path=doc_path, size=21)

    doc_index.index_document(pid, doc["id"])
    blob = "\n".join(doc_index._load_index(pid)["documents"][0]["chunks"]).lower()
    assert "do not state a total" in blob, f"missing no-total guard; chunks={blob[:200]}"


def test_index_project_empty_project(fresh_db, monkeypatch):
    """Empty project produces a valid index file with no documents."""
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    from app.core import doc_index
    importlib.reload(doc_index)
    from app.core.rag import vector_store as _vs
    _vs.reset_store_cache()

    proj = projects_mod.create_project("Empty Project")
    pid = proj["id"]

    result = doc_index.index_project(pid)

    assert result["indexed"] == 0
    assert result["total_chunks"] == 0

    import json
    saved = doc_index._load_index(pid)
    assert saved["documents"] == []
    assert saved["skipped"] == []


def test_index_document_incremental(fresh_db, tmp_path, monkeypatch):
    """index_document adds to an existing index without discarding prior docs."""
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    from app.core import doc_index
    importlib.reload(doc_index)
    from app.core.rag import vector_store as _vs
    _vs.reset_store_cache()

    proj = projects_mod.create_project("Gamma Project")
    pid = proj["id"]

    # Doc A — indexed via index_project first
    content_a = b"Foundation works are complete as per the structural report."
    doc_path_a = _write_txt_doc(tmp_path, "baseline.txt", content_a)
    doc_a = projects_mod.add_document(pid, "baseline.txt", file_path=doc_path_a, size=len(content_a))
    doc_index.index_project(pid)  # establishes doc A in the index

    # Doc B — added afterwards, indexed incrementally
    content_b = b"Concrete pour completed on schedule per daily report."
    doc_path_b = _write_txt_doc(tmp_path, "daily.txt", content_b)
    doc_b = projects_mod.add_document(pid, "daily.txt", file_path=doc_path_b, size=len(content_b))
    did_b = doc_b["id"]

    result = doc_index.index_document(pid, did_b)

    assert result["indexed"] == 1

    import json
    saved = doc_index._load_index(pid)
    ids = [d["document_id"] for d in saved["documents"]]
    # Both docs must be present — proves load-modify-write preserved doc A
    assert doc_a["id"] in ids
    assert did_b in ids


def test_invalidate_project_removes_index(fresh_db, tmp_path, monkeypatch):
    """invalidate_project deletes the index file."""
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    from app.core import doc_index
    importlib.reload(doc_index)
    from app.core.rag import vector_store as _vs
    _vs.reset_store_cache()

    proj = projects_mod.create_project("Delta Project")
    pid = proj["id"]
    doc_index.index_project(pid)  # creates the file

    assert doc_index._load_index(pid) is not None
    doc_index.invalidate_project(pid)
    assert doc_index._load_index(pid) is None


def test_load_index_returns_none_when_absent(fresh_db, monkeypatch):
    """_load_index returns None when no index file exists."""
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    from app.core import doc_index
    importlib.reload(doc_index)

    result = doc_index._load_index("nonexistent-project-id")
    assert result is None


def test_fingerprint_format(fresh_db, tmp_path, monkeypatch):
    """Each indexed document entry has a fingerprint of the form 'uploaded_at:size'."""
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    from app.core import doc_index
    importlib.reload(doc_index)
    from app.core.rag import vector_store as _vs
    _vs.reset_store_cache()

    proj = projects_mod.create_project("Fingerprint Project")
    pid = proj["id"]
    content = b"Fingerprint test document content here."
    doc_path = _write_txt_doc(tmp_path, "fp.txt", content)
    doc = projects_mod.add_document(pid, "fp.txt", file_path=doc_path, size=len(content))

    doc_index.index_project(pid)

    import json
    saved = doc_index._load_index(pid)
    entry = saved["documents"][0]
    expected_fp = f"{doc['uploaded_at']}:{doc['size']}"
    assert entry["fingerprint"] == expected_fp


# ────────────────────────────────────────────────────────────────────────────
# WAVE 4 — Phase C2: search_project_documents
# ────────────────────────────────────────────────────────────────────────────

CONCRETE_TEXT = (
    b"Concrete pour and curing schedule: the contractor must ensure the mix "
    b"reaches the specified compressive strength before formwork is removed. "
    b"The curing time for standard Portland cement concrete is 28 days. "
    b"Water-cement ratio must be controlled. Slump tests are performed on site. "
    b"The concrete curing process requires consistent moisture and temperature."
)

ELECTRICAL_TEXT = (
    b"Electrical wiring and conduit inspection: all cable runs shall be "
    b"installed in rigid conduit per the approved shop drawings. The inspector "
    b"must verify continuity, grounding, and insulation resistance. "
    b"Circuit breaker panels must be clearly labelled. "
    b"Junction boxes require accessible covers for future maintenance."
)

LANDSCAPING_TEXT = (
    b"Landscaping and irrigation layout: the planting plan shows trees, "
    b"shrubs, and ground cover arranged along the building perimeter. "
    b"Drip irrigation lines are routed to each planted zone. "
    b"Soil preparation includes aeration and compost amendment. "
    b"Irrigation controllers are programmed for seasonal watering schedules."
)


@pytest.fixture
def search_project(tmp_path, monkeypatch):
    """Shared fixture: isolated DATA_DIR, fresh DB, project with 3 disjoint docs."""
    monkeypatch.setenv("DATA_DIR", str(tmp_path))
    monkeypatch.setenv("RAG_EMBEDDING_MODEL", "fake")
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    # Disable GK merging: on PostgreSQL the training_material project seeded
    # by other tests persists in the shared DB and would corrupt rankings.
    monkeypatch.setenv("RAG_GENERAL_KNOWLEDGE_PROJECTS", "")
    monkeypatch.setattr(projects_mod, "_initialized", False)
    projects_mod.init_db()

    from app.core import doc_index
    importlib.reload(doc_index)

    # PR #94: search_project_documents now uses the hybrid retriever which
    # caches its VectorStore per (db_url, dim) at module scope. Tests swap
    # DATA_DIR per-test, so the cache must be reset to repoint the store
    # at the new SQLite file, otherwise queries hit the previous test's DB.
    from app.core.rag import vector_store as _vs
    from app.core.rag import embeddings as _emb
    _vs.reset_store_cache()
    _emb.reset_embedder_cache()

    proj = projects_mod.create_project("Search Test Project")
    pid = proj["id"]

    def _add(filename, content):
        p = str(tmp_path / filename)
        file_crypto.write_document(p, content)
        return projects_mod.add_document(pid, filename, file_path=p, size=len(content))

    doc_concrete = _add("concrete.txt", CONCRETE_TEXT)
    doc_elec = _add("electrical.txt", ELECTRICAL_TEXT)
    doc_landscape = _add("landscaping.txt", LANDSCAPING_TEXT)

    return {
        "pid": pid,
        "doc_concrete": doc_concrete,
        "doc_elec": doc_elec,
        "doc_landscape": doc_landscape,
        "tmp_path": tmp_path,
        "doc_index": doc_index,
    }


@pytest.mark.asyncio
async def test_search_ranks_relevant_doc_first(search_project):
    """Query 'concrete curing time' should rank the concrete document first."""
    doc_index = search_project["doc_index"]
    pid = search_project["pid"]
    concrete_id = search_project["doc_concrete"]["id"]

    results = await doc_index.search_project_documents(pid, "concrete curing time")

    assert len(results) >= 1
    assert results[0]["document_id"] == concrete_id


@pytest.mark.asyncio
async def test_search_returns_shape(search_project):
    """Each result has document_id, filename, snippet, score; score is a float."""
    doc_index = search_project["doc_index"]
    pid = search_project["pid"]

    results = await doc_index.search_project_documents(pid, "concrete curing time")

    assert len(results) >= 1
    for r in results:
        assert "document_id" in r
        assert "filename" in r
        assert "snippet" in r
        assert "score" in r
        assert isinstance(r["score"], float)


@pytest.mark.asyncio
async def test_search_top_k(search_project):
    """top_k=1 returns exactly 1 result."""
    doc_index = search_project["doc_index"]
    pid = search_project["pid"]

    results = await doc_index.search_project_documents(pid, "concrete curing time", top_k=1)

    assert len(results) == 1


@pytest.mark.asyncio
async def test_search_empty_project(tmp_path, monkeypatch):
    """A project with no documents returns [] without raising."""
    monkeypatch.setenv("DATA_DIR", str(tmp_path))
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    # Disable GK so training_material chunks from other tests don't leak in.
    monkeypatch.setenv("RAG_GENERAL_KNOWLEDGE_PROJECTS", "")
    monkeypatch.setattr(projects_mod, "_initialized", False)
    projects_mod.init_db()

    from app.core import doc_index
    importlib.reload(doc_index)
    from app.core.rag import vector_store as _vs
    _vs.reset_store_cache()

    proj = projects_mod.create_project("Empty Search Project")
    pid = proj["id"]

    results = await doc_index.search_project_documents(pid, "anything")
    assert results == []


@pytest.mark.asyncio
async def test_search_builds_index_lazily(tmp_path, monkeypatch):
    """Index file must not exist before search; must exist afterwards."""
    monkeypatch.setenv("DATA_DIR", str(tmp_path))
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    # Disable GK so training_material chunks from other tests don't cause
    # search_project_documents to skip the lazy-bootstrap branch.
    monkeypatch.setenv("RAG_GENERAL_KNOWLEDGE_PROJECTS", "")
    monkeypatch.setattr(projects_mod, "_initialized", False)
    projects_mod.init_db()

    from app.core import doc_index
    importlib.reload(doc_index)
    from app.core.rag import vector_store as _vs
    _vs.reset_store_cache()

    proj = projects_mod.create_project("Lazy Build Project")
    pid = proj["id"]

    p = str(tmp_path / "lazy.txt")
    file_crypto.write_document(p, b"Lazy build test content about concrete pouring.")
    projects_mod.add_document(pid, "lazy.txt", file_path=p, size=47)

    # Index must not yet exist
    assert doc_index._load_index(pid) is None

    results = await doc_index.search_project_documents(pid, "concrete")

    # After search, index exists and results are returned
    assert doc_index._load_index(pid) is not None
    assert isinstance(results, list)


@pytest.mark.xfail(
    strict=False,
    reason="GK crowds project docs out of top-5 - tracked by TASK H knobs (RAG_AUDIT_V2)",
)
@pytest.mark.asyncio
async def test_search_uses_hybrid_retriever(search_project):
    """PR #94: search_project_documents must query the same chunks table
    the RAG injection layer queries. Verify by seeding the chunks table
    directly and confirming the function returns results without going
    through the legacy JSON index path.
    """
    doc_index = search_project["doc_index"]
    pid = search_project["pid"]
    concrete_id = search_project["doc_concrete"]["id"]

    # Force a search — this triggers the bootstrap which writes to the
    # chunks table. Then delete the legacy JSON index file to prove the
    # next call goes through the hybrid retriever (which reads the
    # chunks table, not the JSON blob).
    results1 = await doc_index.search_project_documents(pid, "concrete curing")
    assert len(results1) >= 1
    assert results1[0]["document_id"] == concrete_id

    # Wipe the legacy JSON index row but leave the chunks table populated.
    from app.core.db import SessionLocal
    from app.core.models import DocIndex
    from sqlalchemy import delete as _sql_delete
    with SessionLocal() as s:
        s.execute(_sql_delete(DocIndex).where(DocIndex.project_id == pid))
        s.commit()
    assert doc_index._load_index(pid) is None

    # Second call: bootstrap would re-fire (since JSON is gone), but the
    # retriever should still find the chunks already in the chunks
    # table. The contract here: results come from the chunks table,
    # not the JSON blob.
    results2 = await doc_index.search_project_documents(pid, "concrete curing")
    assert len(results2) >= 1
    assert results2[0]["document_id"] == concrete_id


@pytest.mark.asyncio
async def test_search_excludes_deleted_document(tmp_path, monkeypatch):
    """After deleting a document, its content must not appear in search results."""
    monkeypatch.setenv("DATA_DIR", str(tmp_path))
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    monkeypatch.setattr(projects_mod, "_initialized", False)
    projects_mod.init_db()

    from app.core import doc_index
    importlib.reload(doc_index)
    from app.core.rag import vector_store as _vs
    _vs.reset_store_cache()

    proj = projects_mod.create_project("Delete Test Project")
    pid = proj["id"]

    p1 = str(tmp_path / "concrete2.txt")
    file_crypto.write_document(p1, CONCRETE_TEXT)
    doc_keep = projects_mod.add_document(pid, "concrete2.txt", file_path=p1, size=len(CONCRETE_TEXT))

    p2 = str(tmp_path / "electrical2.txt")
    file_crypto.write_document(p2, ELECTRICAL_TEXT)
    doc_del = projects_mod.add_document(pid, "electrical2.txt", file_path=p2, size=len(ELECTRICAL_TEXT))

    # Build the index by searching once
    await doc_index.search_project_documents(pid, "concrete")

    # Delete the electrical document from the DB
    projects_mod.delete_document(doc_del["id"])

    # Search again — deleted doc must not appear in results
    results = await doc_index.search_project_documents(pid, "electrical wiring conduit")
    returned_ids = [r["document_id"] for r in results]
    assert doc_del["id"] not in returned_ids


def test_zero_chunk_indexing_emits_warning(tmp_path, monkeypatch, caplog):
    """A supported file that extracts to zero chunks must WARN with the
    grep-able ZERO_CHUNK marker (TASK C3): three whole projects sat
    silently unretrievable in this state before the 2026-06 audit."""
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    monkeypatch.setenv("RAG_EMBEDDING_MODEL", "fake")
    monkeypatch.setenv("RAG_VECTOR_NAMESPACE", "")
    from app.core import doc_index
    importlib.reload(doc_index)
    projects_mod.init_db()

    proj = projects_mod.create_project("Zero Chunk Project")
    pid = proj["id"]

    doc_path = _write_txt_doc(tmp_path, "empty.txt", b"")
    doc = projects_mod.add_document(pid, "empty.txt", file_path=doc_path, size=0)

    import logging
    with caplog.at_level(logging.WARNING, logger="app.core.doc_index"):
        result = doc_index.index_document(pid, doc["id"])

    assert result["status"] == "error"
    assert result["error"] == "ZERO_CHUNK"
    assert any("ZERO_CHUNK" in r.message for r in caplog.records), (
        f"no ZERO_CHUNK warning; records={[r.message for r in caplog.records]}"
    )
    marker = next(r.message for r in caplog.records if "ZERO_CHUNK" in r.message)
    assert pid in marker and doc["id"] in marker


def test_nonempty_indexing_does_not_emit_zero_chunk_warning(tmp_path, monkeypatch, caplog):
    """The ZERO_CHUNK marker must stay quiet for normal documents, or the
    log signal drowns in noise and stops being a tripwire."""
    monkeypatch.delenv("DATA_ENCRYPTION_KEY", raising=False)
    monkeypatch.setenv("RAG_EMBEDDING_MODEL", "fake")
    monkeypatch.setenv("RAG_VECTOR_NAMESPACE", "")
    from app.core import doc_index
    importlib.reload(doc_index)
    projects_mod.init_db()

    proj = projects_mod.create_project("Normal Project")
    pid = proj["id"]

    content = b"Concrete grade C40 shall be used for all substructure elements."
    doc_path = _write_txt_doc(tmp_path, "spec.txt", content)
    doc = projects_mod.add_document(pid, "spec.txt", file_path=doc_path, size=len(content))

    import logging
    with caplog.at_level(logging.WARNING, logger="app.core.doc_index"):
        doc_index.index_document(pid, doc["id"])

    assert not any("ZERO_CHUNK" in r.message for r in caplog.records)


# ── pdfplumber page-count gate ─────────────────────────────────────────────
# Vol 2 Specification parts 4 and 8 are plain A4 under the 25 MB size gate and
# under the A3 page-size gate, but 738 and 387 pages. Measured 2026-08-23 on
# those exact files, table extraction is ~97% of peak memory:
#
#     part 4 (738 pp)   tables on: 4191 MB / 173s     off: 122 MB / 10s
#     part 8 (387 pp)   tables on: 2094 MB /  93s     off: 154 MB /  6s
#
# 4191 MB is what OOM-killed the 4 GB instance before extraction was isolated,
# and what still breaches the isolated child's budget today.

class _FakeRect:
    def __init__(self, w=595.0, h=842.0):  # A4 portrait, in points
        self.width, self.height = w, h


class _FakePage:
    def __init__(self, w=595.0, h=842.0):
        self.rect = _FakeRect(w, h)


class _FakeDoc:
    """Minimal stand-in for a fitz document used as a context manager."""

    def __init__(self, pages: int, w=595.0, h=842.0):
        self.page_count = pages
        self._pages = [_FakePage(w, h) for _ in range(min(pages, 5))]

    def __iter__(self):
        return iter(self._pages)

    def __enter__(self):
        return self

    def __exit__(self, *exc):
        return False


def _fake_pdf(monkeypatch, pages: int, size_mb: float = 8.0, w=595.0, h=842.0):
    import fitz
    monkeypatch.setattr("os.path.getsize", lambda p: int(size_mb * 1024 * 1024))
    monkeypatch.setattr(fitz, "open", lambda *a, **k: _FakeDoc(pages, w, h))


def test_tables_skipped_for_a_long_ordinary_document(monkeypatch):
    """THE FIX: page count is a cost proxy the size and dimension gates miss."""
    from app.core import doc_index
    _fake_pdf(monkeypatch, pages=738, size_mb=7.5)   # Vol 2 part 4
    assert doc_index._pdf_tables_enabled("spec_4_of_9.pdf") is False

    _fake_pdf(monkeypatch, pages=387, size_mb=12.8)  # Vol 2 part 8
    assert doc_index._pdf_tables_enabled("spec_8_of_9.pdf") is False


def test_tables_kept_for_a_normal_length_document(monkeypatch):
    """The control. Digital table BOQs are what this feature exists for and
    they are tens of pages — the gate must not touch them."""
    from app.core import doc_index
    _fake_pdf(monkeypatch, pages=40, size_mb=2.0)
    assert doc_index._pdf_tables_enabled("boq.pdf") is True


def test_page_gate_is_tunable_and_disablable(monkeypatch):
    from app.core import doc_index
    _fake_pdf(monkeypatch, pages=300, size_mb=5.0)

    monkeypatch.setenv("PDF_TABLES_MAX_PAGES", "500")
    assert doc_index._pdf_tables_enabled("mid.pdf") is True, "raised bar allows it"

    monkeypatch.setenv("PDF_TABLES_MAX_PAGES", "100")
    assert doc_index._pdf_tables_enabled("mid.pdf") is False, "lowered bar blocks it"

    monkeypatch.setenv("PDF_TABLES_MAX_PAGES", "0")
    assert doc_index._pdf_tables_enabled("mid.pdf") is True, "0 disables the gate"


def test_existing_gates_still_win(monkeypatch):
    """The page gate is additional, not a replacement: an oversized file and an
    A1 drawing sheet must still be rejected even at a modest page count."""
    from app.core import doc_index
    _fake_pdf(monkeypatch, pages=10, size_mb=60.0)
    assert doc_index._pdf_tables_enabled("big.pdf") is False, "size gate"

    _fake_pdf(monkeypatch, pages=10, size_mb=5.0, w=2384.0, h=1684.0)  # A1
    assert doc_index._pdf_tables_enabled("drawing.pdf") is False, "page-size gate"
