"""W9 — export endpoint round-trip smoke.

Proves the /export/evm, /export/cost-boq and /export/cost-schedule endpoints
return a downloadable .xlsx that actually OPENS (openpyxl.load_workbook) with the
expected sheets and LIVE formulas — the gap the CLOSING_LEDGER flagged (libs were
numerically proven, but the endpoint->file round-trip was never smoke-tested).
Mirrors tests/test_schedule_export_endpoint.py.
"""
from __future__ import annotations

import io

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.main import app

H = {"Authorization": "Bearer cb_dev_key"}
_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture(scope="module")
def client():
    with TestClient(app) as c:
        yield c


def _pid(client):
    r = client.post("/v1/projects", json={"name": "Export Smoke", "client": "ACME"}, headers=H)
    assert r.status_code == 201, r.text
    return r.json()["id"]


def _has_formula(wb) -> bool:
    """True if any cell in any sheet is a live formula (starts with '=')."""
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    return True
    return False


def test_export_evm_opens_and_is_formula_linked(client):
    pid = _pid(client)
    r = client.post(
        f"/v1/projects/{pid}/export/evm",
        json={"bac": 1_000_000, "currency": "SAR", "periods": [
            {"period": "M1", "pv": 200_000, "ev": 180_000, "ac": 190_000},
            {"period": "M2", "pv": 450_000, "ev": 400_000, "ac": 430_000},
        ]},
        headers=H,
    )
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == _XLSX_MEDIA
    assert r.content[:2] == b"PK"                       # valid xlsx (zip) magic
    wb = openpyxl.load_workbook(io.BytesIO(r.content))  # actually opens
    assert wb.worksheets and _has_formula(wb), "EVM workbook must carry live formulas (CV/SV/CPI/SPI)"


def test_export_cost_boq_opens_with_qty_rate_formulas(client):
    pid = _pid(client)
    r = client.post(
        f"/v1/projects/{pid}/export/cost-boq",
        json={"project_name": "Export Smoke", "currency": "SAR", "ingest": False,
              "categories": [
                  {"name": "Site Works", "items": [
                      {"item_no": "A.1", "description": "Site clearing", "unit": "Lot", "qty": 1, "rate": 120000},
                      {"item_no": "A.2", "description": "Topsoil stripping", "unit": "sqm", "qty": 15000, "rate": 15},
                  ]},
              ]},
        headers=H,
    )
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == _XLSX_MEDIA
    assert r.content[:2] == b"PK"
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert "BOQ_Detail" in wb.sheetnames
    assert _has_formula(wb), "line-item amounts must be live =Qty*Rate formulas"


def test_export_cost_schedule_opens(client):
    pid = _pid(client)
    r = client.post(
        f"/v1/projects/{pid}/export/cost-schedule",
        json={"project_name": "Export Smoke", "currency": "SAR",
              "activities": [
                  {"id": "1", "wbs": "1", "name": "Mobilization", "duration": 10,
                   "predecessors": [], "cost": 50000, "manpower": 8},
                  {"id": "2", "wbs": "2", "name": "Earthworks", "duration": 20,
                   "predecessors": ["1"], "cost": 200000, "manpower": 16},
              ]},
        headers=H,
    )
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == _XLSX_MEDIA
    assert r.content[:2] == b"PK"
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert wb.worksheets, "cost-loaded schedule workbook must open with at least one sheet"
