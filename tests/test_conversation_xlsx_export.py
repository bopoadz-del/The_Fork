"""Conversation message -> xlsx export (Document Export Layer).

The xlsx format exists for table-shaped answers (cost/BOQ breakdowns arrive as
markdown pipe-tables): each markdown table becomes a real worksheet with a
bold header row; prose lands on a Message sheet. PDF remains an honest 501
(operator decision 2026-07-24: docx + xlsx cover the deliverables).

The endpoint dispatch is tested too — a renderer with no endpoint wiring is
exactly the seam class the 2026-07-24 sweep exists to prevent.
"""

from __future__ import annotations

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app
from app.dependencies import require_user
from app.routers.exports import _parse_markdown_tables, _render_message_xlsx


SAMPLE = """Here is the cost summary you asked for.

| Item | Qty | Rate | Amount |
| --- | --- | --- | --- |
| Excavation | 1,200 | 45 | 54,000 |
| Backfill | 800 | 30 | 24,000 |

Notes: rates are from the company priced BOQ.

| Milestone | Date |
| --- | --- |
| Mobilisation | 2026-08-01 |
"""


class TestParseMarkdownTables:
    def test_extracts_tables_and_prose(self):
        tables, prose = _parse_markdown_tables(SAMPLE)
        assert len(tables) == 2
        assert tables[0][0] == ["Item", "Qty", "Rate", "Amount"]
        assert tables[0][1] == ["Excavation", "1,200", "45", "54,000"]
        assert tables[1][0] == ["Milestone", "Date"]
        joined = " ".join(prose)
        assert "cost summary" in joined
        assert "company priced BOQ" in joined
        # Separator rows (---) never leak into data.
        assert not any("---" in c for row in tables[0] for c in row)

    def test_no_tables_returns_prose_only(self):
        tables, prose = _parse_markdown_tables("Just a plain answer.")
        assert tables == []
        assert prose == ["Just a plain answer."]


class TestRenderMessageXlsx:
    def test_tables_become_worksheets(self, tmp_path):
        path = _render_message_xlsx("Project X", SAMPLE, "conv12345678")
        wb = load_workbook(path)
        assert "Table 1" in wb.sheetnames and "Table 2" in wb.sheetnames
        t1 = wb["Table 1"]
        assert t1.cell(row=1, column=1).value == "Item"
        assert t1.cell(row=1, column=1).font.bold
        assert t1.cell(row=2, column=1).value == "Excavation"
        assert t1.cell(row=3, column=4).value == "24,000"
        t2 = wb["Table 2"]
        assert t2.cell(row=2, column=2).value == "2026-08-01"

    def test_prose_lands_on_message_sheet(self):
        path = _render_message_xlsx("Project X", SAMPLE, "conv12345678")
        wb = load_workbook(path)
        ws = wb["Message"]
        texts = [c.value for row in ws.iter_rows() for c in row if c.value]
        assert any("cost summary" in str(t) for t in texts)

    def test_plain_prose_message_still_exports(self):
        path = _render_message_xlsx("P", "No tables here at all.", "conv12345678")
        wb = load_workbook(path)
        assert wb.sheetnames == ["Message"]


class TestExportEndpointFormatDispatch:
    """The endpoint must actually route each format to its renderer."""

    @pytest.fixture(autouse=True)
    def _wired(self, monkeypatch):
        from app.routers import exports as mod

        app.dependency_overrides[require_user] = lambda: {
            "user_id": "u1", "role": "user",
        }
        monkeypatch.setattr(mod, "_check_owner", lambda *_a: {"name": "Proj"})
        monkeypatch.setattr(
            mod.agent_memory, "get_messages",
            lambda *_a, **_k: [{"role": "assistant", "content": SAMPLE}],
        )
        yield
        app.dependency_overrides.clear()

    def _post(self, fmt: str):
        with TestClient(app) as client:
            return client.post(
                f"/v1/projects/p1/conversations/conv12345678/export?format={fmt}"
            )

    def test_xlsx_returns_workbook(self):
        res = self._post("xlsx")
        assert res.status_code == 200
        assert "spreadsheetml" in res.headers["content-type"]
        assert res.headers["content-disposition"].endswith('.xlsx"')

    def test_docx_still_works(self):
        res = self._post("docx")
        assert res.status_code == 200
        assert "wordprocessingml" in res.headers["content-type"]

    def test_pdf_is_501_by_decision(self):
        res = self._post("pdf")
        assert res.status_code == 501
        assert "docx" in res.json()["detail"]
