"""The schedule-from-brief export endpoint backs the chat 'Schedule (Excel)'
download offer: it runs the deterministic generate_wbs, bridges it, and returns
a cost-loaded workbook (CPM + man-days S-curve + manpower histogram + milestones).
"""
from __future__ import annotations

import io

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.main import app

H = {"Authorization": "Bearer cb_dev_key"}
_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture(scope="module")
def client():
    with TestClient(app) as c:
        yield c


def _pid(client):
    r = client.post("/v1/projects", json={"name": "Sched Export", "client": "ACME"}, headers=H)
    assert r.status_code == 201, r.text
    return r.json()["id"]


def test_schedule_from_brief_returns_valid_workbook(client):
    pid = _pid(client)
    r = client.post(
        f"/v1/projects/{pid}/export/schedule-from-brief",
        json={"brief": "hyperscale data center", "target_count": 40,
              "project_type": "data_center", "start_date": "2026-01-05"},
        headers=H,
    )
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == _XLSX_MEDIA
    assert r.content[:2] == b"PK"
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert {"L2 Schedule", "Cost Loading", "Manpower Histogram", "Milestones", "Summary"} <= set(wb.sheetnames)
    # honest S-curve = a line chart on the progress-baseline sheet
    from openpyxl.chart import LineChart
    assert any(isinstance(c, LineChart) for c in wb["Cost Loading"]._charts)
    # milestones are dated when a start_date is supplied
    ms_txt = "\n".join(str(c.value) for row in wb["Milestones"].iter_rows()
                       for c in row if c.value is not None)
    assert "PROJECT COMPLETION" in ms_txt and "2026-" in ms_txt


def test_no_day_rate_means_no_cost_column(client):
    pid = _pid(client)
    r = client.post(
        f"/v1/projects/{pid}/export/schedule-from-brief",
        json={"brief": "data center", "target_count": 30, "project_type": "data_center"},
        headers=H,
    )
    assert r.status_code == 200, r.text
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    summ = "\n".join(str(c.value) for row in wb["Summary"].iter_rows()
                     for c in row if c.value is not None)
    assert "Total Cost" not in summ and "Total Man-days" in summ


def test_schedule_from_document_injects_extracted_lead_times(client, tmp_path):
    """A BOD/RFP with real equipment lead times drives long-lead procurement +
    target milestones in the workbook (the document_engine -> schedule wiring)."""
    from docx import Document
    from app.core import projects as store

    pid = _pid(client)
    doc_path = tmp_path / "BOD.docx"
    d = Document()
    for line in [
        "Anthropic Data Center Basis of Design.",
        "The standby generator SHALL have a 200 day manufacturing lead time.",
        "Chillers require 150 days lead time before installation.",
        "MV switchgear has a 120 day manufacturing lead time.",
        "Initial capacity operational by December 2027.",
    ]:
        d.add_paragraph(line)
    d.save(str(doc_path))
    did = store.add_document(pid, "BOD.docx", doc_path.name, str(doc_path))["id"]

    r = client.post(
        f"/v1/projects/{pid}/export/schedule-from-document",
        json={"document_ids": [did], "target_count": 80,
              "project_type": "data_center", "start_date": "2026-01-05"},
        headers=H,
    )
    assert r.status_code == 200, r.text
    # real lead times extracted + injected (generator/chiller/switchgear)
    assert int(r.headers.get("X-Procurement-Items", "0")) >= 2
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    l2 = "\n".join(str(c.value) for row in wb["L2 Schedule"].iter_rows()
                   for c in row if c.value is not None)
    assert "long-lead" in l2.lower(), "procurement activities must appear in the L2 schedule"
