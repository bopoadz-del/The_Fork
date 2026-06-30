"""The platform GENERATES cost-loaded L2 schedules and EVM workbooks —
formula-linked, matching the Anthropic/Kenya/Canada L2 exemplars.

Construction is cost: the schedule carries cost per activity + a cumulative
cost baseline (S-curve), and EVM derives CPI/SPI/EAC as LIVE formulas from
PV/EV/AC. Nothing pre-computed and pasted.
"""
from __future__ import annotations

import io

# 3 activities, finish-to-start chain A->B->C, with cost + manpower.
ACTIVITIES = [
    {"id": "1", "wbs": "1.1", "name": "Mobilization", "duration": 5, "predecessors": [], "cost": 100000, "manpower": 10},
    {"id": "2", "wbs": "1.2", "name": "Earthworks", "duration": 10, "predecessors": ["1"], "cost": 400000, "manpower": 25},
    {"id": "3", "wbs": "1.3", "name": "Foundations", "duration": 8, "predecessors": ["2"], "cost": 300000, "manpower": 20},
]


def _formulas(ws):
    return [c.value for row in ws.iter_rows() for c in row
            if isinstance(c.value, str) and c.value.startswith("=")]


def test_schedule_computes_cpm_and_critical_path():
    from app.lib.pm_excel import generate_cost_loaded_schedule
    wb = generate_cost_loaded_schedule({"project": "Test DC"}, ACTIVITIES)
    assert {"L2 Schedule", "Cost Loading", "Manpower Histogram", "Summary"} <= set(wb.sheetnames)
    # the FS chain A(5)->B(10)->C(8) is fully critical; project duration 23
    sched = wb["L2 Schedule"]
    txt = "\n".join(str(c.value) for row in sched.iter_rows() for c in row if c.value is not None)
    assert "23" in txt, "project duration (5+10+8=23) must appear"


def test_schedule_cost_loading_is_formula_linked():
    from app.lib.pm_excel import generate_cost_loaded_schedule
    wb = generate_cost_loaded_schedule({"project": "Test DC"}, ACTIVITIES)
    cl = _formulas(wb["Cost Loading"])
    # cumulative cost must be a running formula (=prev+current), not pasted
    assert any("+" in f for f in cl), f"cost loading needs cumulative formulas, got {cl}"
    # total cost is a =SUM
    assert any(f.startswith("=SUM(") for f in cl + _formulas(wb["L2 Schedule"])), "needs a =SUM total cost"
    # man-days = duration * manpower (live)
    md = _formulas(wb["Manpower Histogram"])
    assert any("*" in f for f in md), "man-days must be =Dur*Manpower"


def test_evm_metrics_are_live_formulas():
    from app.lib.pm_excel import generate_evm_workbook
    periods = [
        {"period": "M1", "pv": 100000, "ev": 90000, "ac": 95000},
        {"period": "M2", "pv": 300000, "ev": 280000, "ac": 300000},
    ]
    wb = generate_evm_workbook({"project": "Test DC", "bac": 800000}, periods)
    f = _formulas(wb["EVM"])
    # CPI=EV/AC, SPI=EV/PV, EAC=BAC/CPI — all live (division formulas present)
    assert any("/" in x for x in f), f"EVM needs CPI/SPI/EAC division formulas, got {f}"
    # CV=EV-AC, SV=EV-PV (subtraction)
    assert any("-" in x for x in f), "EVM needs CV/SV variance formulas"


def test_evm_formulas_compute_cpi_correctly():
    """Validation guard: CPI for M2 = EV/AC = 280000/300000 must evaluate."""
    import openpyxl, re
    from app.lib.pm_excel import generate_evm_workbook
    periods = [{"period": "M2", "pv": 300000, "ev": 280000, "ac": 300000}]
    wb = generate_evm_workbook({"project": "X", "bac": 800000}, periods)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    wb2 = openpyxl.load_workbook(buf, data_only=False)
    ws = wb2["EVM"]
    # find a cell whose formula divides two cells (CPI) and evaluate it
    found = None
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and re.fullmatch(r"=[A-Z]+\d+/[A-Z]+\d+", c.value):
                a, b = re.findall(r"[A-Z]+\d+", c.value)
                va, vb = ws[a].value, ws[b].value
                if isinstance(va, (int, float)) and isinstance(vb, (int, float)) and vb:
                    found = round(va / vb, 4)
    assert found == round(280000 / 300000, 4), f"CPI formula didn't compute right: {found}"
